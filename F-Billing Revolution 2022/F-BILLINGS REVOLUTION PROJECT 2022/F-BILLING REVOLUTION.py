from cgitb import text
from distutils import command
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from itertools import count
from pickle import FRAME
from pydoc import describe
from select import select
import smtplib
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkinter.tix import Select
from PIL import ImageTk, Image, ImageFile
from turtle import width
from xml.dom.minidom import Element
from PIL import ImageTk, Image
import PyPDF2
from numpy import empty
import pandas as pd
from tkinter.messagebox import showinfo
import tkinter.scrolledtext as scrolledtext
from tkinter.filedialog import askopenfilename
import datetime as dt
import tkinter as tk
import os, sys
# from pymysql import NULL
#from win32 import win32print
import pdfkit
import base64
import os
from datetime import datetime
from PyPDF2 import PdfFileWriter, PdfFileReader
from tkinter.filedialog import asksaveasfile
import webbrowser
from tkcalendar import Calendar
from tkcalendar import DateEntry
import tkcalendar
import pandas
from datetime import timedelta
from datetime import date
from tkinter import filedialog
import subprocess
import mysql.connector
import io
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import shutil
import csv
# from models import *


fbilldb = mysql.connector.connect(
    host="localhost", user="root", password="", database="fbillingsintgrtd", port="3306"
)
fbcursor = fbilldb.cursor()

ImageFile.LOAD_TRUNCATED_IMAGES = True

def reset():
  global root
  root.destroy()

root=Tk()
root.geometry("1360x730")
root.resizable(False, False)
root.title("F-Billing Revolution 2022(FREE version) | Company database:fbillingdb | User:Administrator")
p1 = PhotoImage(file = 'images/fbicon.png')
root.iconphoto(False, p1)


s = ttk.Style()
s.theme_use('default')
s.configure('TNotebook.Tab', background="#999999", width=20, padding=10)
invoices= PhotoImage(file="images/invoice.png")
orders = PhotoImage(file="images/order.png")
estimates = PhotoImage(file="images/estimate.png")
recurring = PhotoImage(file="images/recurring.png")
purchase = PhotoImage(file="images/purchase.png")
expenses = PhotoImage(file="images/expense.png")
customer = PhotoImage(file="images/customer.png")
product = PhotoImage(file="images/package.png")
reports = PhotoImage(file="images/report.png")
setting = PhotoImage(file="images/setting.png")
tick = PhotoImage(file="images/check.png")
warnin = PhotoImage(file="images/sign_warning.png")
cancel = PhotoImage(file="images/close.png")
saves = PhotoImage(file="images/save.png")
folder = PhotoImage(file="images/folder-black.png")
photo11 = PhotoImage(file = "images/invoice-pvt.png")
customer = PhotoImage(file="images/customer.png")
smslog = PhotoImage(file = "images/smslog.png")
video = PhotoImage(file = "images/video.png")
mark1 = PhotoImage(file="images/mark.png")
mark2 = PhotoImage(file="images/mark2.png")
photo10 = PhotoImage(file = "images/text-message.png")
addnew = PhotoImage(file="images/plus.png")
delete = PhotoImage(file="images/delete_E.png")
tabControl = ttk.Notebook(root)
tab1 = ttk.Frame(tabControl)
tab2 = ttk.Frame(tabControl)
tab3=  ttk.Frame(tabControl)
tab4 = ttk.Frame(tabControl)
tab5 = ttk.Frame(tabControl)
tab6=  ttk.Frame(tabControl)
tab7 = ttk.Frame(tabControl)
tab8 = ttk.Frame(tabControl)
tab9 =  ttk.Frame(tabControl)
tab10=  ttk.Frame(tabControl)
tabControl.add(tab1,image=invoices,compound = LEFT, text ='Invoices',)
tabControl.add(tab2,image=orders,compound = LEFT, text ='Orders')
tabControl.add(tab3,image=estimates,compound = LEFT, text ='Estimates')
tabControl.add(tab4,image=recurring,compound = LEFT, text ='Recurring')
tabControl.add(tab5,image=purchase,compound = LEFT, text ='Purchase Orders') 
tabControl.add(tab6,image=expenses,compound = LEFT, text ='Expenses')
tabControl.add(tab7,image=customer,compound = LEFT, text ='Customers')
tabControl.add(tab8,image=product,compound = LEFT, text ='Product/Services')
tabControl.add(tab9,image=reports,compound = LEFT, text ='Report')
tabControl.add(tab10,image=setting,compound = LEFT, text ='Settings')
tabControl.pack(expand = 1, fill ="both")


selectall = PhotoImage(file="images/table_select_all.png")
cut = PhotoImage(file="images/cut.png")
copy = PhotoImage(file="images/copy.png")
paste = PhotoImage(file="images/paste.png")

undo = PhotoImage(file="images/undo.png")
redo = PhotoImage(file="images/redo.png")
bold = PhotoImage(file="images/bold.png")

italics = PhotoImage(file="images/italics.png")
underline = PhotoImage(file="images/underline.png")
left = PhotoImage(file="images/left.png")

right = PhotoImage(file="images/right.png")
center = PhotoImage(file="images/center.png")
hyperlink = PhotoImage(file="images/hyperlink.png")
remove = PhotoImage(file="images/eraser.png")


photo = PhotoImage(file = "images/plus.png")
photo1 = PhotoImage(file = "images/edit.png")
photo2 = PhotoImage(file = "images/delete_E.png")
photo3 = PhotoImage(file = "images/export-file.png")
photo4 = PhotoImage(file = "images/seo.png")
photo5 = PhotoImage(file = "images/printer.png")
photo6 = PhotoImage(file = "images/gmail.png")
photo7 = PhotoImage(file = "images/priewok.png")
photo8 = PhotoImage(file = "images/refresh_E.png")
photo9 = PhotoImage(file = "images/sum.png")
photo10 = PhotoImage(file = "images/text-message.png")



############################################# ORDER MODULE #############################################
########## CREATE ORDER ###########
def create():
  pop=Toplevel(midFrame)
  pop.title("Orders")
  pop.geometry("950x690+150+0")


  #select customer
 

  def custom():
    global cuselection

    cuselection=Toplevel()
    cuselection.title("Select Customer")
    cuselection.geometry("930x650+240+10")
    cuselection.resizable(False, False)
  


    #add new customer

    
    def create1():
      global checkvar1,checkvar2,custid,bname,baddress,cat,sname,saddress,contp,cemail,ctel,cfax,cmob,scontp,scemail,sctel,scfax,ccountry,ccity,cnotes
      ven=Toplevel(midFrame)
      ven.title("Add new vendor")
      ven.geometry("930x650+240+10")
      def reg_1ord():# Storing values into db (user)contp,cemail,ctel,cfax,cmob,scontp,scemail,sctel,scfax,ccountry,ccity
        customerid = custid.get()
        businessname = bname.get()
        businessaddress = baddress.get()
        category= cat.get()
        shipname= sname.get()
        shipaddress = saddress.get()
        contactperson= contp.get() 
        cpemail = cemail.get()
        cptelno = ctel.get() 
        cpfax = cfax.get()
        cpmobileforsms = cmob.get()
        shipcontactperson= scontp.get()
        discount=e2dsc.get()
        specifictax1=e1tax.get()
        shipcpemail= scemail.get()
        shipcptelno= sctel.get()
        shipcpfax = scfax.get()
        country= ccountry.get()
        city = ccity.get()
        customertype = radio.get()
        notes=cnotes.get()
        status= checkvar1.get()
        taxexempt= checkvar2.get()   
        sql='INSERT INTO customer (customerid,businessname,businessaddress,category,status,shipname,shipaddress,contactperson,cpemail,cptelno,cpfax,cpmobileforsms,shipcontactperson,shipcpemail,shipcptelno,shipcpfax,taxexempt,country,city,notes,discount,specifictax1,customertype) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)' #adding values into db
        val=(customerid,businessname,businessaddress,category,status,shipname,shipaddress,contactperson,cpemail,cptelno,cpfax,cpmobileforsms,shipcontactperson,shipcpemail,shipcptelno,shipcpfax,taxexempt,country,city,notes,discount,specifictax1,customertype)
        fbcursor.execute(sql,val)
        fbilldb.commit()
        for record in slctcstmrtree.get_children():
          slctcstmrtree.delete(record)
        count=0
        fbcursor.execute('SELECT * FROM Customer;')
        for i in fbcursor:
            if True:
             slctcstmrtree.insert(parent='', index='end', iid=i, text='', values=(i[0],i[4],i[10],i[8]))  
            else:
                pass
        count += 1
        messagebox.showinfo('Registration successfull','Registration successfull')
      checkvar1=IntVar()
      checkvar2=IntVar()
      radio=IntVar()
      createFrame=Frame(ven, bg="#f5f3f2", height=650)
      createFrame.pack(side="top", fill="both")
      labelframe1 = LabelFrame(createFrame,text="Customer",bg="#f5f3f2",font=("arial",15))
      labelframe1.place(x=10,y=5,width=910,height=600)
      fbcursor.execute("SELECT * FROM Customer ORDER BY customerid DESC LIMIT 1")
      qury = fbcursor.fetchone()

      text1=Label(labelframe1, text="Customer ID:",bg="#f5f3f2",fg="blue").place(x=5 ,y=10)
      custid=Entry(labelframe1,width=25)
      custid.place(x=150,y=10)
      if not qury== None:
        id00=qury[0]+1
      else:
        id00=1
      custid.insert(0, id00)
      text2=Label(labelframe1, text="Category:",bg="#f5f3f2").place(x=390 ,y=10)
      cat=ttk.Combobox(labelframe1,width=25,value="Default")
      cat.place(x=460 ,y=10)
      text3=Label(labelframe1, text="Status:",bg="#f5f3f2").place(x=710 ,y=10)
      Checkbutton(labelframe1,text="Active",variable=checkvar1,onvalue=1,offvalue=0,bg="#f5f3f2").place(x=760 ,y=10)
      
      labelframe2 = LabelFrame(labelframe1,text="Invoice to (appears on invoices)",bg="#f5f3f2")
      labelframe2.place(x=5,y=40,width=420,height=150)
      name = Label(labelframe2, text="Ship to name:",bg="#f5f3f2",fg="blue").place(x=5,y=5)
      sname = Entry(labelframe2,width=28)
      sname.place(x=130,y=5)
      addr = Label(labelframe2, text="Address:",bg="#f5f3f2",fg="blue").place(x=5,y=40)
      saddress = Entry(labelframe2,width=28)
      saddress.place(x=130,y=40,height=80)

      def btn5cpy():
       tosp=sname.get()
       saddres=saddress.get()
       bname.delete(0, 'end')
       bname.insert(0, tosp)
       baddress.delete(0,'end')
       baddress.insert(0, saddres)
      
      btn1=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>",command=btn5cpy).place(x=440, y=90)

      labelframe3 = LabelFrame(labelframe1,text="Ship to (appears on invoices)",bg="#f5f3f2")
      labelframe3.place(x=480,y=40,width=420,height=150)
      name = Label(labelframe3, text="Business name:",bg="#f5f3f2").place(x=5,y=5)
      bname = Entry(labelframe3,width=28)
      bname.place(x=130,y=5)
      addr = Label(labelframe3, text="Address:",bg="#f5f3f2").place(x=5,y=40)
      baddress = Entry(labelframe3,width=28)
      baddress.place(x=130,y=40,height=80)
      
      labelframe4 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
      labelframe4.place(x=5,y=195,width=420,height=150)
      name = Label(labelframe4, text="Contact person:",bg="#f5f3f2").place(x=5,y=5)
      contp = Entry(labelframe4,width=28)
      contp.place(x=130,y=5)
      email = Label(labelframe4, text="E-mail address:",bg="#f5f3f2",fg="blue").place(x=5,y=35)
      cemail = Entry(labelframe4,width=28)
      cemail.place(x=130,y=35)
      tel = Label(labelframe4, text="Tel.number:",bg="#f5f3f2").place(x=5,y=65)
      ctel = Entry(labelframe4,width=11)
      ctel.place(x=130,y=65)
      fax = Label(labelframe4, text="Fax:",bg="#f5f3f2").place(x=240,y=65)
      cfax = Entry(labelframe4,width=11)
      cfax.place(x=280,y=65)
      sms = Label(labelframe4, text="Mobile number for SMS notifications:",bg="#f5f3f2").place(x=5,y=95)
      cmob = Entry(labelframe4,width=15)
      cmob.place(x=248,y=95)      
      def btncpy1():
       cprsn1=contp.get()
       cemail1=cemail.get()
       no=ctel.get()
       fx=cfax.get()
       scontp.insert(0, cprsn1)
       scemail.insert(0, cemail1)
       sctel.insert(0, no)
       scfax.insert(0,fx)
      btn1=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>",command=btncpy1).place(x=440, y=250)

      
      labelframe5 = LabelFrame(labelframe1,text="Ship to contact",bg="#f5f3f2")
      labelframe5.place(x=480,y=195,width=420,height=125)
      name = Label(labelframe5, text="Contact person:",bg="#f5f3f2").place(x=5,y=5)
      scontp = Entry(labelframe5,width=28)
      scontp.place(x=130,y=5)
      email = Label(labelframe5, text="E-mail address:",bg="#f5f3f2").place(x=5,y=35)
      scemail = Entry(labelframe5,width=28)
      scemail.place(x=130,y=35)
      tel = Label(labelframe5, text="Tel.number:",bg="#f5f3f2").place(x=5,y=65)
      sctel = Entry(labelframe5,width=11)
      sctel.place(x=130,y=65)
      fax = Label(labelframe5, text="Fax:",bg="#f5f3f2").place(x=240,y=65)
      scfax = Entry(labelframe5,width=11)
      scfax.place(x=280,y=65)

      labelframe6 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
      labelframe6.place(x=5,y=350,width=420,height=100)
      Checkbutton(labelframe6,text="Tax Exempt",variable=checkvar2,onvalue=1,offvalue=0,bg="#f5f3f2").place(x=5 ,y=5)
      tax = Label(labelframe6, text="Specific Tax1 %:",bg="#f5f3f2").place(x=180,y=5)
      e1tax = Entry(labelframe6,width=10)
      e1tax.place(x=290,y=5)
      discount = Label(labelframe6, text="Discount%:",bg="#f5f3f2").place(x=5,y=35)
      e2dsc = Entry(labelframe6,width=10)
      e2dsc.place(x=100,y=35)

      labelframe7 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
      labelframe7.place(x=480,y=330,width=420,height=100)
      country = Label(labelframe7, text="country:",bg="#f5f3f2").place(x=5,y=5)
      ccountry = Entry(labelframe7,width=28)
      ccountry.place(x=130,y=5)
      city = Label(labelframe7, text="City:",bg="#f5f3f2").place(x=5,y=35)
      ccity = Entry(labelframe7,width=28)
      ccity.place(x=130,y=35)

      labelframe8 = LabelFrame(labelframe1,text="Customer Type",bg="#f5f3f2")
      labelframe8.place(x=5,y=460,width=420,height=100)
      R1=Radiobutton(labelframe8,text=" Client ",variable=radio,value=1,bg="#f5f3f2").place(x=5,y=15)
      R2=Radiobutton(labelframe8,text=" Vendor ",variable=radio,value=2,bg="#f5f3f2").place(x=150,y=15)
      R3=Radiobutton(labelframe8,text=" Both(client/vendor)",variable=radio,value=3,bg="#f5f3f2").place(x=250,y=15)
      

      labelframe9 = LabelFrame(labelframe1,text="Notes",bg="#f5f3f2")
      labelframe9.place(x=480,y=430,width=420,height=150)
      cnotes = Entry(labelframe9)
      cnotes.place(x=10,y=10,height=100,width=390)

      btn1=Button(ven,width=60,height=10,bg="#f5f3f2",compound = LEFT,image=tick ,text="OK",command=reg_1ord).place(x=20, y=615)
      btn2=Button(ven,width=60,height=10,bg="#f5f3f2",compound = LEFT,image=cancel,text="Cancel",command=lambda:ven.destroy()).place(x=800, y=615)
      ven.mainloop()

#Edit Customer 
    def edit_customer_edtp():
    #  try:
      itemid = slctcstmrtree.item(slctcstmrtree.focus())["values"][0]
      sql = "select * from Customer where customerid = %s"
      val = (itemid, )
      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone()


      def update_customer_edto():# Storing values into db (user)
        
        itemid = slctcstmrtree.item(slctcstmrtree.focus())["values"][0]
        print(itemid)
        customerid = b1.get()
        category = ca.get()
        businessname = b5.get()
        address = b14.get()
        shiptoname = b15.get()
        shipaddress = a23.get()
        contactperson = cs.get()
        cpemail = ct.get()
        cptelno = a61.get() 
        cpfax = a71.get()
        cpmobileforsms = a51.get()
        shipcontactperson = b13.get()
        shipcpemail = a22.get()
        shipcptelno = a31.get()
        shipcpfax = a41.get()
        taxexempt = checkvar1.get()
        specifictax1 = a13.get()
        discount = a12.get()
        country = c.get()
        city = a14.get()
        notes = b12.get()
        sql='UPDATE Customer set category=%s,businessname=%s,businessaddress=%s,shipname=%s,shipaddress=%s,contactperson=%s,cpemail=%s,cptelno=%s,cpfax=%s,cpmobileforsms=%s,shipcontactperson=%s,shipcpemail=%s,shipcptelno=%s,shipcpfax=%s,taxexempt=%s,specifictax1=%s,discount=%s,country=%s,city=%s,notes=%s where customerid=%s'
        val=(category,businessname,address,shiptoname,shipaddress,contactperson,cpemail,cptelno,cpfax,cpmobileforsms,shipcontactperson,shipcpemail,shipcptelno,shipcpfax,taxexempt,specifictax1,discount,country,city,notes,itemid)
        fbcursor.execute(sql,val)
        fbilldb.commit()
        messagebox.showinfo('Update Successfull','Update Successfull')
        edit_customer.destroy()
    





      edit_customer = Toplevel()  
      edit_customer.title("Edit Customer Details ")
      p2 = PhotoImage(file = "images/fbicon.png")
      edit_customer.iconphoto(False, p2)
      edit_customer.geometry("775x580+300+100")
      Labelframe1=LabelFrame(edit_customer,text="Customer")
      Labelframe1.place(x=10,y=10,width=755,height=525)
      a1=Label(Labelframe1,text="Customer ID:",fg="Blue")
      a2=Label(Labelframe1,text="Category:")
      a3=Label(Labelframe1,text="Status :")
      a3.place(x=620,y=7)
      b1=Entry(Labelframe1)
      b1.delete(0,'end')
      b1.insert(0, psdata[0])
      ca=StringVar() 
      b2=ttk.Combobox(Labelframe1,textvariable = ca )
      b2.delete(0,'end')
      b2.insert(0, psdata[2])    
      b2['values'] = ('Default')  
      b2.place(x=390,y=220) 
      b2.current(0)
      a1.place(x=10,y=7)
      a2.place(x=330,y=7)   
      b1.place(x=120,y=7,width=200)
      b2.place(x=390,y=7,width=220)
      checkvar1 = BooleanVar()
      chkbtn2 = Checkbutton(Labelframe1, text = "Active", variable = checkvar1, onvalue = 0, offvalue = 1)
        # chkbtn1.delete(0,'end')
        # chkbtn1.insert(0, psdata[3])
      chkbtn2.place(x=670,y=6)
      act = psdata[3]
      print(act,)
      if act == '1':
        chkbtn2.select()
      else:
        chkbtn2.deselect()
        
      Labelframe2=LabelFrame(Labelframe1,text="Invoice to (appears on invoice)")
      Labelframe2.place(x=10,y=35,width=340,height=125)
      a1=Label(Labelframe2,text="Business Name:",fg="Blue").place(x=10,y=10)
      a2=Label(Labelframe2,text="Address:",fg="Blue").place(x=10,y=35)
      b5=Entry(Labelframe2)
      b5.place(x=110,y=10,width=210)
      b5.delete(0,'end')
      b5.insert(0, psdata[4])
      b14=Entry(Labelframe2)
      b14.place(x=110,y=35,width=210,height=63)
      b14.delete(0,'end')
      b14.insert(0, psdata[5]) 


      def btnedtord1():
          b15.delete(0,END)
          b15.insert(0,psdata[4])
          a23.delete(0,END)
          a23.insert(0, psdata[5]) 
      btn110=Button(Labelframe1,width=3,height=2,compound = LEFT,text=">>", command=btnedtord1).place(x=359,y=85,height=20)
      Labelframe3=LabelFrame(Labelframe1,text="Ship to (appears on invoice)")
      Labelframe3.place(x=400,y=35,width=340,height=125)
      a11=Label(Labelframe3,text="Ship to Name:").place(x=10,y=10)
      a21=Label(Labelframe3,text="Address:").place(x=10,y=35)
      b15=Entry(Labelframe3)
      b15.place(x=110,y=10,width=210)
      b15.delete(0,'end')
      b15.insert(0, psdata[6])
      a23=Entry(Labelframe3)
      a23.place(x=110,y=35,width=210,height=63)
      a23.delete(0,'end')
      a23.insert(0, psdata[7])
      Labelframe4=LabelFrame(Labelframe1,text="Contact")
      Labelframe4.place(x=10,y=170,width=340,height=137)
      a11=Label(Labelframe4,text="Contact Person:").place(x=10,y=10)
      a21=Label(Labelframe4,text="Email Address:",fg="Blue").place(x=10,y=35)
      a31=Label(Labelframe4,text="Tel. No:").place(x=10,y=60)
      a41=Label(Labelframe4,text="Fax:").place(x=200,y=60)
      a51=Label(Labelframe4,text="Mobile number for SMS notification:").place(x=10,y=85)
      cs=StringVar()
      a11=Entry(Labelframe4, textvariable=cs)
      a11.place(x=110,y=10,width=210)
      a11.delete(0,'end')
      a11.insert(0, psdata[8])
      ct=StringVar()
      a21=Entry(Labelframe4, textvariable=ct)
      a21.place(x=110,y=35,width=210)  
      a21.delete(0,'end')
      a21.insert(0, psdata[9])
      a61=Entry(Labelframe4)
      a61.place(x=110,y=60,width=90)
      a61.delete(0,'end')
      a61.insert(0, psdata[10])
      a71=Entry(Labelframe4)
      a71.place(x=230,y=60,width=90)
      a71.delete(0,'end')
      a71.insert(0, psdata[11])
      a51=Entry(Labelframe4)
      a51.place(x=215,y=85,width=105)
      a51.delete(0,'end')
      a51.insert(0, psdata[12])


      def btnedtord():
            b13.delete(0,END)
            b13.insert(0,psdata[8])
            a22.delete(0,END)
            a22.insert(0, psdata[9])
            a31.delete(0,END)
            a31.insert(0,psdata[10])
            a41.delete(0,END)
            a41.insert(0,psdata[11]) 
      btn111=Button(Labelframe1,width=3,height=2,compound = LEFT,text=">>", command=btnedtord).place(x=359,y=220,height=20)
      Labelframe5=LabelFrame(Labelframe1,text="Ship To Contact")
      Labelframe5.place(x=400,y=170,width=340,height=108)
      a11=Label(Labelframe5,text="Contact Person:").place(x=10,y=10)
      a21=Label(Labelframe5,text="Email Address:").place(x=10,y=35)
      a31=Label(Labelframe5,text="Tel. No:").place(x=10,y=60)
      a41=Label(Labelframe5,text="Fax:").place(x=200,y=60)
      b13=Entry(Labelframe5)
      b13.place(x=110,y=10,width=210)
      b13.delete(0,'end')
      b13.insert(0, psdata[13])
      a22=Entry(Labelframe5)
      a22.place(x=110,y=35,width=210)
      a22.delete(0,'end')
      a22.insert(0, psdata[14])
      a31=Entry(Labelframe5)
      a31.place(x=110,y=60,width=90)
      a31.delete(0,'end')
      a31.insert(0, psdata[15])
      a41=Entry(Labelframe5)
      a41.place(x=230,y=60,width=90)
      a41.delete(0,'end')
      a41.insert(0, psdata[16])
      Labelframe6=LabelFrame(Labelframe1,text="Payment Option")
      Labelframe6.place(x=10,y=317,width=340,height=80)
      checkvar1 = IntVar()
      chkbtn1 = Checkbutton(Labelframe6, text = "Tax Exempt", variable = checkvar1, onvalue = 1, offvalue = 0, font=("arial", 8))
      act = psdata[17]
      print(act,)
      if act == '1':
        chkbtn1.select()
      else:
        chkbtn1.deselect()
      chkbtn1.place(x=10,y=6)
        
      a11=Label(Labelframe6,text="Specific Tax1%:").place(x=150,y=7)
      a3label=Label(Labelframe6,text="Discount%:").place(x=10,y=30)
      b11val = IntVar(Labelframe6)
      a13=Entry(Labelframe6)
      a13.place(x=250,y=7,width=70)
      a13.delete(0,'end')
      # a13.insert(0, psdata[18])
      a12=Entry(Labelframe6)
      a12.place(x=80,y=30,width=70)
      a12.delete(0,'end')
      a12.insert(0, psdata[19])
      Labelframe7=LabelFrame(Labelframe1,text="Customer type")
      Labelframe7.place(x=10,y=405,width=340,height=90)
      i=IntVar()
      r1=Radiobutton(Labelframe7, text = "Client", variable = i, value = 1).place(x=5,y=15)
      r2=Radiobutton(Labelframe7, text = "Vender", variable = i, value = 2).place(x=90,y=15)
      r3=Radiobutton(Labelframe7, text = "Both(Client/Vender)", variable = i, value = 3).place(x=180,y=15)
      Labelframe8=LabelFrame(Labelframe1,text="Additional Info")
      Labelframe8.place(x=400,y=288,width=340,height=80)
      a2label=Label(Labelframe8,text="Country:").place(x=10,y=5)
      a1label=Label(Labelframe8,text="City:").place(x=10,y=30)
      c=StringVar() 
      b11=ttk.Combobox(Labelframe8,textvariable=c)
      b11.place(x=110,y=5,width=210)
      b11['values'] = ('India','America')    
      b11.place(x=110,y=5)
      b11.delete(0,'end')
      b11.insert(0, psdata[20]) 
      a14=Entry(Labelframe8)
      a14.place(x=110,y=30,width=210)
      a14.delete(0,'end')
      a14.insert(0, psdata[21])
      Labelframe9=LabelFrame(Labelframe1,text="Notes")
      Labelframe9.place(x=400,y=380,width=340,height=115)
      '''scrollbar = Scrollbar(Labelframe9)
          scrollbar.place(x=300,y=10)
          b12=Entry(Labelframe9,yscrollcommand=scrollbar.set).place(x=10,y=10,width=290,height=70)
          yscrollcommand.config(command=b12.yview)'''
      b12=Entry(Labelframe9)
      b12.place(x=20,y=10,width=295,height=70)
      b12.delete(0,'end')
      b12.insert(0, psdata[23])
      scrollbar = Scrollbar(Labelframe9)
      scrollbar.place(x=295,y=10)
      btn1=Button(edit_customer,width=50,compound = LEFT,image=tick ,text="  OK",command=update_customer_edto).place(x=20, y=545)
      btn2=Button(edit_customer,width=80,compound = LEFT,image=cancel,text="  Cancel",command=lambda:edit_customer.destroy()).place(x=665, y=545)

    #  except:
    #   try:
    #     edit_customer.destroy()
    #   except:
    #     pass
    #     messagebox.showerror('F-Billing Revolution', 'Select a product to edit.')

    #     edit_customer.mainloop()





    global slctcstmrtree
    slctcstmrtree=ttk.Treeview(cuselection, height=27)
    slctcstmrtree["columns"]=["1","2","3","4"]
    slctcstmrtree.column("#0", width=35, anchor="center")
    slctcstmrtree.column("1", width=160, anchor="center")
    slctcstmrtree.column("2", width=160, anchor="center")
    slctcstmrtree.column("3", width=140, anchor="center")
    slctcstmrtree.column("4", width=140, anchor="center")
    slctcstmrtree.heading("#0",text="")
    slctcstmrtree.heading("1",text="Customer/Ventor ID")
    slctcstmrtree.heading("2",text="Customer/Ventor Name")
    slctcstmrtree.heading("3",text="Tel.")
    slctcstmrtree.heading("4",text="Contact Person")
    slctcstmrtree.place(x=5, y=45)


    def filt():
      # global filttxt 
  
      filtr = filttxt.get()
      sqlcusven = "SELECT * FROM Customer WHERE businessname=%s"
      valcusven = (filtr, )
      fbcursor.execute(sqlcusven, valcusven)
      records = fbcursor.fetchall()
      # print(records)
      if records==[]:
        pass
      else:
       for record in slctcstmrtree.get_children():
        slctcstmrtree.delete(record)   
       count=0
       for i in records:
         if True:
          slctcstmrtree.insert(parent='', index='end', iid=i, text='', values=(i[0],i[4],i[10],i[8]))  
         else:
           pass
       count += 1

    # global filttxt
    enter=Label(cuselection, text="Enter filter text").place(x=5, y=10)
    filttxt=Entry(cuselection, width=20)
    filttxt.place(x=110, y=10)
    fltrbutton=Button(cuselection,text = 'Click Here',command=filt)
    fltrbutton.place(x=236,y=7,height=25,width=70)

    # def treefthcng():
    #  itemid = slctcstmrtree.item(slctcstmrtree.focus())["values"][1]
    #  sql = "select * from Customer where customerid = %s"
    #  val = (itemid, )
    #  fbcursor.execute(sql, val)
    #  global slctcstr
    #  slctcstr = fbcursor.fetchone()
    #  e1.delete(0,'end')
    #  e1.insert(0, slctcstr[4])

    def cancelcuselection():
      cuselection.destroy()

 
    
    fbcursor.execute('SELECT * FROM Customer;') 
    j = 0
    for i in fbcursor:
      slctcstmrtree.insert(parent='', index='end', iid=i, text='', values=(i[0],i[4],i[10],i[8]))
      j += 1


    ctegorytree=ttk.Treeview(cuselection, height=27)
    ctegorytree["columns"]=["1"]
    ctegorytree.column("#0", width=35, minwidth=20)
    ctegorytree.column("1", width=205, minwidth=25, anchor=CENTER)    
    ctegorytree.heading("#0",text="", anchor=W)
    ctegorytree.heading("1",text="View filter by category", anchor=CENTER)
    ctegorytree.place(x=660, y=45)

    listbox = Listbox(cuselection,height = 33,  
              width = 40,  
              bg = "white",
              activestyle = 'dotbox',  
              fg = "black",
              highlightbackground="white")
    
    listbox.insert(0, "                               View all records")
    # listbox.insert(1, "                               View all products")
    # listbox.insert(2, "                               View all services")
    listbox.place(x=660,y=65,)
    listbox.bind('<<ListboxSelect>>')


    # scrollbar = Scrollbar(tab2)
    # scrollbar.place(x=640, y=45, height=560)
    # scrollbar.config( command=tree.yview )

    scrollbar = Scrollbar(cuselection)
    scrollbar.place(x=1016+300+25, y=120, height=280+20)

    btn1=Button(cuselection,compound = LEFT,image=tick ,text="Ok", width=60,command=treefthcng).place(x=15, y=610)
    btn1=Button(cuselection,compound = LEFT,image=tick,text="Edit selected customer", width=150,command=edit_customer_edtp).place(x=250, y=610)
    btn1=Button(cuselection,compound = LEFT,image=tick, text="Add new customer", width=150,command=create1).place(x=435, y=610)
    btn1=Button(cuselection,compound = LEFT,image=cancel ,text="Cancel",command= cancelcuselection,width=60).place(x=740, y=610)   



    

  #add new line item
  def newline():

    try:
      if not cmb.get():
        messagebox.showerror('F-Billing Revolution', 'Add a customer before adding new line.') 
      else:

        newselection=Toplevel()
        newselection.title("Select Customer")
        newselection.geometry("930x650+240+10")
        newselection.resizable(False, False)

        #add new product

        def edit_productord():  
         try:
          itemid = addcusventtree.item(addcusventtree.focus())["values"][0]
          
          global filename
          filename = ""
          
          def upload_file():
            global filename,img, b2
            f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
            filename = filedialog.askopenfilename(filetypes=f_types)
            shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            image = Image.open(filename)
            resize_image = image.resize((350, 350))
            img = ImageTk.PhotoImage(resize_image)
            b2 = Button(imageFrame,image=img)
            b2.place(x=130, y=80)
          
          def updateproducts():
            global img , filename 
            sku = codeentry.get()
            status = checkvarStatus.get()
            catgory = n.get()
            name = nameentry.get()
            description = desentry.get()
            unitprice = uval.get()
            peices = pcsentry.get()
            cost = costval.get()
            price_cost = priceval.get()
            taxable = checkvarStatus2.get()
            nostockcontrol = checkvarStatus3.get()
            stock = stockval.get()
            lowstock = lowval.get()
            warehouse = wareentry.get()
            pnotes = sctxt.get("1.0", 'end-1c')
            entries = [sku, name, unitprice, cost]
            entri = []
            for i in entries:
              if i == '':
                entri.append(i)
            if len(entri) == 0:
              if filename == "":
                sql = "update Productservice set sku=%s, category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, privatenote=%s where Productserviceid = %s"
                val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes, itemid)
                fbcursor.execute(sql, val)
                fbilldb.commit()
              else:
                file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
                sql = "update Productservice set category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, image=%s, privatenote=%s where Productserviceid = %s"
                val = (catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse,filename.split('/')[-1], pnotes, itemid)
                fbcursor.execute(sql, val)
                fbilldb.commit()
                messagebox.showinfo("F-Billing Revolution", "Product updated successfully.")
                for record in addcusventtree.get_children():
                  addcusventtree.delete(record)
                fbcursor.execute("select *  from Productservice")
                pandsdata = fbcursor.fetchall()
                countp = 0
                for i in pandsdata:
                  if i[6] == '1':
                    acti = 'Active'
                  else:
                    acti = 'Inactive' 
                  if i[13] > i[14]:
                    addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))
                    countp += 1
                  elif (i[12] =="0") == (i[13] <= i[14]):
                    addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))
                    countp += 1
                  elif i[12] == '1':
                    addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))
                    countp += 1
                  else:
                    pass
              top.destroy()
            else:
              messagebox.showinfo("F-Billing Revolution", "Fields name or SKU entered is already in database.")
              top.destroy()
            for record in addcusventtree.get_children():
                  addcusventtree.delete(record)
            fbcursor.execute("select *  from Productservice")
            pandsdata = fbcursor.fetchall()
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))
                countp += 1
              elif i[12] == '1':
                addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))
                countp += 1
              else:
                pass
            
          sql = "select * from Productservice where Productserviceid = %s"
          val = (itemid, )
          fbcursor.execute(sql, val)
          psdata = fbcursor.fetchone()
          
          
          top = Toplevel()  
          top.title("Edit Product/Service details")
          p3 = PhotoImage(file = 'images/fbicon.png')
          top.iconphoto(False, p1)
          top.geometry("600x550+390+125")
          tabControl = ttk.Notebook(top)
          s = ttk.Style()
          s.theme_use('default')
          s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)

          taba = ttk.Frame(tabControl)
          tabb = ttk.Frame(tabControl)
          
          tabControl.add(taba,compound = LEFT, text ='Product/Service')
          tabControl.add(tabb,compound = LEFT, text ='Product Image')
          
          tabControl.pack(expand = 1, fill ="both")
          
          innerFrame = Frame(taba,bg="#f5f3f2", relief=GROOVE)
          innerFrame.pack(side="top",fill=BOTH)

          updateframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
          updateframe.pack(side="top",fill=BOTH,padx=10)

          code1=Label(updateframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
          code1.place(x=20,y=0)
          codeentry = Entry(updateframe,width=35)
          codeentry.place(x=110,y=8)
          codeentry.insert(0, psdata[2])

          checkvarStatus=IntVar()
          status1=Label(updateframe,text="Status:")
          status1.place(x=380,y=8)
          Button1 = Checkbutton(updateframe, 
                            variable = checkvarStatus,text="Active",compound="right",
                            onvalue =1,
                            offvalue =0,
                            width = 10)
          Button1.place(x=420,y=5)
          sta = psdata[6]
          if sta == '1':
            Button1.select()
          else:
            Button1.deselect()



          category1=Label(updateframe,text="Category:",pady=5,padx=10)
          category1.place(x=20,y=40)
          n = StringVar() 
          category = Entry(updateframe,width=70,textvariable=n) 
          category.place(x=110,y=45)
          category.insert(0, psdata[3])


          name1=Label(updateframe,text="Name :",fg="blue",pady=5,padx=10)
          name1.place(x=20,y=70)
          nameentry = Entry(updateframe,width=70)
          nameentry.place(x=110,y=75)
          nameentry.insert(0, psdata[4])

          des1=Label(updateframe,text="Description :",pady=5,padx=10)
          des1.place(x=20,y=100)
          desentry = Entry(updateframe,width=70)
          desentry.place(x=110,y=105)
          desentry.insert(0, psdata[5])

          def set_label(name, index, mode):
            priceval.set(uval.get() - costval.get())
          
          unit1=Label(updateframe,text="Unit Price:",fg="blue",pady=5,padx=10)
          unit1.place(x=20,y=130)
          
          uval = IntVar()
          unitentry = Entry(updateframe,width=20,textvariable=uval)
          unitentry.place(x=110,y=135)
          unitentry.delete(0,'end')
          unitentry.insert(0, psdata[7])
          

          pcsval = IntVar()
          pcs1=Label(updateframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
          pcs1.place(x=320,y=130)
          pcsentry = Entry(updateframe,width=20,textvariable=pcsval)
          pcsentry.place(x=410,y=135)
          pcsentry.delete(0,'end')
          pcsentry.insert(0, psdata[8])
          

          costval = IntVar()
          cost1=Label(updateframe,text="Cost:",pady=5,padx=10)
          cost1.place(x=20,y=160)
          costentry = Entry(updateframe,width=20,textvariable=costval)
          costentry.place(x=110,y=165)
          costentry.delete(0,'end')
          costentry.insert(0, psdata[9])
          

          priceval = IntVar()
          price1=Label(updateframe,text="(Price-Cost):",pady=5,padx=10)
          price1.place(x=20,y=190)
          priceentry = Entry(updateframe,width=20,textvariable=priceval)
          priceentry.place(x=110,y=195)
          priceentry.delete(0,'end')
          priceentry.insert(0, psdata[11])

          uval.trace('w', set_label)
          costval.trace('w', set_label)
          

          checkvarStatus2=IntVar()
        
          Button2 = Checkbutton(updateframe,variable = checkvarStatus2, 
                            text="Taxable Tax1rate",compound="right",
                            onvalue =1 ,
                            offvalue =0,
                            height=2,
                            width = 12)

          Button2.place(x=415,y=153)
          tax = psdata[10]
          if tax == '1':
            Button2.select()
          else:
            Button2.deselect()
          

          

          checkvarStatus3=IntVar()
          def switch():
            if checkvarStatus3.get():
              stockentry["state"] = DISABLED
              lowentry["state"] = DISABLED
              wareentry["state"] = DISABLED
            else:
              stockentry["state"] = NORMAL
              lowentry["state"] = NORMAL
              wareentry["state"] = NORMAL
        
          Button3 = Checkbutton(updateframe,variable = checkvarStatus3,command=switch, 
                            text="No stock Control", 
                            onvalue =1 ,
                            offvalue = 0,
                            height=3,
                            width = 15)

          Button3.place(x=40,y=220)

          


          stockval = IntVar(updateframe)
          stock1=Label(updateframe,text="Stock:",pady=5,padx=10)
          stock1.place(x=90,y=260)
          stockentry = Entry(updateframe,width=15,textvariable=stockval)
          stockentry.place(x=140,y=265)
          stockentry.delete(0,'end')
          stockentry.insert(0, psdata[13])
          

          lowval = IntVar(updateframe)
          low1=Label(updateframe,text="Low Stock Warning Limit:",pady=5,padx=10)
          low1.place(x=280,y=260)
          lowentry = Entry(updateframe,width=15,textvariable=lowval)
          lowentry.place(x=435,y=265)
          lowentry.delete(0,'end')
          lowentry.insert(0, psdata[14])
          

        
          ware1=Label(updateframe,text="Warehouse:",pady=5,padx=10)
          ware1.place(x=60,y=290)
          wareentry = Entry(updateframe,width=64)
          wareentry.place(x=140,y=295)
          wareentry.insert(0, psdata[15])

          scr = psdata[12]
          if scr == '1':
            Button3.select()
            stockentry["state"] = DISABLED
            lowentry["state"] = DISABLED
            wareentry["state"] = DISABLED
          else:
            Button3.deselect()
            stockentry["state"] = NORMAL
            lowentry["state"] = NORMAL
            wareentry["state"] = NORMAL
          
          

          

          text1=Label(updateframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
          text1.place(x=20,y=320)
          sctxt = scrolledtext.ScrolledText(updateframe, undo=True,width=62,height=4)
          sctxt.place(x=32,y=358)
          try:
            sctxt.insert("1.0", psdata[16])
          except:
            pass

          okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=updateproducts)
          okButton.pack(side=LEFT, padx=(10, 0))

          cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=lambda : top.destroy())
          cancelButton.pack(side=RIGHT, padx=(0, 10))
          
          
          imageFrame = Frame(tabb, relief=GROOVE,height=580)
          imageFrame.pack(side="top",fill=BOTH)

          browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
          browseimg.place(x=15,y=35)

          browsebutton=Button(imageFrame,text = 'Browse', command=upload_file)
          browsebutton.place(x=470,y=30,height=30,width=50)

          try:
            image = Image.open("images/"+psdata[17])
            resize_image = image.resize((350, 350))
            image = ImageTk.PhotoImage(resize_image)
            b2 = Label(imageFrame,image=image,width=350,height=350)
            b2.photo = image
            b2.place(x=130, y=80)
            print(image)
          except:
            pass

          removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT,command=lambda: b2.destroy())
          removeButton.place(x=410,y=460)
         except:
        #   try:
        #     top.destroy()
        #   except:
            pass
            messagebox.showerror('F-Billing Revolution', 'Select a record to edit.')

        def product():  
          global codeentry,nameentry,country,desentry,unitentry,pcsentry,costentry,priceentry,stockentry,lowentry,wareentry,txt,checkvarStatus,checkvarStatus2,checkvarStatus3
          top = Toplevel()  
          top.title("Add a new Product/Service")
          p2 = PhotoImage(file = 'images/fbicon.png')
          top.iconphoto(False, p2)

          top.geometry("700x550+390+15")
          tabControl = ttk.Notebook(top)
          s = ttk.Style()
          s.theme_use('default')
          s.configure('TNotebook.Tab', background="#999999",padding=10,bd=0)


          tabsb1 = ttk.Frame(tabControl)
          tab2 = ttk.Frame(tabControl)

          tabControl.add(tabsb1,compound = LEFT, text ='Product/Service')
          tabControl.add(tab2,compound = LEFT, text ='Product Image')

          tabControl.pack(expand = 1, fill ="both")

          innerFrame = Frame(tabsb1,bg="#f5f3f2", relief=GROOVE)
          innerFrame.pack(side="top",fill=BOTH)

          Customerlabelframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
          Customerlabelframe.pack(side="top",fill=BOTH,padx=10)
          global filename
          filename = ""
          
          def upload_fileord():
            global filename,img, b2
            f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
            filename = filedialog.askopenfilename(filetypes=f_types)
            shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            image = Image.open(filename)
            resize_image = image.resize((350, 350))
            img = ImageTk.PhotoImage(resize_image)
            b2 = Button(imageFrame,image=img)
            b2.place(x=130, y=80)
          
          def addproductsord():
            global img , filename 
            sku = codeentry.get()
            status = checkvarStatus.get()
            catgory = n.get()
            name = nameentry.get()
            description = desentry.get()
            unitprice = uval.get()
            peices = pcsentry.get()
            cost = costval.get()
            price_cost = priceval.get()
            taxable = checkvarStatus2.get()
            nostockcontrol = checkvarStatus3.get()
            stock = stockval.get()
            lowstock = lowval.get()
            warehouse = wareentry.get()
            pnotes = txt.get("1.0",'end-1c')
            entries = [sku, name, unitprice, cost]
            entri = []
            for i in entries:
              if i == '':
                entri.append(i)
            if len(entri) == 0:
              sql = 'select * from Productservice where sku = %s or name = %s'
              val  = (sku, name)
              fbcursor.execute(sql, val)
              fbcursor.fetchall()
              row_count = fbcursor.rowcount
              if row_count == 0:
                if filename == "":
                  sql = 'insert into Productservice(sku,productserviceid, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)'
                  val = (sku,sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes)
                  fbcursor.execute(sql, val)
                  fbilldb.commit()
                else:
                  file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
                  sql = 'insert into Productservice(sku,productserviceid,category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, image, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)'
                  val = (sku,sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, filename.split('/')[-1], pnotes)
                  fbcursor.execute(sql, val)
                  fbilldb.commit()
              else:
                messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
              top.destroy()
              messagebox.showinfo("F-Billing Revolution", "Product added successfully.")
              for record in addcusventtree.get_children():
                addcusventtree.delete(record)
              fbcursor.execute("select *  from Productservice")
              pandsdata = fbcursor.fetchall()
              countp = 0
              for i in pandsdata:
                if i[6] == '1':
                    acti = 'Active'
                else:
                    acti = 'Inactive' 
                if i[13] > i[14]:
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))
                  countp += 1
                elif (i[12] =="0") == (i[13] <= i[14]):
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))
                  countp += 1
                elif i[12] == '1':
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))
                  countp += 1
                else:
                  pass                
              
            else:
              messagebox.showinfo("Alert", "Fields name and SKU should not be empty.\nFill out required fields and try again")
              top.destroy()
      
          fbcursor.execute("SELECT * FROM Productservice ORDER BY sku DESC LIMIT 1")
          skuin = fbcursor.fetchone()
    
    
          code1=Label(Customerlabelframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
          code1.place(x=20,y=0)
          codeentry = Entry(Customerlabelframe,width=35)
          codeentry.place(x=120,y=8)

          checkvarStatus=IntVar()
          status1=Label(Customerlabelframe,text="Status:")
          status1.place(x=500,y=8)
          Button1 = Checkbutton(Customerlabelframe,
                            variable = checkvarStatus,text="Active",compound="right",
                            onvalue =0 ,
                            offvalue = 1,

                            width = 10)

          Button1.place(x=550,y=5)

          category1=Label(Customerlabelframe,text="Category:",pady=5,padx=10)
          category1.place(x=20,y=40)
          n = StringVar()
          catgory = ttk.Combobox(Customerlabelframe, width = 40, textvariable = n ) 
          catgory.place(x=110,y=45)
          catgory.insert(0, 'Default')


          name1=Label(Customerlabelframe,text="Name :",fg="blue",pady=5,padx=10)
          name1.place(x=20,y=70)
          nameentry = Entry(Customerlabelframe,width=60)
          nameentry.place(x=120,y=75)

          des1=Label(Customerlabelframe,text="Description :",pady=5,padx=10)
          des1.place(x=20,y=100)
          desentry = Entry(Customerlabelframe,width=60)
          desentry.place(x=120,y=105)

          uval = IntVar()
          unit1=Label(Customerlabelframe,text="Unit Price:",fg="blue",pady=5,padx=10)
          unit1.place(x=20,y=130)
          unitentry = Entry(Customerlabelframe,width=20,textvariable=uval)
          unitentry.place(x=120,y=135)

          # pcsval = IntVar(Customerlabelframe, value='0')
          pcs1=Label(Customerlabelframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
          pcs1.place(x=320,y=140)
          pcsentry = Entry(Customerlabelframe,width=20)
          pcsentry.place(x=410,y=140)

          costval = IntVar()
          cost1=Label(Customerlabelframe,text="Cost:",pady=5,padx=10)
          cost1.place(x=20,y=160)
          costentry = Entry(Customerlabelframe,width=20,textvariable=costval)
          costentry.place(x=120,y=165)
          def set_label(name, index, mode):
           priceval.set(uval.get() - costval.get())
          priceval = IntVar()
          price1=Label(Customerlabelframe,text="(Price-Cost):",pady=5,padx=10)
          price1.place(x=20,y=190)
          priceentry = Entry(Customerlabelframe,width=20,textvariable=priceval)
          priceentry.place(x=120,y=195)

          uval.trace('w', set_label)
          costval.trace('w', set_label)
          checkvarStatus2=IntVar()

          Button2 = Checkbutton(Customerlabelframe,variable = checkvarStatus2,
                            text="Taxable Tax1rate",compound="right",
                            onvalue =0 ,
                            offvalue = 1,
                            height=2,
                            width = 12)

          Button2.place(x=415,y=170)
          def switch():
           if checkvarStatus3.get():
            stockentry["state"] = DISABLED
            lowentry["state"] = DISABLED
            wareentry["state"] = DISABLED
           else:
            stockentry["state"] = NORMAL
            lowentry["state"] = NORMAL
            wareentry["state"] = NORMAL

          checkvarStatus3=BooleanVar()

          Button3 = Checkbutton(Customerlabelframe,variable = checkvarStatus3,command=switch,
                            text="No stock Control",
                            onvalue =1 ,
                            offvalue = 0,
                            height=3,
                            width = 15)

          Button3.place(x=40,y=220)


          stockval = IntVar()
          stock1=Label(Customerlabelframe,text="Stock:",pady=5,padx=10)
          stock1.place(x=90,y=260)
          stockentry = Entry(Customerlabelframe,width=15,textvariable=stockval)
          stockentry.place(x=150,y=265)

          lowval = IntVar(Customerlabelframe)
          low1=Label(Customerlabelframe,text="Low Stock Warning Limit:",pady=5,padx=10)
          low1.place(x=300,y=260)
          lowentry = Entry(Customerlabelframe,width=10,textvariable=lowval)
          lowentry.place(x=495,y=265)


          ware1=Label(Customerlabelframe,text="Warehouse:",pady=5,padx=10)
          ware1.place(x=60,y=290)
          wareentry = Entry(Customerlabelframe,width=50)
          wareentry.place(x=150,y=295)

          text1=Label(Customerlabelframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
          text1.place(x=20,y=330)

          txt = scrolledtext.ScrolledText(Customerlabelframe, undo=True,width=72,height=4)
          txt.place(x=32,y=358)




          okButton = Button(innerFrame,compound = LEFT,image=tick , text ="Ok",command=addproductsord,width=60)
          okButton.pack(side=LEFT)
          def closetab():
           top.destroy()

          cancelButton = Button(innerFrame,compound = LEFT,image=cancel ,text="Cancel",width=60, command=closetab)
          cancelButton.pack(side=RIGHT)

          imageFrame = Frame(tab2, relief=GROOVE,height=580)
          imageFrame.pack(side="top",fill=BOTH)

          browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
          browseimg.place(x=15,y=35)

          browsebutton=Button(imageFrame,text = 'Browse',command=upload_fileord)
          browsebutton.place(x=580,y=30,height=30,width=50)
          # try:
          #   image = Image.open("images/"+psdata[17])
          #   resize_image = image.resize((350, 350))
          #   image = ImageTk.PhotoImage(resize_image)
          #   b2 = Label(imageFrame,image=image,width=350,height=350)
          #   b2.photo = image
          #   b2.place(x=130, y=80)
          #   print(image)
          # except:
          #   pass

          removeButton = Button(imageFrame,compound = LEFT,image=cancel, text ="Remove Product Image",command=lambda:b2.destroy(),width=150)
          removeButton.place(x=400,y=450)
          top.mainloop()






        # text=Label(newselection, text="Filtered column").place(x=340, y=10)
        # e2=Entry(newselection, width=20).place(x=450, y=10)

        addcusventtree=ttk.Treeview(newselection, height=27)
        addcusventtree["columns"]=["1","2","3", "4","5"]
        addcusventtree.column("#0", width=35, anchor="center")
        addcusventtree.column("1", width=160, anchor="center")
        addcusventtree.column("2", width=160, anchor="center")
        addcusventtree.column("3", width=140, anchor="center")
        addcusventtree.column("4", width=70, anchor="center")
        addcusventtree.column("5", width=70, anchor="center")
        addcusventtree.heading("#0",text="")
        addcusventtree.heading("1",text="ID/SKU")
        addcusventtree.heading("2",text="Product/Service Name")
        addcusventtree.heading("3",text="Unit price")
        addcusventtree.heading("4",text="Service")
        addcusventtree.heading("5",text="Stock")
        addcusventtree.place(x=5, y=45)
        addcusventtree.tag_configure('green', foreground='green')
        addcusventtree.tag_configure('red', foreground='red')
        addcusventtree.tag_configure('blue', foreground='blue')


        def filtm():
        # global filttxt 
     
         filtr = filttxt.get()    
         sqlcusven = "SELECT * FROM Productservice WHERE name=%s"
         valcusven = (filtr, )
         fbcursor.execute(sqlcusven, valcusven)
         records = fbcursor.fetchall()
        #  print(records)
  
         if records==[]:
          pass
         else:  
          for record in addcusventtree.get_children():
           addcusventtree.delete(record)  
          count=0
          for i in records:
            if True:
             addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]))  
            else:
              pass
          count += 1
  
    #   global filttxt
        enter=Label(newselection, text="Enter filter text").place(x=5, y=10)
        filttxt=Entry(newselection, width=20)
        filttxt.place(x=110, y=10)
        fltrbutton=Button(newselection,text = 'Go',command=filtm)
        fltrbutton.place(x=236,y=7,height=25,width=70)
        # fltrbutton=Button(newselection,text = 'Back',command=filtm)
        # fltrbutton.place(x=246,y=7,height=25,width=70)
        


        def cancelnewselection():
         newselection.destroy()

        def selectp1():
            selected = addcusventtree.focus()
            global valuep
            valuep= addcusventtree.item(selected)["values"][0]
            # messagebox.showinfo("",valuep)

            sql = "SELECT * FROM productservice  WHERE productserviceid= %s"
            i=(valuep,)
            fbcursor.execute(sql,i)

            a=0
            j = 0
            for i in fbcursor:
                tree10.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[5],i[7],1,i[8],i[10],(i[7]*1)))
                for line in tree10.get_children():
                  idsave=tree10.item(line)['values'][7]
                  a+=idsave
            j += 1
            pricecol1.config(text=a)





            # rept=fbcursor.execute('"SELECT * FROM Orders"')
            # repeat=rept[0]
            # print(repeat)
            # if not repeat == frck:

            valu10= addcusventtree.item(selected)["values"][0]
            sqllll = "SELECT * FROM productservice  WHERE productserviceid= %s"
            r=(valu10,)
            fbcursor.execute(sqllll,r)
            child=fbcursor.fetchone()

            sql21= 'INSERT INTO storingproduct(Productserviceid,orderid,sku,category,name,description,status,unitprice,peices,cost,taxable,priceminuscost,serviceornot,stock,stocklimit,warehouse,privatenote,quantity) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            vatree=(child[0],frck,child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16],1)
            fbcursor.execute(sql21,vatree,)
            fbilldb.commit()

            # try:
            #   k = 0
            #   for s in fbcursor:
            #     prstree.insert(parent='', index='end', iid=s, text='', values=(s[2],s[4],S[5],s[7],s[13],s[8],S[10],(s[7]*s[13])))
            #   k += 1
            # except:
            #   pass

            newselection.destroy()


            # sql = "SELECT * FROM productservice  WHERE productserviceid= %s"
            # i=(valuep,)
            # x=fbcursor.execute(sql,i)
            # print(x)
            # k = 0
            # for s in x:
            #  prstree.insert(parent='', index='end', iid=s, text='', values=(s[2],s[4],S[5],s[7],s[13],s[8],S[10],(s[7]*s[13])))
            # k += 1


        fbcursor.execute('SELECT * FROM Productservice;') 
        countp = 0
        for i in fbcursor:
                if i[6] == '1':
                    acti = 'Active'
                else:
                    acti = 'Inactive' 
                if i[13] > i[14]:
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))
                  countp += 1
                elif (i[12] =="0") == (i[13] <= i[14]):
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))
                  countp += 1
                elif i[12] == '1':
                  addcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))
                  countp += 1
                else:
                  pass        




        ctegorytree=ttk.Treeview(newselection, height=27)
        ctegorytree["columns"]=["1"]
        ctegorytree.column("#0", width=35, minwidth=20)
        ctegorytree.column("1", width=205, minwidth=25, anchor=CENTER)    
        ctegorytree.heading("#0",text="", anchor=W)
        ctegorytree.heading("1",text="View filter by category", anchor=CENTER)
        ctegorytree.place(x=660, y=45)

     
        def items_selected(event):
          selected_indices = listbox.curselection()
          selected_filter = ",".join([listbox.get(i) for i in selected_indices])

          sql = 'select * from Productservice'
          fbcursor.execute(sql)
          pandsdata = fbcursor.fetchall()
          psql = "select * from Productservice where serviceornot=%s"
          val = ('0', )
          fbcursor.execute(psql, val)
          pdata = fbcursor.fetchall()

          ssql = "select * from Productservice where serviceornot=%s"
          val = ('1', )
          fbcursor.execute(ssql, val)
          sdata = fbcursor.fetchall()

          # pssql = "select * from Productservice where category=%s"
          # psval = (selected_filter, )
          # fbcursor.execute(pssql, psval)
          # pssdata = fbcursor.fetchall()
          if selected_filter == "                           View all records":
            for record in addcusventtree.get_children():
              addcusventtree.delete(record)
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          elif selected_filter == "                           View all products":
            for record in addcusventtree.get_children():
              addcusventtree.delete(record)
            countp = 0
            for i in pdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          elif selected_filter == "                           View all services":
            for record in addcusventtree.get_children():
              addcusventtree.delete(record)
            countp = 0
            for i in sdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                addcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          # elif selected_filter == True:
          #     for record in treeproducts.get_children():
          #       treeproducts.delete(record)
          #     countp = 0
          #     for i in pssdata:
          #       treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i    [2], i[4], i[7], i[13], i[15]))
          #       countp += 1
        # def cat_pro(event):
        #   selected_indices = listbox.curselection()
        #   selected_filt = ",".join([listbox.get(i) for i in selected_indices])
        #   pssql = "select * from Productservice where category=%s"
        #   psval = (selected_filt, )
        #   fbcursor.execute(pssql, psval)
        #   pssdata = fbcursor.fetchall()
        #   print(pssdata)
        #   for record in treeproducts.get_children():
        #     treeproducts.delete(record)
        #   countp = 0
        #   for i in pssdata:
        #     treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i[2], i[4], i[7], i[13], i[15]))
        #     countp += 1


        # sql = "SELECT DISTINCT category FROM Productservice WHERE NOT category = %s AND NOT category = %s"
        # val = ('service','product',)
        # fbcursor.execute(sql,val,)
        # lic = fbcursor.fetchall()
        # print(lic)
        listbox = Listbox(newselection,height = 33,  
                      width = 40,  
                      bg = "white",
                      activestyle = 'dotbox',  
                      fg = "black",
                      highlightbackground="white")  
        listbox.insert(0, "                           View all records")
        listbox.insert(1, "                           View all products")
        listbox.insert(2, "                           View all services")
        listbox.place(x=660,y=65,)
        listbox.bind('<<ListboxSelect>>', items_selected)

        # for nc in lic:
        #     listbox.insert(END, nc)


                



        # scrollbar = Scrollbar(newselection)
        # scrollbar.place(x=640, y=45, height=560)
        # scrollbar.config( command=tree.yview )
        scrollbar = Scrollbar(newselection)
        scrollbar.place(x=1016+300+25, y=120, height=280+20)



        btn1=Button(newselection,compound = LEFT,image=tick ,text="ok", width=60,command=selectp1).place(x=15, y=610)
        btn1=Button(newselection,compound = LEFT,image=tick , text="Edit product/Service", width=150,command=edit_productord).place(x=250, y=610)
        btn1=Button(newselection,compound = LEFT,image=tick , text="Add product/Service", width=150,command=product).place(x=435, y=610)
        btn1=Button(newselection,compound = LEFT,image=cancel ,text="Cancel", width=60,command=cancelnewselection).place(x=740, y=610)
    except:
     try:
      newselection.destroy()
     except:
      pass
    


  #preview new line
  def previewline():
    messagebox.showerror("F-Billing Revolution","line is required,please select customer for this order before printing.")


  
  #sms notification
  def sms1():
    send_SMS=Toplevel()
    send_SMS.geometry("700x480+240+150")
    send_SMS.title("Send SMS notification")

    style = ttk.Style()
    style.theme_use('default')
    style.configure('TNotebook.Tab', background="#999999", padding=5)
    sms_Notebook = ttk.Notebook(send_SMS)
    SMS_Notification = Frame(sms_Notebook, height=470, width=700)
    SMS_Service_Account = Frame(sms_Notebook, height=470, width=700)
    sms_Notebook.add(SMS_Notification, text="SMS Notification")
    sms_Notebook.add(SMS_Service_Account, text="SMS Service Account")
    sms_Notebook.place(x=0, y=0)

    numlbel=Label(SMS_Notification, text="SMS number or comma seperated SMS number list(Please start each SMS number with the country code)")
    numlbel.place(x=10, y=10)
    numentry=Entry(SMS_Notification, width=92).place(x=10, y=30)
    stexbel=Label(SMS_Notification, text="SMS Text").place(x=10, y=60)
    stex=Entry(SMS_Notification, width=40).place(x=10, y=85,height=120)
    
    dclbel=Label(SMS_Notification, text="Double click to insert into text")
    dclbel.place(x=410, y=60)
    dcl=Entry(SMS_Notification, width=30)
    dcl.place(x=400, y=85,height=200)
    
    smstype=LabelFrame(SMS_Notification, text="SMS message type", width=377, height=60)
    smstype.place(x=10, y=223)
    snuvar=IntVar()
    normal_rbtn=Radiobutton(smstype, text="Normal SMS(160 chars)", variable=snuvar, value=1)
    normal_rbtn.place(x=5, y=5)
    unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)", variable=snuvar, value=2)
    unicode_rbtn.place(x=190, y=5)
    tiplbf=LabelFrame(SMS_Notification, text="Tips", width=680, height=120)
    tiplbf.place(x=10, y=290)
    tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="Always start the SMS nymber with the country code. Do not use the + sign at the beginning(example\nUS number:8455807546). Do not use any special characters in your normal SMS text. Please use the\nstndard SMS characters or the English alphabet and numbers only. Otherwise the SMS will be\nunreadable or undeliverable. If you need to enter international characters, accents,email address, or\nspecial characters to the SMS text field then choose the Unicode SMS format.")
    tiplabl.place(x=5, y=5)

    btn1=Button(SMS_Notification, width=20, text="Send SMS notification").place(x=10, y=420)
    btn2=Button(SMS_Notification, width=25, text="Confirm SMS cost before sending").place(x=280, y=420)
    btn3=Button(SMS_Notification, width=15, text="Cancel").place(x=550, y=420)
    

    smstype=LabelFrame(SMS_Service_Account, text="Select the notification service provider", width=670, height=65)
    smstype.place(x=10, y=5)
    snumvar=IntVar()
    normal_rbtn=Radiobutton(smstype,text="BULKSMS(www.bulksms.com)",variable=snumvar,value=1,)
    normal_rbtn.place(x=5, y=5)
    unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)-Recommended", variable=snumvar, value=2)
    unicode_rbtn.place(x=290, y=5)

    sms1type=LabelFrame(SMS_Service_Account, text="Your BULKSMS.COM Account", width=670, height=100)
    sms1type.place(x=10, y=80)
    name=Label(sms1type, text="Username").place(x=10, y=5)
    na=Entry(sms1type, width=20).place(x=100, y=5)
    password=Label(sms1type, text="Password").place(x=10, y=45)
    pas=Entry(sms1type, width=20).place(x=100, y=45)
    combo=Label(sms1type, text="Route").place(x=400, y=5)
    n = StringVar()
    combo1 = ttk.Combobox(sms1type, width = 20, textvariable = n ).place(x=450,y=5)
    btn1=Button(sms1type, width=10, text="Save settings").place(x=550, y=45)

    
    tiplbf=LabelFrame(SMS_Service_Account, text="Terms of service", width=680, height=250)
    tiplbf.place(x=10, y=190)
    tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="The SMS notification service is not free.This service costs you creadit.You must have your own account\nat BULKSMS.COM and you need to have sufficient creadit and an active internet connection to use\nthis feature.Please review all fields in this form for accuracy")
    tiplabl.place(x=0, y=5)
    tiplabl1=Label(tiplbf,justify=LEFT,fg="black",  text="visit www.bulksms.com website to create your own account.please make sure the BULKSMS .COM\n service works well in your country before you busy creadit")
    tiplabl1.place(x=0, y=60)
    tiplabl2=Label(tiplbf,justify=LEFT,fg="black",  text="Our SMS notification tool comes without any warranty.our software only forwards your SMS message\nthe BULKSMS API server .The BULKSMS API server will try to sent SMS message your recipient")
    tiplabl2.place(x=0, y=100)
    tiplabl3=Label(tiplbf,justify=LEFT,fg="red",  text="Please note that you access and use the SMS notification tool your own risk.F-Billing software is not\nresponsible for any type of loss or damage or undelivered SMS massage which you may as a result\nof accessing and using the SMS notification service.")
    tiplabl3.place(x=0, y=140)
    checkvar1=IntVar()
    chkbtn1=Checkbutton(tiplbf,text="I have read and agree to the terms of service above",variable=checkvar1,onvalue=1,offvalue=0).place(x=70, y=200) 



  
  #delete line item  
  def delete1():
      # selected_item = tree10.selection()[0]
      # print=(selected_item)
      # tree10.delete(selected_item)
      
      xyz = tree10.item(tree10.focus())["values"][0]
      quantityidcr = tree10.item(tree10.focus())["values"][4]
      unitpriceidcr = tree10.item(tree10.focus())["values"][3]
      nameidcr = tree10.item(tree10.focus())["values"][1]
      descriidcr = tree10.item(tree10.focus())["values"][2]
      pcsidcr = tree10.item(tree10.focus())["values"][5]
      print(xyz,)
      dltnow=frck
      sql = 'DELETE FROM storingproduct WHERE orderid=%s AND quantity=%s AND Productserviceid=%s AND peices=%s AND unitprice=%s AND name=%s AND description=%s'
      val = (dltnow,quantityidcr,xyz,pcsidcr,unitpriceidcr,nameidcr,descriidcr)
      fbcursor.execute(sql, val)
      fbilldb.commit()
      tree10.delete(tree10.selection()[0])

      for record in tree10.get_children():
                tree10.delete(record)
        
      sql = "SELECT * FROM storingproduct WHERE orderid= %s"
      i=(dltnow,)
      fbcursor.execute(sql,i)
  
      a=0
      j = 0
      for i in fbcursor:
        tree10.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))
        for line in tree10.get_children():
            idsave=tree10.item(line)['values'][7]
        a+=idsave
      j += 1
      pricecol1.config(text=a)
      # delmess = messagebox.askyesno("Delete Product", "Are you sure to delete this Product?")
      # if delmess == True:
      #   itemid = vwedttree1.item(vwedttree1.focus())["values"][1]
      #   print(itemid,)
      #   sql = 'DELETE FROM Orders WHERE orderid=%s'
      #   val = (itemid,)
      #   fbcursor.execute(sql, val)
      #   fbilldb.commit()
      #   ordtree.delete(ordtree.selection()[0])

      #   sqlstrngprdct='DELETE FROM storingproduct WHERE orderid=%s'
      #   valstpr = (itemid,)
      #   fbcursor.execute(sqlstrngprdct, valstpr)
      #   fbilldb.commit()
      #   vwedttree1.delete(vwedttree1.selection()[0])
      # else:
      #  pass

  # #delete orders  
  # def dele():
  #   delmess = messagebox.askyesno("Delete Invoice", "Are you sure to delete this Invoice?")
  #   messagebox.showerror("F-Billing Revolution","Customer is required,please select customer before deleting line item .")

  #   if delmess == True:
  #     itemid = ordtree.item(ordtree.focus())["values"][1]
  #     print(itemid,)
  #     sql = 'DELETE FROM Orders WHERE orderid=%s'
  #     val = (itemid,)
  #     fbcursor.execute(sql, val)
  #     fbilldb.commit()
  #     ordtree.delete(ordtree.selection()[0])
  #   else:
  #     pass  
################################################ CREATE ORDER #############################################
  def creatingorder():
  #  if ot1v.cget("text")==NULL:
  #     messagebox.showerror('F-Billing Revolution', 'Add a customer before adding new line.') 
  #  else: 
    cmbodto=cmb.get()
    addrsfrm=addrs.get('1.0', 'end-1c')
    sptto=spt.get()
    adrsto=adrs1.get('1.0', 'end-1c')
    emlfrm=eml.get()
    smsfrm=smsno.get()
    # ordrid=ord.get()
    orddatein=orddte.get_date()
    ordduein=duedte.get_date()
    trmsin=trms.get()
    extracstnme=xtracstnme.get()
    discountrte=dscntrtecr.get()
    extracst=xtracstcr.get()
    taax=taxxcr.get()
    tplts=tmplte.get()
    slzprzn=salesprsn.get()
    ctgryy=ctgry.get()

    ab="Draft"

  


    # for line in tree10.get_children():
    #   idsave=tree10.item(line)['values'][0]

    #   sql1= 'SELECT * FROM  Productservice WHERE Productserviceid = %s'
    #   val=(idsave,)
    #   print(val)
    #   fbcursor.execute(sql1,val,)
    #   child=fbcursor.fetchone()
    #   print(child)
    #   sql2= 'INSERT INTO storingproduct(Productserviceid,orderid,sku,category,name,description,status,unitprice,peices,cost,taxable,priceminuscost,serviceornot,stock,stocklimit,warehouse,privatenote) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
    #   val1=(child[0],ordrid,child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16])
    #   fbcursor.execute(sql2,val1,)
    #   fbilldb.commit()
      
    # sss='INSERT INTO orders WHERE Order_total=%s'
    # vvv=('NULL')
    # fbcursor.execute(sss,vvv,)
    # fbilldb.commit()
    
    # sql3='UPDATE Orders SET  order_date=%s, due_date=%s, businessname=%s, status=%s, extra_cost_name=%s, extra_cost=%s, template=%s, sales_person=%s, discount_rate=%s, tax1=%s, category=%s, businessaddress=%s, shipname=%s, shipaddress=%s, cpemail=%s, cpmobileforsms=%s,Order_total=%s WHERE orderid=%s'
    # val2=(orddatein,ordduein,cmbodto,ab,extracstnme,extracst,tplts,slzprzn,discountrte,taax,ctgryy,addrsfrm,sptto,adrsto,emlfrm,smsfrm,ot1v,frck)
    # fbcursor.execute(sql3,val2,)
    # fbilldb.commit()


    # def check_empty() :
    #  if taxxcr.get():
    #      pass     #your function where you want to jump
    #  else:
    #     print(' input required')

    ttlprzinptcr=pricecol1.cget("text")
    ratediscntcr=dscntrtecr.get()
    xtaxcr=taxxcr.get()
    cstxtracr=xtracstcr.get()
    print(ttlprzinptcr,ratediscntcr,xtaxcr,cstxtracr)
 
    sdd=round(int(ratediscntcr),3)
    pldd=int(ttlprzinptcr)
    p=round(pldd, 4)
    dsctedd=(sdd*pldd)/100
    dscdd=round(dsctedd, 4)
    discount=Label(summaryfrme, text=str(sdd)+"% Discount").place(x=0 ,y=0)
    discount1cr=Label(summaryfrme,text='Rs. '+str(dscdd))
    discount1cr.place(x=130 ,y=0)
 
    sbtotlhee=p-dscdd
    subtotl=round(sbtotlhee, 4)
    sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
    sub100cr=Label(summaryfrme, text='Rs. '+str(subtotl))
    sub100cr.place(x=130 ,y=21)
 
    taxval91=int(xtaxcr)
    tax1dsplay=(taxval91*subtotl)/100
    taxvalu=round(tax1dsplay,2)
    tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
    tax100cr=Label(summaryfrme, text='Rs. '+str(taxvalu))
    tax100cr.place(x=130 ,y=42)
 
    xtracstvaluzz=int(cstxtracr)
    xtra1cst=round(xtracstvaluzz, 4)
    cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
    cost100cr=Label(summaryfrme, text='Rs. '+str(xtra1cst))
    cost100cr.place(x=130 ,y=63)
  
    otherttlval=subtotl+taxvalu+xtra1cst
    ot1v=round(otherttlval, 4)
    order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
    order100cr=Label(summaryfrme, text='Rs. '+str(ot1v))
    order100cr.place(x=130 ,y=84)
 

    sql3='UPDATE Orders SET  order_date=%s, due_date=%s, businessname=%s, status=%s, extra_cost_name=%s, extra_cost=%s, template=%s, sales_person=%s, discount_rate=%s, tax1=%s, category=%s, businessaddress=%s, shipname=%s, shipaddress=%s, cpemail=%s, cpmobileforsms=%s,Order_total=%s,sum_discount=%s,sum_tax=%s,sum_subtotal=%s WHERE orderid=%s'
    val2=(orddatein,ordduein,cmbodto,ab,extracstnme,extracst,tplts,slzprzn,sdd,taax,ctgryy,addrsfrm,sptto,adrsto,emlfrm,smsfrm,ot1v,dscdd,taxvalu,subtotl,frck)
    fbcursor.execute(sql3,val2,)
    fbilldb.commit()
    vel=addnote.get('1.0','end-1c')
    adntsql="UPDATE orders SET private_notes=%s WHERE orderid=%s"
    adntsval=(vel,frck)
    fbcursor.execute(adntsql,adntsval)
    fbilldb.commit()

    # for recordat in cusventtreead.get_children():
    #  print("sckrr",recordat)

    att="SELECT add_document FROM documents WHERE orderid=%s"
    valat=(frck,)
    fbcursor.execute(att,valat)
    attdta = fbcursor.fetchone()
    print(attdta)
    if attdta == (None,):
     zuk="NULL"
     aat="UPDATE document SET add_document=%s WHERE orderid=%s"
     aats=(zuk,frck)
     fbcursor.execute(aat,aats)
     fbilldb.commit()


    dropordr.delete(0,END)
    notepriv.delete('1.0','end')
    for record in prstree.get_children():
     prstree.delete(record)
    for record in treeblw.get_children():
     treeblw.delete(record)
    for record in ordtree.get_children():
      ordtree.delete(record)
    fbcursor.execute("select * from Orders")
    # pandsdata = fbcursor.fetchall()
    # countp = 0
    # for i in pandsdata:
    #   ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))

    # ordertotalinput=0
    # for line in ordtree.get_children():
    #  idsave1=ordtree.item(line)['values'][9]
    # ordertotalinput += idsave1
    # countp += 1
    j = 0
    for i in fbcursor:
     ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    
     ordertotalinput=0
     for line in ordtree.get_children():
      idsave1=ordtree.item(line)['values'][9]
      ordertotalinput += idsave1
    j += 1
    ordtotalrowcol.config(text=ordertotalinput)

    # ordtotalrowcol.config(text=ordertotalinput)
    messagebox.showinfo('Successfully Added','Successfully Added')

  def addprintsele():
    printordinv = messagebox.askyesno("Download to Print", "Continue to Download?")
    if printordinv == True:
      # itemid = ordtree.item(ordtree.focus())["values"][1]
      # print(itemid,)
      # ordrid=sctxt.get()
      sql = 'SELECT * FROM Orders WHERE orderid=%s'
      val = (frck,)
      fbcursor.execute(sql, val)
      koko=fbcursor.fetchone()

      # # fbilldb.commit()
      # selected = ordtree.focus()
      # selected_prdct= ordtree.item(selected)["values"][1]

      dateee=dt.datetime.now()
      for record in ordtree.get_children():
       ordtree.delete(record)
      sqlq = "UPDATE Orders SET printed_on=%s WHERE orderid = %s"
    # 'UPDATE storingproduct SET Productserviceid=%s
      valq = (dateee,frck, )
      fbcursor.execute(sqlq, valq,)
      fbilldb.commit()
      fbcursor.execute('SELECT * FROM Orders;')
      ordertotalinput=0
      j = 0
      for i in fbcursor:
       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
       for line in ordtree.get_children():
        idsave1=ordtree.item(line)['values'][9]
       ordertotalinput += idsave1
       j += 1
    
      sqlprint='SELECT * FROM storingproduct WHERE orderid=%s'
      valprint=(frck,)
      fbcursor.execute(sqlprint,valprint)
      mama=fbcursor.fetchall()
      # ordtree.delete(ordtree.selection()[0])
      # sqlstrngprdct='SELECT * FROM storingproduct WHERE orderid=%s'
      # valstpr = (itemid,)
      # fbcursor.execute(sqlstrngprdct, valstpr)
      # mama=fbcursor.fetchone()
      # fbilldb.commit()
      # print("HELLO",koko)
      # print("HAI",mama)
      # prdisc=discount1.cget("text")
    str_html = ""

    str_html+="""
                    <!doctype>
                    <html>
                        <head>
                            <style>
                                .header {
                                    text-align:center;
                                }
                                .body {
                                    font-size: 1.5rem;
                                    font-weight:normal;
                                }
                                table {
                                    
                                    width:100%
                                }
                                td{
                                    text-align: center;
                                    padding-top: 1%;
                                }
                                .thead .th{
                                    border-top: 1px solid #333;
                                    border-bottom: 1px solid #333;
                                    text-align: center;
                                    font-weight: bold;
                                }
                            </style>
                        </head>


                        <body>

                            <div class="header">
                                <h2>Order Invoice</h2>
                            </div>
                            <br><br><br><br><br><br><br>
                            <div>
                                <h3 >Order Id &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[0])+"""</h3>      

                                <h3 >Order Date &nbsp;&nbsp;"""+str(koko[1])+"""</h3>                             

                                <h3 >Due Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[2])+"""</h3>                             
                            </div>                                                        
                            <br><br>
                            <div style="display:flex;">
                                <div>
                                <h2>Order To</h2><h3>"""+str(koko[3])+"""<br>"""+str(koko[16])+"""</h3>
                                </div>
                                <div>
                                <h2>Ship To</h2><h3>"""+str(koko[17])+"""<br>"""+str(koko[18])+"""</h3>                             
                                </div>                       
                            </div>
                            

                            <div class="body">

                                <table>
                                    <tr class="thead" ; style="background-color:orange">
                                        <td class="th">ID/SKU</td>
                                        <td class="th">Product/Service</td>
                                        <td class="th">Description</td>
                                        <td class="th">Quantity</td>
                                        <td class="th">Unit Price</td>
                                        <td class="th">Price</td>
                                    </tr>

                                    <tbody>"""

    for pri in mama:
     str_html+="""<tr>

                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[4])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[6])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[7])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[22])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(pri[9])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str((pri[22])*(pri[9]))+"""</td>
 
                        </tr>"""

    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th style="border: 1px solid black;border-collapse: collapse;">"""+str(koko[13])+'%  Discount'+"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[24])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Subtotal'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[26])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'TAX1'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[25])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Extra Cost'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[10])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Order Total'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[8])+"""</td>
                        </tr>"""
  
    str_html+="""</tbody>
                    </table>
                </div>
            </body>
        </html>"""

    
    # import pdb;pdb.set_trace()


    html_data = str_html

    file_path = os.getcwd()+'/images/'


    options = {'margin-top': '15.00mm',
                    'margin-right': '15.00mm',
                    'margin-bottom': '15.00mm',
                    'margin-left': '15.00mm',
                    'dpi':300,
                    }

    lst_pdfs = []
    str_file_name ='ORD_'+str(koko[0])+'.pdf'
    filename_ledger =  file_path+'/'+str_file_name
    path_wkthmltopdf = b'C:\Program Files\wkhtmltopdf\\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)
    pdfkit.from_string(html_data,filename_ledger,options=options,configuration=config)
    lst_pdfs.append(filename_ledger)

    pdf_writer = PdfFileWriter()

    # pop.destroy()

      # save1111=save1111.append(idsave)
      # print(save1111)
    # for i in idsave:
    #   print(idsave[i])
      # for value in tree10.item(line)['values']:
        # print(value[0])
      

    # prdctidvar=valuep
    # print(prdctidvar)
    

      

      #######################################################################################################################################################
  # def Add():
  #     order = e1.get()
  #     address = e2.get()
  #     ship = e3.get()
  #     address1 = e4.get()
  #     email = e5.get()
  #     sms = e6.get()
  
  #     mysqldb=mysql.connector.connect(host="localhost",user="root",password="",database="fbilldb" , port="3306")
  #     fbcursor=mysqldb.cursor()
  
  #     try:
  #       sql = "INSERT INTO  Customer (businessname,businessaddress,shipname,shipaddress,cpemail,cpmobileforsms) VALUES (%s, %s, %s, %s, %s, %s)"
  #       val = (order,address,ship,address1,email,sms)
  #       fbcursor.execute(sql, val)
  #       mysqldb.commit()
  #       lastid = fbcursor.lastrowid
  #       #messagebox.showinfo("information", "Employee inserted successfully...")
  #       e1.delete(0, END)
  #       e2.delete(0, END)
  #       e3.delete(0, END)
  #       e4.delete(0, END)
  #       e5.delete(0, END)
  #       e6.delete(0, END)
  #       e1.focus_set()
  #     except Exception as e:
  #       print(e)
  #       mysqldb.rollback()
  #       mysqldb.close()

  # def addtotalprice():
  #   global Sum
  #   Sum=0.0
  #   for line in tree10.get_children():
  #      pricetotl=tree10.item(line)['values'][7]
  #      Sum+=pricetotl
  #   return Sum   


  def send_mail():

      sender_email = email_from.get()    
      sender_password = email_pswrd.get() 

      server = smtplib.SMTP('smtp.gmail.com', 587)
      print("Login successful1")
      server.starttls()
      print("Login successful2")
      server.login(sender_email, sender_password)
      print("Login successful3")
      
      # global email_address, email_subject

      # email_address = StringVar()  
      # email_subject = StringVar()
      # msg = MIMEMultipart()

      # for i in htcodeframe.curselection():
      #   # print("hloo",htcodeframe.get(i))
      #   ya=htcodeframe.get(i)
      #   print(ya,"THIS")

      #   fo = open(ya, "rb")
      #   filecontent = fo.read()
      #   encodedcontentnw = base64.b64encode(filecontent) 


      # address_info = email_address.get()
      # print(address_info)
      # # msg['Subject'] = email_subject.get()
      # subject_info = email_subject.get()
      # print(subject_info)
      carbcopy_info = carcopyem_address.get()
      print(carbcopy_info)
      # email_content=mframe.get('1.0','end-1c')
      # print(email_content)
      
      
      msg = MIMEMultipart()
      msg['Subject'] = email_subject.get() 
      mail_content  = mframe.get('1.0','end-1c') 
      msg['From'] = email_from.get()
      msg['To'] = email_address.get()
      # msg.attach(MIMEText(file("text.txt").read()))

      # msg.attach(MIMEImage(open('images/'+filenamez.split('/')[-1],"rb").read()))
      # for i in htcodeframe.curselection():
      
      gettingimg=lstfrm.get()
      lst_data = gettingimg[1:-1].split(',')
      print(lst_data,"happy")
      # print(gettingimg)
      msg.attach(MIMEText(mail_content, 'plain'))

      for i in lst_data:
        # print(i[0],"IMAGE")
        with open('images/'+ i.strip()[1:-1], "rb") as attachment:
              # MIME attachment is a binary file for that content type "application/octet-stream" is used
          part = MIMEBase("application", "octet-stream")
          part.set_payload(attachment.read())
          # encode into base64 
        encoders.encode_base64(part) 

        part.add_header('Content-Disposition', "attachment; filename= %s" % 'images/'+ i.strip()[1:-1]) 

          # attach the instance 'part' to instance 'message' 
        msg.attach(part)
      # message_body = email_body.get()

      server.sendmail(email_from.get(),email_address.get(),msg.as_string())
      server.sendmail(email_from.get(), carbcopy_info,msg.as_string())


      dateeem=dt.datetime.now()
      emitemid = ordtree.item(ordtree.focus())["values"][1]
      for record in ordtree.get_children():
        ordtree.delete(record)
      sqlq = "UPDATE Orders SET emailed_on=%s WHERE orderid = %s"
      valq = (dateeem,emitemid, )
      fbcursor.execute(sqlq, valq,)
      fbilldb.commit()
      fbcursor.execute('SELECT * FROM Orders;')
      ordertotalinput=0
      j = 0
      for i in fbcursor:
       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
      for line in ordtree.get_children():
        idsave1=ordtree.item(line)['values'][9]
      ordertotalinput += idsave1
      j += 1
      ordtotalrowcol.config(text=ordertotalinput)

      print("message sent")
      
      # subent.delete(0, END)
      # emailtoent.delete(0, END)

  def empsfile_image(event):
          global yawn
          for i in htcodeframe.curselection():
            print("hloo",htcodeframe.get(i))
            yawn=htcodeframe.get(i)        
            edit_window_img = Toplevel()
            edit_window_img.title("View Image")
            edit_window_img.geometry("700x500")
            image = Image.open("images/"+yawn)
            resize_image = image.resize((700, 500))
            image = ImageTk.PhotoImage(resize_image)
            psimage = Label(edit_window_img,image=image)
            psimage.photo = image
            psimage.pack()

  def UploadAction(event=None):
      global filenamez
      # filename = filedialog.askopenfilename()
      # print('Selected:', filename)
      # name = askopenfilename(filetypes=[('PDF', '*.pdf',)])

      filenamez = askopenfilename(filetypes=(("png file ",'.png'),("jpg file", ".jpg"), ('PDF', '*.pdf',), ("All files", "*.*"),))
      shutil.copyfile(filenamez, os.getcwd()+'/images/'+filenamez.split('/')[-1])
      htcodeframe.insert(0, filenamez.split('/')[-1])


      # add this
    
      # def upload_file1():
      #  global filename,img, b1
      #  f_types =[('Png files','.png'),('Jpg Files', '.jpg')]
      #  filename = filedialog.askopenfilename(filetypes=f_types)
      #  print(filename, 'name')
      #  #import pdb; pdb.set_trace()
      #  shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
      #  image = Image.open(filename)
      #  resize_image = image.resize((120, 120))
      #  img = ImageTk.PhotoImage(resize_image)
      #  b1 = Label(expenselabelframe,image=img, height=120, width=120)
      #  b1.place(x=450 , y=240)


    ###################DB
    #file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
      #query 
      #val=filename.split('/')[-1]
    
  def addemail_order():
    emitid = frck
    sql = "select * from Orders where orderid = %s"
    val = (emitid, )
    fbcursor.execute(sql, val)
    emailnow = fbcursor.fetchone()
    mailDetail=Toplevel()
    mailDetail.title("Order E-Mail")
    mailDetail.geometry("1080x550")
    mailDetail.resizable(False, False)
    # def my_SMTP():
    #     if True:
    #         em_ser_conbtn.destroy()
    #         mysmtpservercon=LabelFrame(account_Frame,text="SMTP server connection(ask your ISP for your SMTP settings)", height=165, width=380)
    #         mysmtpservercon.place(x=610, y=110)
    #         lbl_hostn=Label(mysmtpservercon, text="Hostname").place(x=5, y=10)
    #         hostnent=Entry(mysmtpservercon, width=30).place(x=80, y=10)
    #         lbl_portn=Label(mysmtpservercon, text="Port").place(x=5, y=35)
    #         portent=Entry(mysmtpservercon, width=30).place(x=80, y=35)
    #         lbl_usn=Label(mysmtpservercon, text="Username").place(x=5, y=60)
    #         unament=Entry(mysmtpservercon, width=30).place(x=80, y=60)
    #         lbl_pasn=Label(mysmtpservercon, text="Password").place(x=5, y=85)
    #         pwdent=Entry(mysmtpservercon, width=30).place(x=80, y=85)
    #         ssl_chkvar=IntVar()
    #         ssl_chkbtn=Checkbutton(mysmtpservercon, variable=ssl_chkvar, text="This server requires a secure connection(SSL)", onvalue=1, offvalue=0)
    #         ssl_chkbtn.place(x=50, y=110)
    #         em_ser_conbtn1=Button(account_Frame, text="Test E-mail Server Connection").place(x=610, y=285)
    #     else:
    #         pass

  
    style = ttk.Style()
    style.theme_use('default')
    style.configure('TNotebook.Tab', background="#999999", padding=5)
    email_Notebook = ttk.Notebook(mailDetail)
    email_Frame = Frame(email_Notebook, height=500, width=1080)
    account_Frame = Frame(email_Notebook, height=550, width=1080)
    email_Notebook.add(email_Frame, text="E-mail")
    email_Notebook.add(account_Frame, text="Account")
    email_Notebook.place(x=0, y=0)

    messagelbframe=LabelFrame(email_Frame,text="Message", height=500, width=730)
    messagelbframe.place(x=5, y=5)
    global email_address, email_subject, email_from,email_pswrd,carcopyem_address,mframe,htcodeframe,lstfrm,langs
    email_address = StringVar() 
    email_subject = StringVar()
    # email_body = StringVar()
    email_from = StringVar()
    email_pswrd = StringVar()
    carcopyem_address = StringVar()
    # content_email = StringVar()
    lbl_emailtoaddr=Label(messagelbframe, text="Email to address").place(x=5, y=5)
    emailtoent=Entry(messagelbframe, width=50,textvariable=email_address)
    emailtoent.place(x=120, y=5)
    emailtoent.delete(0,'end')
    emailtoent.insert(0, emailnow[19])
    #email2 = email_address.get()
    #print(email2)
    sendemail_btn=Button(messagelbframe, text="Send Email", width=10, height=1, command=send_mail).place(x=600, y=10)#,command=addacnt

    lbl_carcopyto=Label(messagelbframe, text="Carbon copy to").place(x=5, y=32)
    carcopyent=Entry(messagelbframe, width=50,textvariable=carcopyem_address)
    carcopyent.place(x=120, y=32)
    # stopemail_btn=Button(messagelbframe, text="Stop sending", width=10, height=1).place(x=600, y=40)
    lbl_subject=Label(messagelbframe, text="Subject").place(x=5, y=59)
    subent=Entry(messagelbframe, width=50, textvariable=email_subject)
    subent.place(x=120, y=59)
    subjectinsrt='ORD_'+str(emailnow[0])
    subent.delete(0,'end')
    subent.insert(0, subjectinsrt)

    
    style = ttk.Style()
    style.theme_use('default')
    style.configure('TNotebook.Tab', background="#999999", width=20, padding=5)
    mess_Notebook = ttk.Notebook(messagelbframe)
    emailmessage_Frame =Frame(mess_Notebook, height=350, width=710)
    htmlsourse_Frame = Frame(mess_Notebook, height=350, width=710)
    mess_Notebook.add(emailmessage_Frame, text="E-mail message")
    # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
    # mess_Notebook.place(x=120, y=59)
    mess_Notebook.add(htmlsourse_Frame, )#text="Html source code")
    mess_Notebook.place(x=5, y=90)

    # btn1=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

    
    # btn2=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
    # btn3=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
    # btn4=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
    # btn5=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=undo).place(x=140, y=1)
    # btn6=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=redo).place(x=175, y=1)
    # btn7=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=bold).place(x=210, y=1)
    # btn8=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=italics).place(x=245, y=1)
    # btn9=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=underline).place(x=280, y=1)
    # btn10=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=left).place(x=315, y=1)
    # btn11=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=right).place(x=350, y=1)
    # btn12=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=center).place(x=385, y=1)
    # btn13=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=hyperlink).place(x=420, y=1)
    
    # btn14=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=remove).place(x=455, y=1)


    # dropcomp = ttk.Combobox(emailmessage_Frame, width=12, height=3).place(x=500, y=5)
    # dropcompo = ttk.Combobox(emailmessage_Frame, width=6, height=3).place(x=600, y=5)
    mframe=scrolledtext.Text(emailmessage_Frame,  undo=True,width=88, bg="white", height=22)
    mframe.place(x=0, y=10)
    # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
    # mess_Notebook.place(x=120, y=59)


    # btn1=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

    
    # btn2=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
    # btn3=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
    # btn4=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
    # mframe1=Frame(htmlsourse_Frame, height=350, width=710, bg="white")
    # mframe1.place(x=0, y=28)
    attachlbframe=LabelFrame(email_Frame,text="Attachment(s)", height=350, width=280)
    attachlbframe.place(x=740, y=5)
    # langs=[]
    lstfrm=StringVar()  
    htcodeframe=Listbox(attachlbframe, height=13, width=43,listvariable=lstfrm, bg="white")
    htcodeframe.place(x=5, y=5)
    htcodeframe.bind('<Double-Button-1>' , empsfile_image)

    def deslist():
        laa=htcodeframe.curselection()
        print("hloo",htcodeframe.get(laa))
        yawn=htcodeframe.get(laa)        
        htcodeframe.delete(ACTIVE)

    lbl_btn_info=Label(attachlbframe, text="Double click on attachment to view").place(x=30, y=230)
    btn17=Button(attachlbframe, width=20, text="Add attachment file...", command=UploadAction).place(x=60, y=260)
    btn18=Button(attachlbframe, width=20, text="Remove attachment",command=deslist).place(x=60, y=295)
    lbl_tt_info=Label(email_Frame, text="You can create predefined invoice, order, estimate\nand payment receipt email templates under Main\nmenu/Settings/E-Mail templates tab")
    lbl_tt_info.place(x=740, y=370)

    ready_frame=Frame(mailDetail, height=20, width=1080, bg="#b3b3b3").place(x=0,y=530)
    
    sendatalbframe=LabelFrame(account_Frame,text="E-Mail(Sender data)",height=140, width=600)
    sendatalbframe.place(x=5, y=5)
    lbl_sendermail=Label(sendatalbframe, text="Company email address").place(x=5, y=10)
    sentent=Entry(sendatalbframe, width=40, textvariable=email_from)
    sentent.place(x=195, y=10)
    #############################################
    lbl_senderpswrd=Label(sendatalbframe, text="Email Password").place(x=5, y=40)
    pswrdent=Entry(sendatalbframe, width=40, textvariable=email_pswrd,show="*")
    pswrdent.place(x=195, y=40)
    #############################################
    # lbl_orcompanyname=Label(sendatalbframe, text="Your name or company name").place(x=5, y=70)
    # nament=Entry(sendatalbframe, width=40).place(x=195, y=70)
    # lbl_reply=Label(sendatalbframe, text="Reply to email address").place(x=5, y=100)
    # replyent=Entry(sendatalbframe, width=40).place(x=195, y=100)
    # lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=140)
    # signent=Entry(sendatalbframe,width=50).place(x=100, y=140,height=75)
    # @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    # lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=95)
    # signent=Entry(sendatalbframe,width=50).place(x=195, y=70,height=95)
    # confirm_chkvar=IntVar()
    # confirm_chkbtn=Checkbutton(sendatalbframe, variable=confirm_chkvar, text="Confirmation reading", onvalue=1, offvalue=0)
    # confirm_chkbtn.place(x=200, y=215)
    # btn18=Button(account_Frame, width=15, text="Save settings",command=savesettings).place(x=25, y=285)

    # sendatalbframe=LabelFrame(account_Frame,text="SMTP Server",height=100, width=380)
    # sendatalbframe.place(x=610, y=5)
    # servar=IntVar()
    # SMTP_rbtn=Radiobutton(sendatalbframe, text="Use the Built-In SMTP Server Settings", variable=servar, value=1)
    # SMTP_rbtn.place(x=10, y=10)
    # MySMTP_rbtn=Radiobutton(sendatalbframe, text="Use My Own SMTP Server Settings(Recommended)", variable=servar, value=2, command=my_SMTP)
    # MySMTP_rbtn.place(x=10, y=40)
    # em_ser_conbtn=Button(account_Frame, text="Test E-mail Server Connection")
    # em_ser_conbtn.place(x=710, y=110)

        
      

  firFrame=Frame(pop, bg="#f5f3f2", height=60)
  firFrame.pack(side="top", fill=X)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  create = Button(firFrame,compound="top", text="Select\nCustomer",relief=RAISED, image=customer,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=custom)
  create.pack(side="left", pady=3, ipadx=4)


  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  add= Button(firFrame,compound="top", text="Add new\nline item",relief=RAISED, image=photo,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=newline)
  add.pack(side="left", pady=3, ipadx=4)

  dele= Button(firFrame,compound="top", text="Delete line\nitem",relief=RAISED, image=photo2,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=delete1)
  dele.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  # prev= Button(firFrame,compound="top", text="Preview\nOrder",relief=RAISED, image=photo5,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=previewline)
  # prev.pack(side="left", pady=3, ipadx=4)

  prin= Button(firFrame,compound="top", text="Print \nOrder",relief=RAISED, image=photo4,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=addprintsele)
  prin.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  mail= Button(firFrame,compound="top", text="Email\nOrder",relief=RAISED, image=photo6,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=addemail_order)
  mail.pack(side="left", pady=3, ipadx=4)

  sms1= Button(firFrame,compound="top", text="Send SMS\nnotification",relief=RAISED, image=photo10,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=sms1)
  sms1.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  calc= Button(firFrame,compound="top", text="Open\nCalculator",relief=RAISED, image=photo9,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=calcu)
  calc.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="right", padx=5)

  Createorder= Button(firFrame,compound="top", text="Save",relief=RAISED, image=tick,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=creatingorder)
  Createorder.pack(side="right", pady=3, ipadx=4)

  fir1Frame=Frame(pop, height=180,bg="#f5f3f2")
  fir1Frame.pack(side="top", fill=X)

    
  # fbcursor.execute('SELECT * FROM Productservice;') 
  # j = 0
  # for i in fbcursor:
  #     labelframe1.insert(parent='', index='end', iid=i, text='', values=(' ',i[0],i[4],i[5],i[7],i[8],i[10],))
  #     j += 1


  labelframe1 = LabelFrame(fir1Frame,text="Customers",font=("arial",15))
  labelframe1.place(x=10,y=5,width=640,height=160)





  # sql = "select Customer.businessname from Customer where Customer.customerid = Orders.customerid"
  # sql = "select Customer.businessname from Customer INNER JOIN Orders ON Customer.customerid = Orders.customerid"


#  global slctcstmrtree
  def treefthcng():
    global slctcstr
    itemid = slctcstmrtree.item(slctcstmrtree.focus())["values"][0]
    sql = "select * from Customer where customerid = %s"
    val = (itemid, )
    fbcursor.execute(sql, val)
    slctcstr = fbcursor.fetchone()
    cmb.delete(0,'end')
    cmb.insert(0, slctcstr[4])
    addrs.delete('1.0','end')
    addrs.insert("1.0", slctcstr[5])
    spt.delete(0,'end')
    spt.insert(0, slctcstr[6])
    adrs1.delete('1.0','end')
    adrs1.insert('1.0', slctcstr[7])
    eml.delete(0,'end')
    eml.insert(0, slctcstr[9])
    smsno.delete(0,'end')
    smsno.insert(0, slctcstr[10])

    cmbodto=cmb.get()
    addrsfrm=addrs.get('1.0', 'end-1c')
    sptto=spt.get()
    adrsto=adrs1.get('1.0', 'end-1c')
    emlfrm=eml.get()
    smsfrm=smsno.get()
    orddatein1=orddte.get_date()
    ordduein1=duedte.get_date()
    trmsin=trms.get()
    extracstnme=xtracstnme.get()
    discountrte=dscntrtecr.get()
    extracst=xtracstcr.get()
    taax=taxxcr.get()
    tplts=tmplte.get()
    slzprzn=salesprsn.get()
    ctgryy=ctgry.get()
    # ordrid=ord.get()
    


    insordttl=0
    sql3='INSERT INTO Orders ( extra_cost,orderid,businessname, businessaddress, shipname, shipaddress, cpemail, cpmobileforsms, Order_total,order_date,due_date) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
    val2=(extracst,frck,cmbodto,addrsfrm,sptto,adrsto,emlfrm,smsfrm,insordttl,orddatein1,ordduein1)
    fbcursor.execute(sql3,val2,)
    fbilldb.commit()
    
    cuselection.destroy()

  # def autocstmr(event):
  #  zoo = cmb.get()
  #  sql = "SELECT * FROM Customer WHERE businessname = %s"
  #  val = (zoo, )
  #  fbcursor.execute(sql, val)
  #  cstmrauto = fbcursor.fetchone()
  #  print(cstmrauto)
 
  #  addrs.delete('1.0','end')
  #  addrs.insert("1.0", cstmrauto[5])
  #  spt.delete(0,'end')
  #  spt.insert(0, cstmrauto[6])
  #  adrs1.delete('1.0','end')
  #  adrs1.insert('1.0', cstmrauto[7])
  #  eml.delete(0,'end')
  #  eml.insert(0, cstmrauto[9])
  #  smsno.delete(0,'end')
  #  smsno.insert(0, cstmrauto[10])
  #  sqlau='INSERT INTO Orders ( orderid,businessname, businessaddress, shipname, shipaddress, cpemail, cpmobileforsms) VALUES (%s,%s,%s,%s,%s,%s,%s)'
  #  valau=(frck,cmb,addrs,spt,adrs1,eml,smsno)
  #  fbcursor.execute(sqlau,valau,)
  #  fbilldb.commit()

  sql = "select businessname from Customer"
  fbcursor.execute(sql,)
  pdata = fbcursor.fetchall()
  order = Label(labelframe1, text="Order to").place(x=10,y=5)
  # cmb = ttk.Combobox(labelframe1,width=31)
  # cmb['values']=pdata
  cmb = Entry(labelframe1,width=31)
  cmb.place(x=80,y=5)
  # cmb.bind("<<ComboboxSelected>>",autocstmr)
  ## q = e1.focus()

  # sql = "select * from Customer where businessname=%s"
  # i=(e1)
  # fbcursor.execute(sql,i,)
  # sltn = fbcursor.fetchone()

  # sql = "select Customer.businessaddress from Customer where pdata = Orders.businessname"
  # fbcursor.execute(sql,)
  # addrs = fbcursor.fetchall()
  # print(addrs)                                values="addrs",

  # fbcursor=mysql.cursor()
  # fbcursor.execute("SELECT businessname FROM Customer")
  # myresult= fbcursor.fetchone()
  # for row in myresult:
  #   print(row)

  # sql = "select businessname from Customer"
  # sql1 = "select businessname from Order"

  # if sql == sql1:
  #     itemid = ordtree.item(ordtree.focus())["values"][1]
  #     print(itemid,)
  #     sql = 'SELECT FROM Customer.businessaddress WHERE customerid=%s'
  #     val = (itemid,)
  #     fbcursor.execute(sql, val)
  #     fbilldb.commit()
  #     ordtree.selection_get(ordtree.selection()[0])
  # else:
  #     pass  


  # sql    = "SELECT * FROM Customer.businessaddress WHERE Address='$businessaddress'"


  # sql = "select Customer.businessaddress from Customer where pdata = Orders.businessname"
  # fbcursor.execute(sql,)
  # addrs = fbcursor.fetchone()
  # print(addrs)

  # sql = "SELECT * from Customer where businessaddress =%s"
  # val = ("businessaddress", )
  # fbcursor.execute(sql,val)
  # pdata = fbcursor.fetchone()
  # print(pdata)
  #     entry.insert(0, pdata[3])






  address=Label(labelframe1,text="Address").place(x=10,y=30)
  addrs=scrolledtext.Text(labelframe1, undo=True,width=23)
  addrs.place(x=80,y=30,height=70)
 
  
  ship=Label(labelframe1,text="Ship to").place(x=342,y=5)
  spt=Entry(labelframe1,width=31)
  spt.place(x=402,y=3)

  address1=Label(labelframe1,text="Address").place(x=340,y=30)
  adrs1=scrolledtext.Text(labelframe1, undo=True,width=23)
  adrs1.place(x=402,y=30,height=70)

##
  def copying1():
    # shpdlt=spt1.get(osdata[17])
    spt.delete(0,END)
    spt.insert(0,slctcstr[4])
    adrs1.delete('1.0',END)
    adrs1.insert('1.0',slctcstr[5])
##
  btn1=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>",command=copying1)
  btn1.place(x=290, y=48)
  
  labelframe2 = LabelFrame(fir1Frame,text="")
  labelframe2.place(x=10,y=130,width=640,height=42)

  email=Label(labelframe2,text="Email").place(x=10,y=5)
  eml=Entry(labelframe2,width=31)
  eml.place(x=80,y=5)

  sms=Label(labelframe2,text="SMS No.").place(x=340,y=5)
  smsno=Entry(labelframe2,width=31)
  smsno.place(x=402,y=5)
    
  labelframe = LabelFrame(fir1Frame,text="Order",font=("arial",15))
  labelframe.place(x=652,y=5,width=290,height=170)
 
 ##
  fbcursor.execute("SELECT * FROM Orders ORDER BY orderid DESC LIMIT 1")
  result = fbcursor.fetchone()
 ##
  order=Label(labelframe,text="Order#").place(x=5,y=5)
  ord=Entry(labelframe,width=25)
  ord.place(x=100,y=5,)
  ord.delete(0,'end')
  if not result== None:
   frck=result[0]+1
  else:
    frck=1
  ord.insert(0, frck)
 ##
  def duecheckord1():
     if checkvarStatus5.get()==0:
       duedte['state'] = DISABLED  
     else:
       duedte['state'] = NORMAL

  orderdate=Label(labelframe,text="Order date").place(x=5,y=33)
  orddte=DateEntry(labelframe,width=25)
  orddte.place(x=100,y=33)
  checkvarStatus5=IntVar()
  duedate=Checkbutton(labelframe,variable = checkvarStatus5,text="Due date",onvalue =1 ,offvalue = 0, command=duecheckord1).place(x=5,y=62)
  duedte=DateEntry(labelframe,width=25)
  duedte.place(x=100,y=62)
  terms=Label(labelframe,text="Terms").place(x=5,y=92)
  trms=ttk.Combobox(labelframe, value="",width=25)
  trms.place(x=100,y=92)
  # ref=Label(labelframe,text="Order ref#").place(x=5,y=118)
  # ordref=Entry(labelframe,width=27).place(x=100,y=118)

  fir2Frame=Frame(pop, height=150,width=100,bg="#f5f3f2")
  fir2Frame.pack(side="top", fill=X)
  listFrame = Frame(fir2Frame, bg="white", height=140,borderwidth=5,  relief=RIDGE)


  
  tree10=ttk.Treeview(listFrame)
  tree10["columns"]=["1","2","3","4","5","6","7","8"]

  tree10.column("#0", width=40, anchor="center")
  tree10.column("1", width=80, anchor="center")
  tree10.column("2", width=190, anchor="center")
  tree10.column("3", width=190, anchor="center")
  tree10.column("4", width=80, anchor="center")
  tree10.column("5", width=60, anchor="center")
  tree10.column("6", width=60, anchor="center")
  tree10.column("7", width=60, anchor="center")
  tree10.column("8", width=80, anchor="center")

  tree10.heading("#0")
  tree10.heading("1",text="ID/SKU")
  tree10.heading("2",text="Product/Service")
  tree10.heading("3",text="Description")
  tree10.heading("4",text="Unit Price")
  tree10.heading("5",text="Quantity")
  tree10.heading("6",text="Pcs/Weight")
  tree10.heading("7",text="Tax1")
  tree10.heading("8",text="Price")

  
  pricecol1 = Label(listFrame,bg="#f5f3f2")
  pricecol1.place(x=850,y=200,width=78,height=18)
  # pricecoltotl1 = Entry(listFrame,width=28)

  # fbcursor.execute('SELECT * FROM Productservice;') 
  # j = 0
  # for i in fbcursor:
  #   viwedttree.insert(parent='', index='end', iid=i, text='', values=(' ',i[2],i[12],i[5],i[7],i[13],i[8],i[10]))
  #   j += 1

  
  tree10.pack(fill="both", expand=1)
  listFrame.pack(side="top", fill="both", padx=5, pady=3, expand=1)
  
  new_value = tk.StringVar()

  def edit_window_box(val):
    
    edit_window = Toplevel(root)
    edit_window.title("Edit the value or cancel")
    edit_window.geometry("400x200")
    label_edit = tk.Label(edit_window , text='Enter value to edit', 
    font = ("Times New Roman", 12)).place(x=68,y=60)
    #create edit box
    edit_box = tk.Entry(edit_window)
    edit_box.insert(0,val)
    edit_box.place(x=200,y=63)
    #auto select edit window 
    edit_window.focus()
    
    def value_assignment(event):
        printing = edit_box.get()
        # print(printing)
        indexcolaa=int(cola)-1
        new_value.set(printing)
        # print(col)
        #only destroy will not update the value (perhaps event keeps running in background)
        #quit allows event to stop n update value in tree but does not close the window in single click 
        #rather on dbl click shuts down entire app 

        selected0 = tree10.focus()
        valuz1= tree10.item(selected0)["values"]
        idgettingprdt=valuz1[0]
        valuz1[indexcolaa]=printing
        # print(valuz1)

        sql2z0= 'UPDATE storingproduct SET sku=%s,name=%s,description=%s,unitprice=%s,quantity=%s,peices=%s,taxable=%s WHERE orderid=%s AND Productserviceid=%s'
        val0z1=(valuz1[0],valuz1[1],valuz1[2],valuz1[3],valuz1[4],valuz1[5],valuz1[6],frck,idgettingprdt)
        # dnz=(ordrid,sql2z0,)
        # print(val0z1)
        fbcursor.execute(sql2z0,val0z1)
        fbilldb.commit()

        qntsql= 'UPDATE Productservice SET quantity=%s WHERE Productserviceid=%s'
        qntval=(valuz1[4],idgettingprdt)
        fbcursor.execute(qntsql,qntval)
        fbilldb.commit()

        # for record in tree10.get_children():
        #   tree10.delete(record)
        # tree10.delete(tree10.get_children())

        fbcursor.execute("SELECT * FROM storingproduct ORDER BY orderid DESC LIMIT 1")
        bol = fbcursor.fetchone()[0]
        # print(bol)
        # print(frck)

        if bol==frck:        
         for record in tree10.get_children():
          tree10.delete(record)
         m="SELECT * FROM storingproduct  WHERE orderid=%s"
         i=(frck,)
         fbcursor.execute(m,i)
         panrefrdata = fbcursor.fetchall()
         az=0
         countp = 0
         for i in panrefrdata:
           tree10.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))

           az+=(i[9]*i[22])
         countp += 1
         pricecol1.config(text=az)
        else:
          pass
        edit_window.quit()
        edit_window.destroy()
    
    edit_window.bind('<Return>', value_assignment )
    


    B1 = tk.Button(edit_window, text=" Okay ")
    B1.bind('<Button-1>',value_assignment)
    B1.place(x=70,y=130)
    
    
    B2 = tk.Button(edit_window, text="Cancel", command = edit_window.destroy).place(x=276,y=130)
    edit_window.mainloop()
    
  #will explain
  #variable to hold col value (col clicked)
  shape1 = tk.IntVar()
  #tracks both col , row on mouse click
  def tree_click_handler(event):
      global cola
      cur_item = tree10.item(tree10.focus())
      cola = tree10.identify_column(event.x)[1:]
      rowid = tree10.identify_row(event.y)[1:]
     
      # print(rowid)
      #updates list
      shape1.set(cola)
      try:
          x,y,w,h = tree10.bbox('I'+rowid,'#'+cola)
      except:pass
      #tree.tag_configure("highlight", background="yellow")
      return(cola)

  #code linked to event    
  tree10.bind('<ButtonRelease-1>', tree_click_handler)

  #edit a value in a clicked cell
  def edit(event):
      try:
          selected_item = tree10.selection()[0]
          temp = list(tree10.item(selected_item , 'values'))
          tree_click_handler
          col_selected = int(shape1.get())-1
          edit_window_box(temp[col_selected])
          #do not run if edit window is open
          #use edit_window.mainloop() so value assign after window closes
          temp[col_selected] = new_value.get()
          tree10.item(selected_item, values= temp)
      except: pass

   
  #binding allows to edit on screen double click
  tree10.bind('<Double-Button-1>' , edit)

  # sql = "SELECT * FROM productservice  WHERE productserviceid= %s"
  # i=(valuep)
  # fbcursor.execute(sql,i)
  
  # j = 0
  # for i in fbcursor:
  #         tree10.insert(parent='', index='end', iid=i, text='', values=(i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9]))
  # j += 1

  fir3Frame=Frame(pop,height=200,width=700,bg="#f5f3f2")
  fir3Frame.place(x=0,y=490)

  tabStyle = ttk.Style()
  tabStyle.theme_use('default')
  tabStyle.configure('TNotebook.Tab', background="#999999", width=12, padding=5)
  myNotebook=ttk.Notebook(fir3Frame)
  orderFrame = Frame(myNotebook, height=200, width=800)
  headerFrame = Frame(myNotebook, height=200, width=800)
  commentFrame = Frame(myNotebook, height=200, width=800)
  termsFrame = Frame(myNotebook, height=200, width=800)
  noteFrame = Frame(myNotebook, height=200, width=800)
  documentFrame = Frame(myNotebook, height=200, width=800)
  
  myNotebook.add(orderFrame,compound="left", text="Order")
  myNotebook.add(headerFrame,compound="left",  text="Header/Footer")
  myNotebook.add(commentFrame,compound="left",  text="Comments")
  myNotebook.add(termsFrame,compound="left", text="Terms")
  myNotebook.add(noteFrame,compound="left",  text="Private notes")
  myNotebook.add(documentFrame,compound="left",  text="Documents")
  myNotebook.pack(expand = 1, fill ="both") 

##
  sql = "select extra_cost_name from Orders"
  fbcursor.execute(sql,)
  excodata = fbcursor.fetchall()  
## 

  labelframe1 = LabelFrame(orderFrame,text="",font=("arial",15))
  labelframe1.place(x=1,y=1,width=800,height=170)
  cost1=Label(labelframe1,text="Extra cost name").place(x=2,y=5)
  xtracstnme=ttk.Combobox(labelframe1, value=excodata,width=20)
  xtracstnme.place(x=115,y=5)
  rate=Label(labelframe1,text="Discount rate").place(x=370,y=5)
  dscntrtecr=Spinbox(labelframe1,width=6,  from_=0, to=100, font="italic 10")
  dscntrtecr.place(x=460,y=5)
  cost2=Label(labelframe1,text="Extra cost").place(x=35,y=35)
  xtracstcr=Entry(labelframe1,width=10)
  xtracstcr.place(x=115,y=35)
  tax=Label(labelframe1,text="Tax1").place(x=420,y=35)
  taxxcr=Entry(labelframe1,width=7)
  taxxcr.place(x=460,y=35)
##
  sql = "select template from Orders"
  fbcursor.execute(sql,)
  tmpdata = fbcursor.fetchall()  
##
  template=Label(labelframe1,text="Template").place(x=37,y=70)
  tmplte=ttk.Combobox(labelframe1, value=tmpdata,width=25)
  tmplte.place(x=115,y=70)
  sales=Label(labelframe1,text="Sales Person").place(x=25,y=100)
  salesprsn=Entry(labelframe1,width=18)
  salesprsn.place(x=115,y=100)
  category=Label(labelframe1,text="Category").place(x=300,y=100)
  ctgry=Entry(labelframe1,width=22)
  ctgry.place(x=370,y=100)
  
  statusfrme = LabelFrame(labelframe1,text="Status",font=("arial",15))
  statusfrme.place(x=540,y=0,width=160,height=160)
  draft=Label(statusfrme, text="Draft",font=("arial", 15, "bold"), fg="grey").place(x=50, y=3)
  on1=Label(statusfrme, text="Emailed on:").place( y=50)
  nev1=Label(statusfrme, text="Never").place(x=100,y=50)
  on2=Label(statusfrme, text="Printed on:").place( y=90)
  nev2=Label(statusfrme, text="Never").place(x=100,y=90)

  text1=Label(headerFrame,text="Title text").place(x=50,y=5)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=5)
  text2=Label(headerFrame,text="Page header text").place(x=2,y=45)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=45)
  text3=Label(headerFrame,text="Footer text").place(x=35,y=85)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=85)


  def addnotepri():
    vel=addnote.get('1.0','end-1c')
    adntsql="UPDATE orders SET private_notes=%s WHERE orderid=%s"
    adntsval=(vel,frck)
    fbcursor.execute(adntsql,adntsval)
    fbilldb.commit()
  text=Label(noteFrame,text="Private notes (not shown on invoice/order/estimates)").place(x=10,y=10)
  addnote=scrolledtext.Text(noteFrame,width=90, undo=True,height=7)
  addnote.place(x=10,y=32)
  addnotesave=Button(noteFrame,height=1,width=8,text="Save Note",command=addnotepri)
  addnotesave.place(x=600,y=5)

  e1=Text(termsFrame,width=100,height=9).place(x=10,y=10)

  e1=Text(commentFrame,width=100,height=9).place(x=10,y=10)
  def ordupload_file1():
        global filename
        import shutil
        f_types =[('Png files','.png'),('Jpg Files', '.jpg'),('All Files', '*.*')]
        
        filename = filedialog.askopenfilename(filetypes=f_types)
        shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
        # sql = 'INSERT INTO documents (add_document) VALUES (%s) WHERE orderid=%s'
        sql= 'INSERT INTO documents(add_document,orderid) VALUES(%s,%s)'
        val = (filename.split('/')[-1],frck)
        fbcursor.execute(sql,val)
        fbilldb.commit()
        sqldoc='SELECT * FROM documents Where orderid=%s'
        vall=(frck,)
        fbcursor.execute(sqldoc,vall)
        for record in cusventtreead.get_children():
              cusventtreead.delete(record) 
        j = 0
        for i in fbcursor:
         cusventtreead.insert(parent='', index='end', iid=i, text='', values=(i[2]))
        j += 1
   

  def rmvattad():
    itemidcs = cusventtreead.item(cusventtreead.focus())["values"][0]
    # print(itemid,)
    # Nullified="NULL"
    sql91 = "DELETE FROM documents WHERE add_document=%s AND orderid=%s"
    # val91 = (Nullified,frck,)
    val91=(itemidcs,frck,)
    fbcursor.execute(sql91, val91)
    fbilldb.commit()
    cusventtreead.delete(cusventtreead.selection()[0])
  btn1=Button(documentFrame,height=2,width=3,text="+",command=ordupload_file1).place(x=5,y=10)
  btn2=Button(documentFrame,height=2,width=3,text="-",command=rmvattad).place(x=5,y=50)
  # text=Label(documentFrame,text="Attached documents or image files.If you attach large email then email taken long time to send").place(x=50,y=10)
  def addpsfile_image1(event):
      edit_window_img1 = Toplevel()
      edit_window_img1.title("View Image")
      edit_window_img1.geometry("700x500")
      
      
      itemid = cusventtreead.item(cusventtreead.focus())["values"][0]
      sql = "select * from documents where orderid = %s"
      val = (frck, )

      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone() 
      image = Image.open("images/"+psdata[2])
      resize_image = image.resize((700, 500))
      image = ImageTk.PhotoImage(resize_image)
      psimage = Label(edit_window_img1,image=image)
      psimage.photo = image
      psimage.pack()

  cusventtreead=ttk.Treeview(documentFrame, height=5)
  cusventtreead["columns"]=["1"]
  cusventtreead.column("#0", width=20, anchor="center")
  cusventtreead.column("1", width=650, anchor="center")
  # cusventtreead.column("2", width=250, anchor="center")
  # cusventtreead.column("2", width=200, anchor="center")
  cusventtreead.heading("#0",text="", anchor=W)
  # cusventtreead.heading("1",text="Attach to Email")
  cusventtreead.heading("1",text="Filename")
  # cusventtreeeddc.heading("3",text="Filesize")  
  cusventtreead.place(x=50, y=20)
  cusventtreead.bind('<Double-Button-1>' ,addpsfile_image1)

  

  fir4Frame=Frame(pop,height=190,width=210,bg="#f5f3f2")
  fir4Frame.place(x=740,y=520)
  summaryfrme = LabelFrame(fir4Frame,text="Summary",font=("arial",15))
  summaryfrme.place(x=0,y=0,width=200,height=140)
  discount=Label(summaryfrme, text="Discount").place(x=0 ,y=0)
  discount1=Label(summaryfrme, text="0.00").place(x=130 ,y=0)
  sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
  sub001=Label(summaryfrme, text="0.00").place(x=130 ,y=21)
  tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
  tax001=Label(summaryfrme, text="0.00").place(x=130 ,y=42)
  cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
  cost001=Label(summaryfrme, text="0.00").place(x=130 ,y=63)
  order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
  order001=Label(summaryfrme, text="0.00").place(x=130 ,y=84)
  # total=Label(summaryfrme, text="Total paid").place(x=0 ,y=105)
  # total1=Label(summaryfrme, text="$0.00").place(x=130 ,y=105)
  # balance=Label(summaryfrme, text="Balance").place(x=0 ,y=126)
  # balance1=Label(summaryfrme, text="$0.00").place(x=130 ,y=126)

  fir5Frame=Frame(pop,height=38,width=210)
  fir5Frame.place(x=735,y=485)

  def recalcultncr():
   global ot1v
  #  global discount1,sub1,tax1,cost1z,order1
  #  totalpriceinput.config(text="")
   ttlprzinptcr=pricecol1.cget("text")
   ratediscntcr=dscntrtecr.get()
   xtaxcr=taxxcr.get()
   cstxtracr=xtracstcr.get()
   print(ttlprzinptcr,ratediscntcr,xtaxcr,cstxtracr)


   sdd=round(int(ratediscntcr),3)
   pldd=int(ttlprzinptcr)
   p=round(pldd, 4)
   dsctedd=(sdd*pldd)/100
   dscdd=round(dsctedd, 4)
   discount=Label(summaryfrme, text=str(sdd)+"% Discount").place(x=0 ,y=0)
   discount1cr=Label(summaryfrme,text='Rs. '+str(dscdd))
   discount1cr.place(x=130 ,y=0)

   sbtotlhee=p-dscdd
   subtotl=round(sbtotlhee, 4)
   sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
   sub100cr=Label(summaryfrme, text='Rs. '+str(subtotl))
   sub100cr.place(x=130 ,y=21)

   taxval91=int(xtaxcr)
   tax1dsplay=(taxval91*subtotl)/100
   taxvalu=round(tax1dsplay,2)
   tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
   tax100cr=Label(summaryfrme, text='Rs. '+str(taxvalu))
   tax100cr.place(x=130 ,y=42)

   xtracstvaluzz=int(cstxtracr)
   xtra1cst=round(xtracstvaluzz, 4)
   cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
   cost100cr=Label(summaryfrme, text='Rs. '+str(xtra1cst))
   cost100cr.place(x=130 ,y=63)
 
   otherttlval=subtotl+taxvalu+xtra1cst
   ot1v=round(otherttlval, 2)
   order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
   order100cr=Label(summaryfrme, text='Rs. '+str(ot1v))
   order100cr.place(x=130 ,y=84)

   orddateins=orddte.get_date()
   orddueins=duedte.get_date()
   trmsin=trms.get()
   extracstnme=xtracstnme.get()
   discountrte=dscntrtecr.get()
   extracst=xtracstcr.get()
   taax=taxxcr.get()
   tplts=tmplte.get()
   slzprzn=salesprsn.get()
   ctgryy=ctgry.get()
   itemdot1=frck
   ordttl=ot1v
   sumdisc=dscdd
   sumtax=taxvalu
   sumsub=subtotl
   dsrtedit=sdd
  #  for record in ordtree.get_children():
  #     ordtree.delete(record)
   sqlordttl = "UPDATE Orders SET extra_cost_name=%s,extra_cost=%s,template=%s,discount_rate=%s,tax1=%s,category=%s,sales_person=%s,order_date=%s,due_date=%s,discount_rate=%s,sum_discount=%s,sum_tax=%s,sum_subtotal=%s,Order_total=%s WHERE orderid = %s"
   valqordttl = (extracstnme,extracst,tplts,discountrte,taax,ctgryy,slzprzn,orddateins,orddueins,dsrtedit,sumdisc,sumtax,ot1v,sumsub,itemdot1, )
   fbcursor.execute(sqlordttl, valqordttl,)
   fbilldb.commit()

  #  fbcursor.execute("select *  from Orders")
  #  j = 0
  #  for i in fbcursor:
  #    ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8]))
    
  #    ordertotalinput=0
  #    for line in ordtree.get_children():
  #     idsave1=ordtree.item(line)['values'][9]
  #     ordertotalinput += idsave1
  #  j += 1
  #  ordtotalrowcol.config(text=ordertotalinput)


  recalbtn1=Button(fir5Frame, compound="left", text="Show Summary",command=recalcultncr)
  recalbtn1.place(x=75, y=0,height=35)
  # btnup=Button(fir5Frame, compound="left", text="Line Up").place(x=150, y=0)




#printselected order
  
# def printsele():
  # subprocess.Popen('C:\\Windows\\System32\\spoolsv.exe')
    
    # # Ask for file (Which you want to print)
    # file_to_print = filedialog.askopenfilename(
    #   initialdir="/", title="Select file", 
    #   filetypes=(("Text files", "*.txt"), ("all files", "*.*")))
      
    # if file_to_print:
        
    #     # Print Hard Copy of File
    #     win32api.ShellExecute(0, "print", file_to_print, None, ".", 0)
    #     Button(root, text="Print FIle", command=printsele).pack()

  # def property1():
  #   propert=Toplevel()
  #   propert.title("Microsoft Print To PDF Advanced Document Settings")
  #   propert.geometry("670x500+240+150")

  #   def property2():
  #     propert1=Toplevel()
  #     propert1.title("Microsoft Print To PDF Advanced Document Settings")
  #     propert1.geometry("670x500+240+150")

  #     name=Label(propert1, text="Microsoft Print To PDF Advanced Document Settings").place(x=10, y=5)
  #     paper=Label(propert1, text="Paper/Output").place(x=30, y=35)
  #     size=Label(propert1, text="Paper size").place(x=55, y=65)
  #     n = StringVar()
  #     search = ttk.Combobox(propert1, width = 15, textvariable = n )
  #     search['values'] = ('letter')
  #     search.place(x=150,y=65)
  #     search.current(0)
  #     copy=Label(propert1, text="Copy count:").place(x=55, y=95)

  #     okbtn=Button(propert1,compound = LEFT,image=tick , text="Ok", width=60).place(x=460, y=450)
  #     canbtn=Button(propert1,compound = LEFT,image=cancel, text="Cancel", width=60).place(x=570, y=450)
      
      


  #   style = ttk.Style()
  #   style.theme_use('default')
  #   style.configure('TNotebook.Tab', background="#999999", padding=5)
  #   property_Notebook = ttk.Notebook(propert)
  #   property_Frame = Frame(property_Notebook, height=500, width=670)
  #   property_Notebook.add(property_Frame, text="Layout")
  #   property_Notebook.place(x=0, y=0)

  #   name=Label(property_Frame, text="Orientation:").place(x=10, y=5)
  #   n = StringVar()
  #   search = ttk.Combobox(property_Frame, width = 23, textvariable = n )
  #   search['values'] = ('Portrait')
  #   search.place(x=10,y=25)
  #   search.current(0)

  #   text=Text(property_Frame,width=50).place(x=250, y=5,height=350)

  #   btn=Button(property_Frame, text="Advanced",command=property2).place(x=550, y=380)
  #   btn=Button(property_Frame,compound = LEFT,image=tick  ,text="OK", width=60,).place(x=430, y=420)
  #   btn=Button(property_Frame,compound = LEFT,image=cancel , text="Cancel", width=60,).place(x=550, y=420)     


    
  # if(False):
  #     messagebox.showwarning("FBilling Revelution 2020", "Customer is required, Please select customer for this invoice\nbefore printing")
  # elif(False):
  #     messagebox.showinfo("FBilling Revelution 2020", "Print job has been completed.")
  # else:
  #     print1=Toplevel()
  #     print1.title("Print")
  #     print1.geometry("670x400+240+150")
      
  #     printerframe=LabelFrame(print1, text="Printer", height=80, width=650)
  #     printerframe.place(x=7, y=5)      
  #     name=Label(printerframe, text="Name:").place(x=10, y=5)
  #     e1= ttk.Combobox(printerframe, width=40).place(x=70, y=5)
  #     where=Label(printerframe, text="Where:").place(x=10, y=30)
  #     printocheckvar=IntVar()
  #     printochkbtn=Checkbutton(printerframe,text="Print to file",variable=printocheckvar,onvalue=1,offvalue=0,height=1,width=10)
  #     printochkbtn.place(x=450, y=30)
  #     btn=Button(printerframe, text="Properties", width=10,command=property1).place(x=540, y=5)

  #     pageslblframe=LabelFrame(print1, text="Pages", height=140, width=320)
  #     pageslblframe.place(x=10, y=90)
  #     radvar=IntVar()
  #     radioall=Radiobutton(pageslblframe, text="All", variable=radvar, value="1").place(x=10, y=5)
  #     radiocpage=Radiobutton(pageslblframe, text="Current Page", variable=radvar, value="2").place(x=10, y=25)
  #     radiopages=Radiobutton(pageslblframe, text="Pages: ", variable=radvar, value="3").place(x=10, y=45)
  #     pagecountentry = Entry(pageslblframe, width=23).place(x=80, y=47)
  #     pageinfolabl=Label(pageslblframe, text="Enter page numbers and/or page ranges\nseperated by commas. For example:1,3,5-12")
  #     pageinfolabl.place(x=5, y=75)

  #     copylblframe=LabelFrame(print1, text="Copies", height=140, width=320)
  #     copylblframe.place(x=335, y=90)
  #     nolabl=Label(copylblframe, text="Number of copies").place(x=5, y=5)      
  #     noentry = Entry(copylblframe, width=18).place(x=130, y=5)      
  #     one=Frame(copylblframe, width=30, height=40, bg="black").place(x=20, y=40)     
  #     two=Frame(copylblframe, width=30, height=40, bg="grey").place(x=15, y=45)     
  #     three=Frame(copylblframe, width=30, height=40, bg="white").place(x=10, y=50)      
  #     four=Frame(copylblframe, width=30, height=40, bg="black").place(x=80, y=40)      
  #     fiv=Frame(copylblframe, width=30, height=40, bg="grey").place(x=75, y=45)      
  #     six=Frame(copylblframe, width=30, height=40, bg="white").place(x=70, y=50)      
  #     collatecheckvar=IntVar()
  #     collatechkbtn=Checkbutton(copylblframe,text="Collate",variable=collatecheckvar,onvalue=1,offvalue=0,height=1,width=10)
  #     collatechkbtn.place(x=130, y=70)

  #     othrlblframe=LabelFrame(print1, text="Other", height=120, width=320)
  #     othrlblframe.place(x=10, y=235)
  #     printlb=Label(othrlblframe, text="Print").place(x=5, y=0)
  #     dropprint = ttk.Combobox(othrlblframe, width=23).place(x=80, y=0)
  #     orderlb=Label(othrlblframe, text="Order").place(x=5, y=25)
  #     dropord = ttk.Combobox(othrlblframe, width=23).place(x=80, y=25)
  #     duplexlb=Label(othrlblframe, text="Duplex").place(x=5, y=50)
  #     droplex = ttk.Combobox(othrlblframe, width=23).place(x=80, y=50)

  #     prmodelblframe=LabelFrame(print1, text="Print mode", height=120, width=320)
  #     prmodelblframe.place(x=335, y=235)
  #     dropscal = ttk.Combobox(prmodelblframe, width=30).place(x=5, y=5)
  #     poslb=Label(prmodelblframe, text="Print on sheet").place(x=5, y=35)
  #     droppos = ttk.Combobox(prmodelblframe, width=10).place(x=155, y=35)

  #     okbtn=Button(print1,compound = LEFT,image=tick , text="Ok", width=60).place(x=460, y=370)
  #     canbtn=Button(print1,compound = LEFT,image=cancel, text="Cancel", width=60).place(x=570, y=370)
      


#email
      
# def email_order():
#   mailDetail=Toplevel()
#   mailDetail.title("Invoice E-Mail")
#   mailDetail.geometry("1080x550")
#   mailDetail.resizable(False, False)
#   def my_SMTP():
#       if True:
#           em_ser_conbtn.destroy()
#           mysmtpservercon=LabelFrame(account_Frame,text="SMTP server connection(ask your ISP for your SMTP settings)", height=165, width=380)
#           mysmtpservercon.place(x=610, y=110)
#           lbl_hostn=Label(mysmtpservercon, text="Hostname").place(x=5, y=10)
#           hostnent=Entry(mysmtpservercon, width=30).place(x=80, y=10)
#           lbl_portn=Label(mysmtpservercon, text="Port").place(x=5, y=35)
#           portent=Entry(mysmtpservercon, width=30).place(x=80, y=35)
#           lbl_usn=Label(mysmtpservercon, text="Username").place(x=5, y=60)
#           unament=Entry(mysmtpservercon, width=30).place(x=80, y=60)
#           lbl_pasn=Label(mysmtpservercon, text="Password").place(x=5, y=85)
#           pwdent=Entry(mysmtpservercon, width=30).place(x=80, y=85)
#           ssl_chkvar=IntVar()
#           ssl_chkbtn=Checkbutton(mysmtpservercon, variable=ssl_chkvar, text="This server requires a secure connection(SSL)", onvalue=1, offvalue=0)
#           ssl_chkbtn.place(x=50, y=110)
#           em_ser_conbtn1=Button(account_Frame, text="Test E-mail Server Connection").place(x=610, y=285)
#       else:
#           pass

 
#   style = ttk.Style()
#   style.theme_use('default')
#   style.configure('TNotebook.Tab', background="#999999", padding=5)
#   email_Notebook = ttk.Notebook(mailDetail)
#   email_Frame = Frame(email_Notebook, height=500, width=1080)
#   account_Frame = Frame(email_Notebook, height=550, width=1080)
#   email_Notebook.add(email_Frame, text="E-mail")
#   email_Notebook.add(account_Frame, text="Account")
#   email_Notebook.place(x=0, y=0)

#   messagelbframe=LabelFrame(email_Frame,text="Message", height=500, width=730)
#   messagelbframe.place(x=5, y=5)
#   global email_address, email_subject, email_from
#   email_address = StringVar() 
#   email_subject = StringVar()
#   # email_body = StringVar()
#   email_from = StringVar()

#   lbl_emailtoaddr=Label(messagelbframe, text="Email to address").place(x=5, y=5)
#   emailtoent=Entry(messagelbframe, width=50,  textvariable=email_address)
#   emailtoent.place(x=120, y=5)
#   #email2 = email_address.get()
#   #print(email2)
#   sendemail_btn=Button(messagelbframe, text="Send Email", width=10, height=1, command=send_mail).place(x=600, y=10)#,command=addacnt

#   lbl_carcopyto=Label(messagelbframe, text="Carbon copy to").place(x=5, y=32)
#   carcopyent=Entry(messagelbframe, width=50).place(x=120, y=32)
#   stopemail_btn=Button(messagelbframe, text="Stop sending", width=10, height=1).place(x=600, y=40)
#   lbl_subject=Label(messagelbframe, text="Subject").place(x=5, y=59)
#   subent=Entry(messagelbframe, width=50, textvariable=email_subject)
#   subent.place(x=120, y=59)

  
#   style = ttk.Style()
#   style.theme_use('default')
#   style.configure('TNotebook.Tab', background="#999999", width=20, padding=5)
#   mess_Notebook = ttk.Notebook(messagelbframe)
#   emailmessage_Frame =Frame(mess_Notebook, height=350, width=710)
#   htmlsourse_Frame = Frame(mess_Notebook, height=350, width=710)
#   mess_Notebook.add(emailmessage_Frame, text="E-mail message")
#   mess_Notebook.add(htmlsourse_Frame, text="Html source code")
#   mess_Notebook.place(x=5, y=90)

#   btn1=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

  
#   btn2=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
#   btn3=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
#   btn4=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
#   btn5=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=undo).place(x=140, y=1)
#   btn6=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=redo).place(x=175, y=1)
#   btn7=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=bold).place(x=210, y=1)
#   btn8=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=italics).place(x=245, y=1)
#   btn9=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=underline).place(x=280, y=1)
#   btn10=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=left).place(x=315, y=1)
#   btn11=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=right).place(x=350, y=1)
#   btn12=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=center).place(x=385, y=1)
#   btn13=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=hyperlink).place(x=420, y=1)
  
#   btn14=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=remove).place(x=455, y=1)


#   dropcomp = ttk.Combobox(emailmessage_Frame, width=12, height=3).place(x=500, y=5)
#   dropcompo = ttk.Combobox(emailmessage_Frame, width=6, height=3).place(x=600, y=5)
#   mframe=Frame(emailmessage_Frame, height=350, width=710, bg="white")
#   mframe.place(x=0, y=28)



#   btn1=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

  
#   btn2=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
#   btn3=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
#   btn4=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
#   mframe=Frame(htmlsourse_Frame, height=350, width=710, bg="white")
#   mframe.place(x=0, y=28)
#   attachlbframe=LabelFrame(email_Frame,text="Attachment(s)", height=350, width=280)
#   attachlbframe.place(x=740, y=5)
#   htcodeframe=Frame(attachlbframe, height=220, width=265, bg="white")
#   htcodeframe.place(x=5, y=5)

#   lbl_btn_info=Label(attachlbframe, text="Double click on attachment to view").place(x=30, y=230)
#   btn17=Button(attachlbframe, width=20, text="Add attachment file...", )
#   btn17.place(x=60, y=260)
#   btn18=Button(attachlbframe, width=20, text="Remove attachment").place(x=60, y=295)
#   lbl_tt_info=Label(email_Frame, text="You can create predefined invoice, order, estimate\nand payment receipt email templates under Main\nmenu/Settings/E-Mail templates tab")
#   lbl_tt_info.place(x=740, y=370)

#   ready_frame=Frame(mailDetail, height=20, width=1080, bg="#b3b3b3").place(x=0,y=530)
  
#   sendatalbframe=LabelFrame(account_Frame,text="E-Mail(Sender data)",height=270, width=600)
#   sendatalbframe.place(x=5, y=5)
#   lbl_sendermail=Label(sendatalbframe, text="Your company email address").place(x=5, y=30)
#   sentent=Entry(sendatalbframe, width=40, textvariable=email_from)
#   sentent.place(x=195, y=30)
#   lbl_orcompanyname=Label(sendatalbframe, text="Your name or company name").place(x=5, y=60)
#   nament=Entry(sendatalbframe, width=40).place(x=195, y=60)
#   lbl_reply=Label(sendatalbframe, text="Reply to email address").place(x=5, y=90)
#   replyent=Entry(sendatalbframe, width=40).place(x=195, y=90)
#   lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=120)
#   signent=Entry(sendatalbframe,width=50).place(x=100, y=120,height=75)
#   confirm_chkvar=IntVar()
#   confirm_chkbtn=Checkbutton(sendatalbframe, variable=confirm_chkvar, text="Confirmation reading", onvalue=1, offvalue=0)
#   confirm_chkbtn.place(x=200, y=215)
#   btn18=Button(account_Frame, width=15, text="Save settings",command=savesettings).place(x=25, y=285)

#   sendatalbframe=LabelFrame(account_Frame,text="SMTP Server",height=100, width=380)
#   sendatalbframe.place(x=610, y=5)
#   servar=IntVar()
#   SMTP_rbtn=Radiobutton(sendatalbframe, text="Use the Built-In SMTP Server Settings", variable=servar, value=1)
#   SMTP_rbtn.place(x=10, y=10)
#   MySMTP_rbtn=Radiobutton(sendatalbframe, text="Use My Own SMTP Server Settings(Recommended)", variable=servar, value=2, command=my_SMTP)
#   MySMTP_rbtn.place(x=10, y=40)
#   em_ser_conbtn=Button(account_Frame, text="Test E-mail Server Connection")
#   em_ser_conbtn.place(x=710, y=110)
#   def send_mail():

#     sender_email =  email_from.get()     #"myownworkrelated@gmail.com"   
#     sender_password = "Myownworkrelated1234"

#     server = smtplib.SMTP('smtp.gmail.com', 587) ##########################################################################################
#     print("Login successful1")
#     server.starttls()
#     print("Login successful2")
#     server.login(sender_email, sender_password)
#     print("Login successful3")
    
#     # global email_address, email_subject

#     # email_address = StringVar()  
#     # email_subject = StringVar()
    
#     address_info = email_address.get()
#     print(address_info)
#     subject_info = email_subject.get()
#     # message_body = email_body.get()

#     server.sendmail(sender_email, address_info, subject_info)

#     print("message sent")
    
#     # subent.delete(0, END)
#     # emailtoent.delete(0, END)


#sms notification order
  
def sms():
  send_SMS=Toplevel()
  send_SMS.geometry("700x480+240+150")
  send_SMS.title("Send SMS notification")

  style = ttk.Style()
  style.theme_use('default')
  style.configure('TNotebook.Tab', background="#999999", padding=5)
  sms_Notebook = ttk.Notebook(send_SMS)
  SMS_Notification = Frame(sms_Notebook, height=470, width=700)
  SMS_Service_Account = Frame(sms_Notebook, height=470, width=700)
  sms_Notebook.add(SMS_Notification, text="SMS Notification")
  sms_Notebook.add(SMS_Service_Account, text="SMS Service Account")
  sms_Notebook.place(x=0, y=0)

  numlbel=Label(SMS_Notification, text="SMS number or comma seperated SMS number list(Please start each SMS number with the country code)")
  numlbel.place(x=10, y=10)
  numentry=Entry(SMS_Notification, width=92).place(x=10, y=30)
  stexbel=Label(SMS_Notification, text="SMS Text").place(x=10, y=60)
  stex=Entry(SMS_Notification, width=40).place(x=10, y=85,height=120)
  
  dclbel=Label(SMS_Notification, text="Double click to insert into text")
  dclbel.place(x=410, y=60)
  dcl=Entry(SMS_Notification, width=30)
  dcl.place(x=400, y=85,height=200)
  
  smstype=LabelFrame(SMS_Notification, text="SMS message type", width=377, height=60)
  smstype.place(x=10, y=223)
  snuvar=IntVar()
  normal_rbtn=Radiobutton(smstype, text="Normal SMS(160 chars)", variable=snuvar, value=1)
  normal_rbtn.place(x=5, y=5)
  unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)", variable=snuvar, value=2)
  unicode_rbtn.place(x=190, y=5)
  tiplbf=LabelFrame(SMS_Notification, text="Tips", width=680, height=120)
  tiplbf.place(x=10, y=290)
  tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="Always start the SMS nymber with the country code. Do not use the + sign at the beginning(example\nUS number:8455807546). Do not use any special characters in your normal SMS text. Please use the\nstndard SMS characters or the English alphabet and numbers only. Otherwise the SMS will be\nunreadable or undeliverable. If you need to enter international characters, accents,email address, or\nspecial characters to the SMS text field then choose the Unicode SMS format.")
  tiplabl.place(x=5, y=5)

  btn1=Button(SMS_Notification, width=20, text="Send SMS notification").place(x=10, y=420)
  btn2=Button(SMS_Notification, width=25, text="Confirm SMS cost before sending").place(x=280, y=420)
  btn3=Button(SMS_Notification, width=15, text="Cancel").place(x=550, y=420)
  

  smstype=LabelFrame(SMS_Service_Account, text="Select the notification service provider", width=670, height=65)
  smstype.place(x=10, y=5)
  snumvar=IntVar()
  normal_rbtn=Radiobutton(smstype,text="BULKSMS(www.bulksms.com)",variable=snumvar,value=1,)
  normal_rbtn.place(x=5, y=5)
  unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)-Recommended", variable=snumvar, value=2)
  unicode_rbtn.place(x=290, y=5)

  sms1type=LabelFrame(SMS_Service_Account, text="Your BULKSMS.COM Account", width=670, height=100)
  sms1type.place(x=10, y=80)
  name=Label(sms1type, text="Username").place(x=10, y=5)
  na=Entry(sms1type, width=20).place(x=100, y=5)
  password=Label(sms1type, text="Password").place(x=10, y=45)
  pas=Entry(sms1type, width=20).place(x=100, y=45)
  combo=Label(sms1type, text="Route").place(x=400, y=5)
  n = StringVar()
  combo1 = ttk.Combobox(sms1type, width = 20, textvariable = n ).place(x=450,y=5)
  btn1=Button(sms1type, width=10, text="Save settings").place(x=550, y=45)

  
  tiplbf=LabelFrame(SMS_Service_Account, text="Terms of service", width=680, height=250)
  tiplbf.place(x=10, y=190)
  tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="The SMS notification service is not free.This service costs you creadit.You must have your own account\nat BULKSMS.COM and you need to have sufficient creadit and an active internet connection to use\nthis feature.Please review all fields in this form for accuracy")
  tiplabl.place(x=0, y=5)
  tiplabl1=Label(tiplbf,justify=LEFT,fg="black",  text="visit www.bulksms.com website to create your own account.please make sure the BULKSMS .COM\n service works well in your country before you busy creadit")
  tiplabl1.place(x=0, y=60)
  tiplabl2=Label(tiplbf,justify=LEFT,fg="black",  text="Our SMS notification tool comes without any warranty.our software only forwards your SMS message\nthe BULKSMS API server .The BULKSMS API server will try to sent SMS message your recipient")
  tiplabl2.place(x=0, y=100)
  tiplabl3=Label(tiplbf,justify=LEFT,fg="red",  text="Please note that you access and use the SMS notification tool your own risk.F-Billing software is not\nresponsible for any type of loss or damage or undelivered SMS massage which you may as a result\nof accessing and using the SMS notification service.")
  tiplabl3.place(x=0, y=140)
  checkvar1=IntVar()
  chkbtn1=Checkbutton(tiplbf,text="I have read and agree to the terms of service above",variable=checkvar1,onvalue=1,offvalue=0).place(x=70, y=200)  



#print preview order
def printpreview():
  messagebox.showerror("F-Billing Revolution","Customer is required,please select customer for this order before printing.")





############## View/Edit Orders #############

def edit():
  try:
    itemid = ordtree.item(ordtree.focus())["values"][1]
  # sql = "select * from Orders where orderid = %s"
  # val = (itemid, )
  # fbcursor.execute(sql, val)
  # osdata = fbcursor.fetchone()

  # # itemid = ordtree.item(ordtree.focus())["values"][1]
  # sql = "select * from Orders where orderid = %s"
  # # val = (itemid, )
  # fbcursor.execute(sql,)
  # psdata = fbcursor.fetchone()
  # print(psdata)

    

        
    pop=Toplevel(midFrame)
    pop.title("Orders")
    pop.geometry("950x690+150+0")

    


    #select customer
    def custom():
      global cuselection
      cuselection=Toplevel()
      cuselection.title("Select Customer")
      cuselection.geometry("930x650+240+10")
      cuselection.resizable(False, False)


      #add new customer
      def create1():
        global checkvar1,checkvar2,custid,bname,baddress,cat,sname,saddress,contp,cemail,ctel,cfax,cmob,scontp,scemail,sctel,scfax,ccountry,ccity,cnotes
        ven=Toplevel(midFrame)
        ven.title("Add new vendor")
        ven.geometry("930x650+240+10")
        checkvar1=IntVar()
        checkvar2=IntVar()
        radio=IntVar()
        createFrame=Frame(ven, bg="#f5f3f2", height=650)
        createFrame.pack(side="top", fill="both")
        labelframe1 = LabelFrame(createFrame,text="Customer",bg="#f5f3f2",font=("arial",15))
        labelframe1.place(x=10,y=5,width=910,height=600)
        text1=Label(labelframe1, text="Customer ID:",bg="#f5f3f2",fg="blue").place(x=5 ,y=10)
        custid=Entry(labelframe1,width=25)
        custid.place(x=150,y=10)
        text2=Label(labelframe1, text="Category:",bg="#f5f3f2").place(x=390 ,y=10)
        cat=ttk.Combobox(labelframe1,width=25,value="Default")
        cat.place(x=460 ,y=10)
        text3=Label(labelframe1, text="Status:",bg="#f5f3f2").place(x=710 ,y=10)
        Checkbutton(labelframe1,text="Active",variable=checkvar1,onvalue=1,offvalue=0,bg="#f5f3f2").place(x=760 ,y=10)
        
        labelframe2 = LabelFrame(labelframe1,text="Invoice to (appears on invoices)",bg="#f5f3f2")
        labelframe2.place(x=5,y=40,width=420,height=150)
        name = Label(labelframe2, text="Ship to name:",bg="#f5f3f2",fg="blue").place(x=5,y=5)
        sname = Entry(labelframe2,width=28)
        sname.place(x=130,y=5)
        addr = Label(labelframe2, text="Address:",bg="#f5f3f2",fg="blue").place(x=5,y=40)
        saddress = Entry(labelframe2,width=28)
        saddress.place(x=130,y=40,height=80)
        
        btn1=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>").place(x=440, y=90)

        labelframe3 = LabelFrame(labelframe1,text="Ship to (appears on invoices)",bg="#f5f3f2")
        labelframe3.place(x=480,y=40,width=420,height=150)
        name = Label(labelframe3, text="Business name:",bg="#f5f3f2").place(x=5,y=5)
        bname = Entry(labelframe3,width=28)
        bname.place(x=130,y=5)
        addr = Label(labelframe3, text="Address:",bg="#f5f3f2").place(x=5,y=40)
        baddress = Entry(labelframe3,width=28)
        baddress.place(x=130,y=40,height=80)
        
        labelframe4 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
        labelframe4.place(x=5,y=195,width=420,height=150)
        name = Label(labelframe4, text="Contact person:",bg="#f5f3f2").place(x=5,y=5)
        contp = Entry(labelframe4,width=28)
        contp.place(x=130,y=5)
        email = Label(labelframe4, text="E-mail address:",bg="#f5f3f2",fg="blue").place(x=5,y=35)
        cemail = Entry(labelframe4,width=28)
        cemail.place(x=130,y=35)
        tel = Label(labelframe4, text="Tel.number:",bg="#f5f3f2").place(x=5,y=65)
        ctel = Entry(labelframe4,width=11)
        ctel.place(x=130,y=65)
        fax = Label(labelframe4, text="Fax:",bg="#f5f3f2").place(x=240,y=65)
        cfax = Entry(labelframe4,width=11)
        cfax.place(x=280,y=65)
        sms = Label(labelframe4, text="Mobile number for SMS notifications:",bg="#f5f3f2").place(x=5,y=95)
        cmob = Entry(labelframe4,width=15)
        cmob.place(x=248,y=95)      

        btn1=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>").place(x=440, y=250)

        
        labelframe5 = LabelFrame(labelframe1,text="Ship to contact",bg="#f5f3f2")
        labelframe5.place(x=480,y=195,width=420,height=125)
        name = Label(labelframe5, text="Contact person:",bg="#f5f3f2").place(x=5,y=5)
        scontp = Entry(labelframe5,width=28)
        scontp.place(x=130,y=5)
        email = Label(labelframe5, text="E-mail address:",bg="#f5f3f2").place(x=5,y=35)
        scemail = Entry(labelframe5,width=28)
        scemail.place(x=130,y=35)
        tel = Label(labelframe5, text="Tel.number:",bg="#f5f3f2").place(x=5,y=65)
        sctel = Entry(labelframe5,width=11)
        sctel.place(x=130,y=65)
        fax = Label(labelframe5, text="Fax:",bg="#f5f3f2").place(x=240,y=65)
        scfax = Entry(labelframe5,width=11)
        scfax.place(x=280,y=65)

        labelframe6 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
        labelframe6.place(x=5,y=350,width=420,height=100)
        Checkbutton(labelframe6,text="Tax Exempt",variable=checkvar2,onvalue=1,offvalue=0,bg="#f5f3f2").place(x=5 ,y=5)
        tax = Label(labelframe6, text="Specific Tax1 %:",bg="#f5f3f2").place(x=180,y=5)
        e1 = Entry(labelframe6,width=10).place(x=290,y=5)
        discount = Label(labelframe6, text="Discount%:",bg="#f5f3f2").place(x=5,y=35)
        e2 = Entry(labelframe6,width=10).place(x=100,y=35)

        labelframe7 = LabelFrame(labelframe1,text="Contact",bg="#f5f3f2")
        labelframe7.place(x=480,y=330,width=420,height=100)
        country = Label(labelframe7, text="country:",bg="#f5f3f2").place(x=5,y=5)
        ccountry = Entry(labelframe7,width=28)
        ccountry.place(x=130,y=5)
        city = Label(labelframe7, text="City:",bg="#f5f3f2").place(x=5,y=35)
        ccity = Entry(labelframe7,width=28)
        ccity.place(x=130,y=35)

        labelframe8 = LabelFrame(labelframe1,text="Customer Type",bg="#f5f3f2")
        labelframe8.place(x=5,y=460,width=420,height=100)
        R1=Radiobutton(labelframe8,text=" Client ",variable=radio,value=1,bg="#f5f3f2").place(x=5,y=15)
        R2=Radiobutton(labelframe8,text=" Vendor ",variable=radio,value=2,bg="#f5f3f2").place(x=150,y=15)
        R3=Radiobutton(labelframe8,text=" Both(client/vendor)",variable=radio,value=3,bg="#f5f3f2").place(x=250,y=15)
        

        labelframe9 = LabelFrame(labelframe1,text="Notes",bg="#f5f3f2")
        labelframe9.place(x=480,y=430,width=420,height=150)
        cnotes = Entry(labelframe9)
        cnotes.place(x=10,y=5,height=100,width=390)

        btn1=Button(ven,width=60,height=10,bg="#f5f3f2",compound = LEFT,image=tick ,text="OK",command=ord_cust_reg).place(x=20, y=615)
        btn2=Button(ven,width=60,height=10,bg="#f5f3f2",compound = LEFT,image=cancel,text="Cancel").place(x=800, y=615)
          
                


      # text=Label(cuselection, text="Filtered column").place(x=340, y=10)
      # e2=Entry(cuselection, width=20).place(x=450, y=10)

      def cancelcuselection1():
        cuselection.destroy()

      global cusventtree1
      cusventtree1=ttk.Treeview(cuselection, height=27)
      cusventtree1["columns"]=["1","2","3","4"]
      cusventtree1.column("#0", width=35, anchor="center")
      cusventtree1.column("1", width=160, anchor="center")
      cusventtree1.column("2", width=160, anchor="center")
      cusventtree1.column("3", width=140, anchor="center")
      cusventtree1.column("4", width=140, anchor="center")
      cusventtree1.heading("#0",text="")
      cusventtree1.heading("1",text="Customer/Ventor ID")
      cusventtree1.heading("2",text="Customer/Ventor Name")
      cusventtree1.heading("3",text="Tel.")
      cusventtree1.heading("4",text="Contact Person")
      cusventtree1.place(x=5, y=45)
      def filt():
      # global filttxt 
   
       filtr = filttxt.get()

  
       sqlcusven = "SELECT * FROM Customer WHERE businessname=%s"
       valcusven = (filtr, )
       fbcursor.execute(sqlcusven, valcusven)
       records = fbcursor.fetchall()
      #  print(records)
       if records==[]:
         pass
       else:
        for record in cusventtree1.get_children():
         cusventtree1.delete(record)
        count=0
        for i in records:
          if True:
           cusventtree1.insert(parent='', index='end', iid=i, text='', values=(i[0],i[4],i[10],i[8]))  
          else:
            pass
        count += 1

    # global filttxt
      enter=Label(cuselection, text="Enter filter text").place(x=5, y=10)
      filttxt=Entry(cuselection, width=20)
      filttxt.place(x=110, y=10)
      fltrbutton=Button(cuselection,text = 'Click Here',command=filt)
      fltrbutton.place(x=236,y=7,height=25,width=70)

      
      fbcursor.execute('SELECT * FROM Customer;') 
      j = 0
      for i in fbcursor:
        cusventtree1.insert(parent='', index='end', iid=i, text='', values=(i[0],i[4],i[10],i[8]))
        j += 1
    


      ctegorytree=ttk.Treeview(cuselection, height=27)
      ctegorytree["columns"]=["1"]
      ctegorytree.column("#0", width=35, minwidth=20)
      ctegorytree.column("1", width=205, minwidth=25, anchor=CENTER)    
      ctegorytree.heading("#0",text="", anchor=W)
      ctegorytree.heading("1",text="View filter by category", anchor=CENTER)
      ctegorytree.place(x=660, y=45)



      def items_selected(event):
        selected_indices = listbox.curselection()
        selected_filter = ",".join([listbox.get(i) for i in selected_indices])

        sqloooo = 'SELECT * FROM Productservice'
        fbcursor.execute(sqloooo)
        pandsdatazbn = fbcursor.fetchall()
        print(pandsdatazbn)

      #   psql = "select * from Productservice where serviceornot = %s"
      #   val = ('Not Service', )
      #   fbcursor.execute(psql, val)
      #   pdata = fbcursor.fetchall()
    
      #   ssql = "select * from Productservice where serviceornot = %s"
      #   val = ('Service', )
      #   fbcursor.execute(ssql, val)
      #   sdata = fbcursor.fetchall()

        if selected_filter == "View all records":
          for record in edtcusventtree.get_children():
            edtcusventtree.delete(record)
          countp = 0
          for i in pandsdatazbn:
            edtcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]))
          countp += 1
      #   elif selected_filter == "View all products":
      #     for record in edtcusventtree.get_children():
      #       edtcusventtree.delete(record)
      #     countp = 0
      #     for i in pdata:
      #       edtcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]))
      #     countp += 1
      #   else: 
      #     if selected_filter == "View all services":
      #       for record in edtcusventtree.get_children():
      #         edtcusventtree.delete(record)
      #       countp = 0
      #       for i in sdata:
      #         edtcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]))
      #       countp += 1
      #   # mydb.close()
  
      listbox = Listbox(cuselection,height = 33,  
                            width = 40,  
                            bg = "white",
                            activestyle = 'dotbox',  
                            fg = "black",
                            highlightbackground="white")
     
      listbox.insert(0, "                               View all records")
      # listbox.insert(1, "                               View all products")
      # listbox.insert(2, "                               View all services")
      listbox.place(x=660,y=65,)
      # listbox.bind('<<ListboxSelect>>', items_selected)


      # scrollbar = Scrollbar(cuselection)
      # scrollbar.place(x=640, y=45, height=560)
      # scrollbar.config( command=tree.yview )
      scrollbar = Scrollbar(cuselection)
      scrollbar.place(x=1016+300+25, y=120, height=280+20)

      btn1=Button(cuselection,compound = LEFT,image=tick ,text="ok",command=treefthcng1, width=60).place(x=15, y=610)
      btn1=Button(cuselection,compound = LEFT,image=tick,text="Edit selected customer", width=150,command=create1).place(x=250, y=610)
      btn1=Button(cuselection,compound = LEFT,image=tick, text="Add new customer", width=150,command=create1).place(x=435, y=610)
      btn1=Button(cuselection,compound = LEFT,image=cancel ,text="Cancel", width=60,command=cancelcuselection1).place(x=740, y=610)   



      

    #add new line item
    def newline():
      newselection=Toplevel()
      newselection.title("Select Customer")
      newselection.geometry("930x650+240+10")
      newselection.resizable(False, False)

      
      def edit_productord1():  
         try:
          itemid = edtcusventtree.item(edtcusventtree.focus())["values"][0]
          
          global filename
          filename = ""
          
          def upload_file():
            global filename,img, b2
            f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
            filename = filedialog.askopenfilename(filetypes=f_types)
            shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            image = Image.open(filename)
            resize_image = image.resize((350, 350))
            img = ImageTk.PhotoImage(resize_image)
            b2 = Button(imageFrame,image=img)
            b2.place(x=130, y=80)
          
          def updateproducts():
            global img , filename 
            sku = codeentry.get()
            status = checkvarStatus.get()
            catgory = n.get()
            name = nameentry.get()
            description = desentry.get()
            unitprice = uval.get()
            peices = pcsentry.get()
            cost = costval.get()
            price_cost = priceval.get()
            taxable = checkvarStatus2.get()
            nostockcontrol = checkvarStatus3.get()
            stock = stockval.get()
            lowstock = lowval.get()
            warehouse = wareentry.get()
            pnotes = sctxt.get("1.0", 'end-1c')
            entries = [sku, name, unitprice, cost]
            entri = []
            for i in entries:
              if i == '':
                entri.append(i)
            if len(entri) == 0:
              if filename == "":
                sql = "update Productservice set sku=%s, category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, privatenote=%s where Productserviceid = %s"
                val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes, itemid)
                fbcursor.execute(sql, val)
                fbilldb.commit()
              else:
                file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
                sql = "update Productservice set category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, image=%s, privatenote=%s where Productserviceid = %s"
                val = (catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse,filename.split('/')[-1], pnotes, itemid)
                fbcursor.execute(sql, val)
                fbilldb.commit()
                messagebox.showinfo("F-Billing Revolution", "Product updated successfully.")
                for record in edtcusventtree.get_children():
                  edtcusventtree.delete(record)
                fbcursor.execute("select *  from Productservice")
                pandsdata = fbcursor.fetchall()
                countp = 0
                for i in pandsdata:
                  if i[6] == '1':
                    acti = 'Active'
                  else:
                    acti = 'Inactive' 
                  if i[13] > i[14]:
                    edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                    countp += 1
                  elif (i[12] =="0") == (i[13] <= i[14]):
                    edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                    countp += 1
                  elif i[12] == '1':
                    edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                    countp += 1
                  else:
                    pass
              top.destroy()
            else:
              messagebox.showinfo("F-Billing Revolution", "Fields name or SKU entered is already in database.")
              top.destroy()
            for record in edtcusventtree.get_children():
                  edtcusventtree.delete(record)
            fbcursor.execute("select *  from Productservice")
            pandsdata = fbcursor.fetchall()
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
            
          sql = "select * from Productservice where Productserviceid = %s"
          val = (itemid, )
          fbcursor.execute(sql, val)
          psdata = fbcursor.fetchone()
          
          
          top = Toplevel()  
          top.title("Edit Product/Service details")
          p3 = PhotoImage(file = 'images/fbicon.png')
          top.iconphoto(False, p1)
          top.geometry("600x550+390+125")
          tabControl = ttk.Notebook(top)
          s = ttk.Style()
          s.theme_use('default')
          s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)

          taba = ttk.Frame(tabControl)
          tabb = ttk.Frame(tabControl)
          
          tabControl.add(taba,compound = LEFT, text ='Product/Service')
          tabControl.add(tabb,compound = LEFT, text ='Product Image')
          
          tabControl.pack(expand = 1, fill ="both")
          
          innerFrame = Frame(taba,bg="#f5f3f2", relief=GROOVE)
          innerFrame.pack(side="top",fill=BOTH)

          updateframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
          updateframe.pack(side="top",fill=BOTH,padx=10)

          code1=Label(updateframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
          code1.place(x=20,y=0)
          codeentry = Entry(updateframe,width=35)
          codeentry.place(x=110,y=8)
          codeentry.insert(0, psdata[2])

          checkvarStatus=IntVar()
          status1=Label(updateframe,text="Status:")
          status1.place(x=380,y=8)
          Button1 = Checkbutton(updateframe, 
                            variable = checkvarStatus,text="Active",compound="right",
                            onvalue =1,
                            offvalue =0,
                            width = 10)
          Button1.place(x=420,y=5)
          sta = psdata[6]
          if sta == '1':
            Button1.select()
          else:
            Button1.deselect()



          category1=Label(updateframe,text="Category:",pady=5,padx=10)
          category1.place(x=20,y=40)
          n = StringVar() 
          category = Entry(updateframe,width=70,textvariable=n)
          category.place(x=110,y=45)
          category.insert(0, psdata[3])


          name1=Label(updateframe,text="Name :",fg="blue",pady=5,padx=10)
          name1.place(x=20,y=70)
          nameentry = Entry(updateframe,width=70)
          nameentry.place(x=110,y=75)
          nameentry.insert(0, psdata[4])

          des1=Label(updateframe,text="Description :",pady=5,padx=10)
          des1.place(x=20,y=100)
          desentry = Entry(updateframe,width=70)
          desentry.place(x=110,y=105)
          desentry.insert(0, psdata[5])

          def set_label(name, index, mode):
            priceval.set(uval.get() - costval.get())
          
          unit1=Label(updateframe,text="Unit Price:",fg="blue",pady=5,padx=10)
          unit1.place(x=20,y=130)
          
          uval = IntVar()
          unitentry = Entry(updateframe,width=20,textvariable=uval)
          unitentry.place(x=110,y=135)
          unitentry.delete(0,'end')
          unitentry.insert(0, psdata[7])
          

          pcsval = IntVar()
          pcs1=Label(updateframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
          pcs1.place(x=320,y=130)
          pcsentry = Entry(updateframe,width=20,textvariable=pcsval)
          pcsentry.place(x=410,y=135)
          pcsentry.delete(0,'end')
          pcsentry.insert(0, psdata[8])
          

          costval = IntVar()
          cost1=Label(updateframe,text="Cost:",pady=5,padx=10)
          cost1.place(x=20,y=160)
          costentry = Entry(updateframe,width=20,textvariable=costval)
          costentry.place(x=110,y=165)
          costentry.delete(0,'end')
          costentry.insert(0, psdata[9])
          

          priceval = IntVar()
          price1=Label(updateframe,text="(Price-Cost):",pady=5,padx=10)
          price1.place(x=20,y=190)
          priceentry = Entry(updateframe,width=20,textvariable=priceval)
          priceentry.place(x=110,y=195)
          priceentry.delete(0,'end')
          priceentry.insert(0, psdata[11])

          uval.trace('w', set_label)
          costval.trace('w', set_label)
          

          checkvarStatus2=IntVar()
        
          Button2 = Checkbutton(updateframe,variable = checkvarStatus2, 
                            text="Taxable Tax1rate",compound="right",
                            onvalue =1 ,
                            offvalue =0,
                            height=2,
                            width = 12)

          Button2.place(x=415,y=153)
          tax = psdata[10]
          if tax == '1':
            Button2.select()
          else:
            Button2.deselect()
          

          

          checkvarStatus3=IntVar()
          def switch():
            if checkvarStatus3.get():
              stockentry["state"] = DISABLED
              lowentry["state"] = DISABLED
              wareentry["state"] = DISABLED
            else:
              stockentry["state"] = NORMAL
              lowentry["state"] = NORMAL
              wareentry["state"] = NORMAL
        
          Button3 = Checkbutton(updateframe,variable = checkvarStatus3,command=switch, 
                            text="No stock Control", 
                            onvalue =1 ,
                            offvalue = 0,
                            height=3,
                            width = 15)

          Button3.place(x=40,y=220)

          


          stockval = IntVar(updateframe)
          stock1=Label(updateframe,text="Stock:",pady=5,padx=10)
          stock1.place(x=90,y=260)
          stockentry = Entry(updateframe,width=15,textvariable=stockval)
          stockentry.place(x=140,y=265)
          stockentry.delete(0,'end')
          stockentry.insert(0, psdata[13])
          

          lowval = IntVar(updateframe)
          low1=Label(updateframe,text="Low Stock Warning Limit:",pady=5,padx=10)
          low1.place(x=280,y=260)
          lowentry = Entry(updateframe,width=15,textvariable=lowval)
          lowentry.place(x=435,y=265)
          lowentry.delete(0,'end')
          lowentry.insert(0, psdata[14])
          

        
          ware1=Label(updateframe,text="Warehouse:",pady=5,padx=10)
          ware1.place(x=60,y=290)
          wareentry = Entry(updateframe,width=64)
          wareentry.place(x=140,y=295)
          wareentry.insert(0, psdata[15])

          scr = psdata[12]
          if scr == '1':
            Button3.select()
            stockentry["state"] = DISABLED
            lowentry["state"] = DISABLED
            wareentry["state"] = DISABLED
          else:
            Button3.deselect()
            stockentry["state"] = NORMAL
            lowentry["state"] = NORMAL
            wareentry["state"] = NORMAL
          
          

          

          text1=Label(updateframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
          text1.place(x=20,y=320)
          sctxt = scrolledtext.ScrolledText(updateframe, undo=True,width=62,height=4)
          sctxt.place(x=32,y=358)
          try:
            sctxt.insert("1.0", psdata[16])
          except:
            pass

          okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=updateproducts)
          okButton.pack(side=LEFT, padx=(10, 0))

          cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=lambda : top.destroy())
          cancelButton.pack(side=RIGHT, padx=(0, 10))
          
          
          imageFrame = Frame(tabb, relief=GROOVE,height=580)
          imageFrame.pack(side="top",fill=BOTH)

          browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
          browseimg.place(x=15,y=35)

          browsebutton=Button(imageFrame,text = 'Browse', command=upload_file)
          browsebutton.place(x=470,y=30,height=30,width=50)

          try:
            image = Image.open("images/"+psdata[17])
            resize_image = image.resize((350, 350))
            image = ImageTk.PhotoImage(resize_image)
            b2 = Label(imageFrame,image=image,width=350,height=350)
            b2.photo = image
            b2.place(x=130, y=80)
            print(image)
          except:
            pass

          removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT,command=lambda: b2.destroy())
          removeButton.place(x=410,y=460)
         except:
        #   try:
        #     top.destroy()
        #   except:
            pass
            messagebox.showerror('F-Billing Revolution', 'Select a record to edit.')


      #add new product
      def product():  
          global codeentry,nameentry,country,desentry,unitentry,pcsentry,costentry,priceentry,stockentry,lowentry,wareentry,txt,checkvarStatus,checkvarStatus2,checkvarStatus3
          top = Toplevel()  
          top.title("Add a new Product/Service")
          p2 = PhotoImage(file = 'images/fbicon.png')
          top.iconphoto(False, p2)

          top.geometry("700x550+390+15")
          tabControl = ttk.Notebook(top)
          s = ttk.Style()
          s.theme_use('default')
          s.configure('TNotebook.Tab', background="#999999",padding=10,bd=0)


          tabsb1 = ttk.Frame(tabControl)
          tab2 = ttk.Frame(tabControl)

          tabControl.add(tabsb1,compound = LEFT, text ='Product/Service')
          tabControl.add(tab2,compound = LEFT, text ='Product Image')

          tabControl.pack(expand = 1, fill ="both")

          innerFrame = Frame(tabsb1,bg="#f5f3f2", relief=GROOVE)
          innerFrame.pack(side="top",fill=BOTH)

          Customerlabelframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
          Customerlabelframe.pack(side="top",fill=BOTH,padx=10)
          global filename
          filename = ""
          
          def upload_fileord():
            global filename,img, b2
            f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
            filename = filedialog.askopenfilename(filetypes=f_types)
            shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            image = Image.open(filename)
            resize_image = image.resize((350, 350))
            img = ImageTk.PhotoImage(resize_image)
            b2 = Button(imageFrame,image=img)
            b2.place(x=130, y=80)
          
          def addproductsord():
            global img , filename 
            sku = codeentry.get()
            status = checkvarStatus.get()
            catgory = n.get()
            name = nameentry.get()
            description = desentry.get()
            unitprice = uval.get()
            peices = pcsentry.get()
            cost = costval.get()
            price_cost = priceval.get()
            taxable = checkvarStatus2.get()
            nostockcontrol = checkvarStatus3.get()
            stock = stockval.get()
            lowstock = lowval.get()
            warehouse = wareentry.get()
            pnotes = txt.get("1.0",'end-1c')
            entries = [sku, name, unitprice, cost]
            entri = []
            for i in entries:
              if i == '':
                entri.append(i)
            if len(entri) == 0:
              sql = 'select * from Productservice where sku = %s or name = %s'
              val  = (sku, name)
              fbcursor.execute(sql, val)
              fbcursor.fetchall()
              row_count = fbcursor.rowcount
              if row_count == 0:
                if filename == "":
                  sql = 'insert into Productservice(sku,productserviceid,category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)'
                  val = (sku,sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes)
                  fbcursor.execute(sql, val)
                  fbilldb.commit()
                else:
                  file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
                  sql = 'insert into Productservice(sku,productservice, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, image, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)'
                  val = (sku,sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, filename.split('/')[-1], pnotes)
                  fbcursor.execute(sql, val)
                  fbilldb.commit()
              else:
                messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
              top.destroy()
              messagebox.showinfo("F-Billing Revolution", "Product added successfully.")
              for record in edtcusventtree.get_children():
                edtcusventtree.delete(record)
              fbcursor.execute("select *  from Productservice")
              pandsdata = fbcursor.fetchall()
              countp = 0
              for i in pandsdata:
                if i[6] == '1':
                    acti = 'Active'
                else:
                    acti = 'Inactive' 
                if i[13] > i[14]:
                  edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                  countp += 1
                elif (i[12] =="0") == (i[13] <= i[14]):
                  edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                  countp += 1
                elif i[12] == '1':
                  edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                  countp += 1
                else:
                  pass                
              
            else:
              messagebox.showinfo("Alert", "Fields name and SKU should not be empty.\nFill out required fields and try again")
              top.destroy()
      
          fbcursor.execute("SELECT * FROM Productservice ORDER BY sku DESC LIMIT 1")
          skuin = fbcursor.fetchone()
    
    
          code1=Label(Customerlabelframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
          code1.place(x=20,y=0)
          codeentry = Entry(Customerlabelframe,width=35)
          codeentry.place(x=120,y=8)

          checkvarStatus=IntVar()
          status1=Label(Customerlabelframe,text="Status:")
          status1.place(x=500,y=8)
          Button1 = Checkbutton(Customerlabelframe,
                            variable = checkvarStatus,text="Active",compound="right",
                            onvalue =0 ,
                            offvalue = 1,

                            width = 10)

          Button1.place(x=550,y=5)

          category1=Label(Customerlabelframe,text="Category:",pady=5,padx=10)
          category1.place(x=20,y=40)
          n = StringVar()
          catgory = ttk.Combobox(Customerlabelframe, width = 40, textvariable = n ) 
          catgory.place(x=110,y=45)
          catgory.insert(0, 'Default')


          name1=Label(Customerlabelframe,text="Name :",fg="blue",pady=5,padx=10)
          name1.place(x=20,y=70)
          nameentry = Entry(Customerlabelframe,width=60)
          nameentry.place(x=120,y=75)

          des1=Label(Customerlabelframe,text="Description :",pady=5,padx=10)
          des1.place(x=20,y=100)
          desentry = Entry(Customerlabelframe,width=60)
          desentry.place(x=120,y=105)

          uval = IntVar()
          unit1=Label(Customerlabelframe,text="Unit Price:",fg="blue",pady=5,padx=10)
          unit1.place(x=20,y=130)
          unitentry = Entry(Customerlabelframe,width=20,textvariable=uval)
          unitentry.place(x=120,y=135)

          # pcsval = IntVar(Customerlabelframe, value='0')
          pcs1=Label(Customerlabelframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
          pcs1.place(x=320,y=140)
          pcsentry = Entry(Customerlabelframe,width=20)
          pcsentry.place(x=410,y=140)

          costval = IntVar()
          cost1=Label(Customerlabelframe,text="Cost:",pady=5,padx=10)
          cost1.place(x=20,y=160)
          costentry = Entry(Customerlabelframe,width=20,textvariable=costval)
          costentry.place(x=120,y=165)
          def set_label(name, index, mode):
           priceval.set(uval.get() - costval.get())
          priceval = IntVar()
          price1=Label(Customerlabelframe,text="(Price-Cost):",pady=5,padx=10)
          price1.place(x=20,y=190)
          priceentry = Entry(Customerlabelframe,width=20,textvariable=priceval)
          priceentry.place(x=120,y=195)

          uval.trace('w', set_label)
          costval.trace('w', set_label)
          checkvarStatus2=IntVar()

          Button2 = Checkbutton(Customerlabelframe,variable = checkvarStatus2,
                            text="Taxable Tax1rate",compound="right",
                            onvalue =0 ,
                            offvalue = 1,
                            height=2,
                            width = 12)

          Button2.place(x=415,y=170)
          def switch():
           if checkvarStatus3.get():
            stockentry["state"] = DISABLED
            lowentry["state"] = DISABLED
            wareentry["state"] = DISABLED
           else:
            stockentry["state"] = NORMAL
            lowentry["state"] = NORMAL
            wareentry["state"] = NORMAL

          checkvarStatus3=BooleanVar()

          Button3 = Checkbutton(Customerlabelframe,variable = checkvarStatus3,command=switch,
                            text="No stock Control",
                            onvalue =1 ,
                            offvalue = 0,
                            height=3,
                            width = 15)

          Button3.place(x=40,y=220)


          stockval = IntVar()
          stock1=Label(Customerlabelframe,text="Stock:",pady=5,padx=10)
          stock1.place(x=90,y=260)
          stockentry = Entry(Customerlabelframe,width=15,textvariable=stockval)
          stockentry.place(x=150,y=265)

          lowval = IntVar(Customerlabelframe)
          low1=Label(Customerlabelframe,text="Low Stock Warning Limit:",pady=5,padx=10)
          low1.place(x=300,y=260)
          lowentry = Entry(Customerlabelframe,width=10,textvariable=lowval)
          lowentry.place(x=495,y=265)


          ware1=Label(Customerlabelframe,text="Warehouse:",pady=5,padx=10)
          ware1.place(x=60,y=290)
          wareentry = Entry(Customerlabelframe,width=50)
          wareentry.place(x=150,y=295)

          text1=Label(Customerlabelframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
          text1.place(x=20,y=330)

          txt = scrolledtext.ScrolledText(Customerlabelframe, undo=True,width=72,height=4)
          txt.place(x=32,y=358)




          okButton = Button(innerFrame,compound = LEFT,image=tick , text ="Ok",command=addproductsord,width=60)
          okButton.pack(side=LEFT)
          def closetab():
           top.destroy()

          cancelButton = Button(innerFrame,compound = LEFT,image=cancel ,text="Cancel",width=60, command=closetab)
          cancelButton.pack(side=RIGHT)

          imageFrame = Frame(tab2, relief=GROOVE,height=580)
          imageFrame.pack(side="top",fill=BOTH)

          browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
          browseimg.place(x=15,y=35)

          browsebutton=Button(imageFrame,text = 'Browse',command=upload_fileord)
          browsebutton.place(x=580,y=30,height=30,width=50)

          removeButton = Button(imageFrame,compound = LEFT,image=cancel, text ="Remove Product Image",command=lambda:b2.destroy(),width=150)
          removeButton.place(x=400,y=450)
          # top.mainloop()


      # text=Label(newselection, text="Filtered column").place(x=340, y=10)
      # e2=Entry(newselection, width=20).place(x=450, y=10)
      global edtcusventtree
      edtcusventtree=ttk.Treeview(newselection, height=27)
      edtcusventtree["columns"]=["1","2","3", "4","5"]
      edtcusventtree.column("#0", width=35, anchor="center")
      edtcusventtree.column("1", width=160, anchor="center")
      edtcusventtree.column("2", width=160, anchor="center")
      edtcusventtree.column("3", width=140, anchor="center")
      edtcusventtree.column("4", width=70, anchor="center")
      edtcusventtree.column("5", width=70, anchor="center")
      edtcusventtree.heading("#0",text="")
      edtcusventtree.heading("1",text="ID/SKU")
      edtcusventtree.heading("2",text="Product/Service Name")
      edtcusventtree.heading("3",text="Unit price")
      edtcusventtree.heading("4",text="Service")
      edtcusventtree.heading("5",text="Stock")
      edtcusventtree.place(x=5, y=45)
      edtcusventtree.tag_configure('green', foreground='green')
      edtcusventtree.tag_configure('red', foreground='red')
      edtcusventtree.tag_configure('blue', foreground='blue')
      def filtm():
      # global filttxt 
   
       filtr = filttxt.get()

       sqlcusven = "SELECT * FROM Productservice WHERE name=%s"
       valcusven = (filtr, )
       fbcursor.execute(sqlcusven, valcusven)
       records = fbcursor.fetchall()
      #  print(records)
       if records==[]:
         pass
       else:
        for record in edtcusventtree.get_children():
          edtcusventtree.delete(record)
      
        count=0
        for i in records:
          if True:
           edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]))  
          else:
            pass
        count += 1



    # global filttxt
      enter=Label(newselection, text="Enter filter text").place(x=5, y=10)
      filttxt=Entry(newselection, width=20)
      filttxt.place(x=110, y=10)
      fltrbutton=Button(newselection,text = 'Go',command=filtm)
      fltrbutton.place(x=236,y=7,height=25,width=60)
      fltrbutton=Button(newselection,text = 'Back',command=filtm)
      fltrbutton.place(x=296,y=7,height=25,width=70)
      


      def cancelnewselection():
       newselection.destroy()
       
      global prstree
      def selectp2():
          # selected = edtcusventtree.focus()
          # global valuep10,totalpriceinput
          
          # # # rept = ordtree.item(ordtree.focus())["values"][1]
          # # rept=sctxt.get()
          # # # edvwrept= vwedttree1.item(selected)["values"][1]
          # # norpt="SELECT * FROM storingproduct WHERE orderid=%s "
          # # valrpt=(rept,)
          # # fbcursor.execute(norpt,valrpt)
          # # kooi=fbcursor.fetchall()
          # # # print("koiiiiiiiiiiii",kooi)
          # # # print(rept)

          # valuep10= edtcusventtree.item(selected)["values"][0]
          # sql = "SELECT * FROM productservice WHERE productserviceid=%s"
          # i=(valuep10,)
          # fbcursor.execute(sql,i,)
          # zxy=fbcursor.fetchall()
          # # print("mnnnnnnnnniiii",valuep10)
          # # print(filter(valuep10,kooi[1]))
          
          # # emt=[]
          # # for i in kooi:
          # #   emt.append(i[21])
          #   # print (i[0])
          #   # print (i,"dddddd")
            
          # # if i[0]==valuep10 in rept:
          # totalpriceinput=0
          # j = 0
          # for i in fbcursor:
          #  vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[5],i[7],1,i[8],i[10],(i[7]*1)))
          #  for line in vwedttree1.get_children():
          #     idsave1=vwedttree1.item(line)['values'][7]
          #     totalpriceinput+=idsave1
          #  j += 1
          #  pricecol.config(text=totalpriceinput)



          # valu10= edtcusventtree.item(selected)["values"][0]
          # sqllll = "SELECT * FROM productservice  WHERE productserviceid= %s"
          # r=(valu10,)
          # fbcursor.execute(sqllll,r)
          # child=fbcursor.fetchone() 
          # sql21= 'INSERT INTO storingproduct(Productserviceid,orderid,sku,category,name,description,status,unitprice,peices,cost,taxable,priceminuscost,serviceornot,stock,stocklimit,warehouse,privatenote,quantity) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
          # vatree=(child[0],osdata[0],child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16],1)
          # fbcursor.execute(sql21,vatree,)
          # fbilldb.commit()
          # # else:
          # #     pass
          # # for i in zxy:
          # #   print(i[0],emt)
          # #   if i[0] in emt:
          # #    print("IN")    

          selected = edtcusventtree.focus()
          global valuep10,totalpriceinput
          valuep10= edtcusventtree.item(selected)["values"][0]
          sql = "SELECT * FROM productservice  WHERE productserviceid= %s"
          i=(valuep10,)
          fbcursor.execute(sql,i)
          totalpriceinput=0
          j = 0
          for i in fbcursor:
            vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[5],i[7],1,i[8],i[10],(i[7]*1)))
            for line in vwedttree1.get_children():
              idsave1=vwedttree1.item(line)['values'][7]
              totalpriceinput+=idsave1
          j += 1
          pricecol.config(text=totalpriceinput)

          valu10= edtcusventtree.item(selected)["values"][0]
          sqllll = "SELECT * FROM productservice  WHERE productserviceid= %s"
          r=(valu10,)
          fbcursor.execute(sqllll,r)
          child=fbcursor.fetchone()

          sql21= 'INSERT INTO storingproduct(Productserviceid,orderid,sku,category,name,description,status,unitprice,peices,cost,taxable,priceminuscost,serviceornot,stock,stocklimit,warehouse,privatenote,quantity) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
          vatree=(child[0],osdata[0],child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16],1)
          fbcursor.execute(sql21,vatree,)
          fbilldb.commit()


          newselection.destroy()

      # def selectp2():
      #     selected = edtcusventtree.focus()
      #     valuesp= edtcusventtree.item(selected)["values"][0]
  
      #     sql = "SELECT * FROM productservice  WHERE productserviceid= %s"
      #     i=(valuesp,)
      #     x=fbcursor.execute(sql,i)
      #     print(x)

      #     k = 0
      #     for i in fbcursor:
      #      prstree.insert(parent='', index='end', iid=i, text='', values=(i[4], i[14], i[7], i[9], i[15],i[10],i[12],(i[9]*i[15])))
      #     k += 1

      #     newselection.destroy()

      # def selectp1_and_selectp2():
      #   selectp1()
      #   selectp2()
        
  
  
  
      fbcursor.execute('SELECT * FROM Productservice;') 
      countp = 0
      for i in fbcursor:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
  
  


      ctegorytree=ttk.Treeview(newselection, height=27)
      ctegorytree["columns"]=["1"]
      ctegorytree.column("#0", width=35, minwidth=20)
      ctegorytree.column("1", width=205, minwidth=25, anchor=CENTER)    
      ctegorytree.heading("#0",text="", anchor=W)
      ctegorytree.heading("1",text="View filter by category", anchor=CENTER)
      ctegorytree.place(x=660, y=45)
      ctegorytree.column("#0", width=35, minwidth=20)

      def items_selected(event):
        selected_indices = listbox.curselection()
        selected_filter = ",".join([listbox.get(i) for i in selected_indices])
        mydb = mysql.connector.connect(host='localhost', user='root', password='', port='3306', database='fbillingsintgrtd')
        # fbcursor = fbcursor()

        sqloooo = 'SELECT * FROM Productservice'
        fbcursor.execute(sqloooo)
        pandsdatazbn = fbcursor.fetchall()
        print(pandsdatazbn)



        if selected_filter == "                               View all records":
          # sqloooo = 'SELECT * FROM Productservice'
          # fbcursor.execute(sqloooo)
          # pandsdatazbn = fbcursor.fetchall()
          for record in edtcusventtree.get_children():
           edtcusventtree.delete(record)
          countp = 0
          for i in pandsdatazbn:
            edtcusventtree.insert(parent='', index='end', iid=countp, text='', values=(i[2],i[4],i[7],i[12],i[13]))
          countp += 1
# def ftr(event):
#   filt = dropordr.get()
#   for record in ordtree.get_children():
#     ordtree.delete(record)
    
#   sql = "select * from orders where category = %s"
#   val = (filt, )
#   fbcursor.execute(sql, val)
#   records = fbcursor.fetchall()
  
#   count=0
#   for i in records:
#     if True:
#       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
#     else:
#       pass
#   count += 1
# sql = "SELECT DISTINCT category FROM orders"
# fbcursor.execute(sql,)
# rec = fbcursor.fetchall()
      def items_selected(event):
          selected_indices = listbox.curselection()
          selected_filter = ",".join([listbox.get(i) for i in selected_indices])

          sql = 'select * from Productservice'
          fbcursor.execute(sql)
          pandsdata = fbcursor.fetchall()
          psql = "select * from Productservice where serviceornot=%s"
          val = ('0', )
          fbcursor.execute(psql, val)
          pdata = fbcursor.fetchall()

          ssql = "select * from Productservice where serviceornot=%s"
          val = ('1', )
          fbcursor.execute(ssql, val)
          sdata = fbcursor.fetchall()

          # pssql = "select * from Productservice where category=%s"
          # psval = (selected_filter, )
          # fbcursor.execute(pssql, psval)
          # pssdata = fbcursor.fetchall()
          if selected_filter == "                           View all records":
            for record in edtcusventtree.get_children():
              edtcusventtree.delete(record)
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          elif selected_filter == "                           View all products":
            for record in edtcusventtree.get_children():
              edtcusventtree.delete(record)
            countp = 0
            for i in pdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          elif selected_filter == "                           View all services":
            for record in edtcusventtree.get_children():
              edtcusventtree.delete(record)
            countp = 0
            for i in sdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('green',))  
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('red',))  
                countp += 1
              elif i[12] == '1':
                edtcusventtree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[7],i[12],i[13]),tags=('blue',))  
                countp += 1
              else:
                pass
          # elif selected_filter == True:
          #     for record in treeproducts.get_children():
          #       treeproducts.delete(record)
          #     countp = 0
          #     for i in pssdata:
          #       treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i    [2], i[4], i[7], i[13], i[15]))
          #       countp += 1
        # def cat_pro(event):
        #   selected_indices = listbox.curselection()
        #   selected_filt = ",".join([listbox.get(i) for i in selected_indices])
        #   pssql = "select * from Productservice where category=%s"
        #   psval = (selected_filt, )
        #   fbcursor.execute(pssql, psval)
        #   pssdata = fbcursor.fetchall()
        #   print(pssdata)
        #   for record in treeproducts.get_children():
        #     treeproducts.delete(record)
        #   countp = 0
        #   for i in pssdata:
        #     treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i[2], i[4], i[7], i[13], i[15]))
        #     countp += 1


        # sql = "SELECT DISTINCT category FROM Productservice WHERE NOT category = %s AND NOT category = %s"
        # val = ('service','product',)
        # fbcursor.execute(sql,val,)
        # lic = fbcursor.fetchall()
        # print(lic)
      listbox = Listbox(newselection,height = 33,  
                      width = 40,  
                      bg = "white",
                      activestyle = 'dotbox',  
                      fg = "black",
                      highlightbackground="white")  
      listbox.insert(0, "                           View all records")
      listbox.insert(1, "                           View all products")
      listbox.insert(2, "                           View all services")
      listbox.place(x=660,y=65,)
      listbox.bind('<<ListboxSelect>>', items_selected)

  
      

      # scrollbar = Scrollbar(newselection)
      # scrollbar.place(x=640, y=45, height=560)
      # scrollbar.config( command=tree.yview )
      scrollbar = Scrollbar(newselection)
      scrollbar.place(x=1016+300+25, y=120, height=280+20)
    

      btn1=Button(newselection,compound = LEFT,image=tick ,text="ok",command=selectp2,width=60).place(x=15, y=610)
      btn1=Button(newselection,compound = LEFT,image=tick , text="Edit product/Service", width=150,command=edit_productord1).place(x=250, y=610)
      btn1=Button(newselection,compound = LEFT,image=tick , text="Add product/Service", width=150,command=product).place(x=435, y=610)
      btn1=Button(newselection,compound = LEFT,image=cancel ,text="Cancel", width=60,command=cancelnewselection).place(x=740, y=610)



    #preview new line
    def previewline():
      messagebox.showerror("F-Billing Revolution","line is required,please select customer for this order before printing.")


    
    #sms notification
    def sms1():
      send_SMS=Toplevel()
      send_SMS.geometry("700x480+240+150")
      send_SMS.title("Send SMS notification")

      style = ttk.Style()
      style.theme_use('default')
      style.configure('TNotebook.Tab', background="#999999", padding=5)
      sms_Notebook = ttk.Notebook(send_SMS)
      SMS_Notification = Frame(sms_Notebook, height=470, width=700)
      SMS_Service_Account = Frame(sms_Notebook, height=470, width=700)
      sms_Notebook.add(SMS_Notification, text="SMS Notification")
      sms_Notebook.add(SMS_Service_Account, text="SMS Service Account")
      sms_Notebook.place(x=0, y=0)

      numlbel=Label(SMS_Notification, text="SMS number or comma seperated SMS number list(Please start each SMS number with the country code)")
      numlbel.place(x=10, y=10)
      numentry=Entry(SMS_Notification, width=92).place(x=10, y=30)
      stexbel=Label(SMS_Notification, text="SMS Text").place(x=10, y=60)
      stex=Entry(SMS_Notification, width=40).place(x=10, y=85,height=120)
      
      dclbel=Label(SMS_Notification, text="Double click to insert into text")
      dclbel.place(x=410, y=60)
      dcl=Entry(SMS_Notification, width=30)
      dcl.place(x=400, y=85,height=200)
      
      smstype=LabelFrame(SMS_Notification, text="SMS message type", width=377, height=60)
      smstype.place(x=10, y=223)
      snuvar=IntVar()
      normal_rbtn=Radiobutton(smstype, text="Normal SMS(160 chars)", variable=snuvar, value=1)
      normal_rbtn.place(x=5, y=5)
      unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)", variable=snuvar, value=2)
      unicode_rbtn.place(x=190, y=5)
      tiplbf=LabelFrame(SMS_Notification, text="Tips", width=680, height=120)
      tiplbf.place(x=10, y=290)
      tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="Always start the SMS nymber with the country code. Do not use the + sign at the beginning(example\nUS number:8455807546). Do not use any special characters in your normal SMS text. Please use the\nstndard SMS characters or the English alphabet and numbers only. Otherwise the SMS will be\nunreadable or undeliverable. If you need to enter international characters, accents,email address, or\nspecial characters to the SMS text field then choose the Unicode SMS format.")
      tiplabl.place(x=5, y=5)

      btn1=Button(SMS_Notification, width=20, text="Send SMS notification").place(x=10, y=420)
      btn2=Button(SMS_Notification, width=25, text="Confirm SMS cost before sending").place(x=280, y=420)
      btn3=Button(SMS_Notification, width=15, text="Cancel").place(x=550, y=420)
      

      smstype=LabelFrame(SMS_Service_Account, text="Select the notification service provider", width=670, height=65)
      smstype.place(x=10, y=5)
      snumvar=IntVar()
      normal_rbtn=Radiobutton(smstype,text="BULKSMS(www.bulksms.com)",variable=snumvar,value=1,)
      normal_rbtn.place(x=5, y=5)
      unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)-Recommended", variable=snumvar, value=2)
      unicode_rbtn.place(x=290, y=5)

      sms1type=LabelFrame(SMS_Service_Account, text="Your BULKSMS.COM Account", width=670, height=100)
      sms1type.place(x=10, y=80)
      name=Label(sms1type, text="Username").place(x=10, y=5)
      na=Entry(sms1type, width=20).place(x=100, y=5)
      password=Label(sms1type, text="Password").place(x=10, y=45)
      pas=Entry(sms1type, width=20).place(x=100, y=45)
      combo=Label(sms1type, text="Route").place(x=400, y=5)
      n = StringVar()
      combo1 = ttk.Combobox(sms1type, width = 20, textvariable = n ).place(x=450,y=5)
      btn1=Button(sms1type, width=10, text="Save settings").place(x=550, y=45)

      
      tiplbf=LabelFrame(SMS_Service_Account, text="Terms of service", width=680, height=250)
      tiplbf.place(x=10, y=190)
      tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="The SMS notification service is not free.This service costs you creadit.You must have your own account\nat BULKSMS.COM and you need to have sufficient creadit and an active internet connection to use\nthis feature.Please review all fields in this form for accuracy")
      tiplabl.place(x=0, y=5)
      tiplabl1=Label(tiplbf,justify=LEFT,fg="black",  text="visit www.bulksms.com website to create your own account.please make sure the BULKSMS .COM\n service works well in your country before you busy creadit")
      tiplabl1.place(x=0, y=60)
      tiplabl2=Label(tiplbf,justify=LEFT,fg="black",  text="Our SMS notification tool comes without any warranty.our software only forwards your SMS message\nthe BULKSMS API server .The BULKSMS API server will try to sent SMS message your recipient")
      tiplabl2.place(x=0, y=100)
      tiplabl3=Label(tiplbf,justify=LEFT,fg="red",  text="Please note that you access and use the SMS notification tool your own risk.F-Billing software is not\nresponsible for any type of loss or damage or undelivered SMS massage which you may as a result\nof accessing and using the SMS notification service.")
      tiplabl3.place(x=0, y=140)
      checkvar1=IntVar()
      chkbtn1=Checkbutton(tiplbf,text="I have read and agree to the terms of service above",variable=checkvar1,onvalue=1,offvalue=0).place(x=70, y=200) 



    
    #delete line item  
    def delete2():
        # selected_item = vwedttree1.selection()[0]
        # print(selected_item)
        # vwedttree1.delete(selected_item)

        itemid = vwedttree1.item(vwedttree1.focus())["values"][0]
        quantityid = vwedttree1.item(vwedttree1.focus())["values"][4]
        unitpriceid = vwedttree1.item(vwedttree1.focus())["values"][3]
        nameid = vwedttree1.item(vwedttree1.focus())["values"][1]
        descriid = vwedttree1.item(vwedttree1.focus())["values"][2]
        pcsid = vwedttree1.item(vwedttree1.focus())["values"][5]
        
        print(itemid,)
        dlting=osdata[0]
        sql = 'DELETE FROM storingproduct WHERE orderid=%s AND quantity=%s AND Productserviceid=%s AND peices=%s AND unitprice=%s AND name=%s AND description=%s'
        val = (dlting,quantityid,itemid,pcsid,unitpriceid,nameid,descriid)
        fbcursor.execute(sql, val)
        fbilldb.commit()
        vwedttree1.delete(vwedttree1.selection()[0])

        for record in vwedttree1.get_children():
          vwedttree1.delete(record)
        
        sql = "SELECT * FROM storingproduct WHERE orderid= %s"
        i=(dlting,)
        fbcursor.execute(sql,i)
    
        a=0
        j = 0
        for i in fbcursor:
          vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))
          for line in vwedttree1.get_children():
              idsave=vwedttree1.item(line)['values'][7]
          a+=idsave
        j += 1
        pricecol.config(text=a)
        # vwedttree1.delete(vwedttree1.selection()[0])
        # messdel = messagebox.askyesno("Delete Product", "Are you sure to delete this Product?")
        # if messdel == True:
        #  itemidz = vwedttree1.item(vwedttree1.focus())["values"][0]
        # sqlstrngprdct='DELETE FROM storingproduct WHERE orderid=%s'
        # valstpr = (selected_item,)
        # fbcursor.execute(sqlstrngprdct, valstpr)
        # fbilldb.commit()

        # else:
        #   pass
  
  except:
    try:
      pop.destroy()
    except:
      pass
      messagebox.showerror('F-Billing Revolution', 'Select a record to edit.') 

  #delete orders  
  # def dele():
  #   delmess = messagebox.askyesno("Delete Order", "Are you sure to delete this Order?")
  #   messagebox.showerror("F-Billing Revolution","Customer is required,please select customer before deleting line item .")

  #   if delmess == True:
  #     itemid = ordtree.item(ordtree.focus())["values"][1]
  #     print(itemid,)
  #     sql = 'DELETE FROM Orders WHERE orderid=%s'
  #     val = (itemid,)
  #     fbcursor.execute(sql, val)
  #     fbilldb.commit()
  #     ordtree.delete(ordtree.selection()[0])
  #   else:
  #     pass  

######################################## UPDATING ORDER ##############################################
  def updateorder():
    cmbodto=cmb1.get()
    addrsfrm=addrs1.get('1.0', 'end-1c')
    sptto=spt1.get('1.0', 'end-1c')
    adrsto=adrs.get('1.0', 'end-1c')
    emlfrm=eml.get('1.0', 'end-1c')
    smsfrm=smsno.get('1.0', 'end-1c')
    ordrid=sctxt.get()
    orddatein=e2.get_date()
    ordduein=e3.get_date()
    trmsin=e4.get()
    extracstnme=xtracstnme.get()
    discountrte=dscntrte.get()
    extracst=extra.get()
    taax=taxx.get()
    tplts=tmplte.get()
    slzprzn=salesprsn.get()
    ctgryy=ctgry.get()
    bc="Draft"
    idorder= osdata[0]


    for line in vwedttree1.get_children():
      idsave=vwedttree1.item(line)['values'][0]

      sql1= 'SELECT * FROM  Productservice WHERE Productserviceid = %s'
      val=(idsave,)
      # print(val)
      fbcursor.execute(sql1,val,)
      child=fbcursor.fetchone()
      # print(child)
      # sql2= 'UPDATE storingproduct SET Productserviceid=%s,orderid=%s,sku=%s,category=%s,name=%s,description=%s,status=%s,unitprice=%s,peices=%s,cost=%s,taxable=%s,priceminuscost=%s,serviceornot=%s,stock=%s,stocklimit=%s,warehouse=%s,privatenote=%s'
      # val1=(child[0],sctxt,child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16])
      # fbcursor.execute(sql2,val1,)
      # fbilldb.commit()

    sql3='UPDATE Orders SET orderid=%s,order_date=%s,due_date=%s,businessname=%s,status=%s,extra_cost_name=%s,extra_cost=%s,template=%s,sales_person=%s,discount_rate=%s,tax1=%s,category=%s,businessaddress=%s,shipname=%s,shipaddress=%s,cpemail=%s,cpmobileforsms=%s WHERE orderid=%s'
    val2=(ordrid,orddatein,ordduein,cmbodto,bc,extracstnme,extracst,tplts,slzprzn,discountrte,taax,ctgryy,addrsfrm,sptto,adrsto,emlfrm,smsfrm,idorder)
    fbcursor.execute(sql3,val2,)
    fbilldb.commit()

    sqlup='UPDATE storingproduct SET orderid=%s WHERE orderid=%s'
    valup=(ordrid,idorder,)
    fbcursor.execute(sqlup,valup,)
    fbilldb.commit()

    for record in ordtree.get_children():
      ordtree.delete(record)
    fbcursor.execute("select *  from Orders")
    pandsdata = fbcursor.fetchall()
    countp = 0
    for i in pandsdata:
      ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    countp += 1

    vel=edtnote.get('1.0','end-1c')
    edtntsql="UPDATE orders SET private_notes=%s WHERE orderid=%s"
    edtntsval=(vel,osdata[0])
    fbcursor.execute(edtntsql,edtntsval)
    fbilldb.commit()
    
    notefchz="SELECT * FROM orders WHERE orderid=%s"
    fchz = (osdata[0],)
    fbcursor.execute(notefchz, fchz)
    zoomb=fbcursor.fetchone()   
    notepriv.delete('1.0','end')
    notepriv.insert('1.0', zoomb[27])

    ttlprzinpt=pricecol.cget("text")
    ratediscnt=dscntrte.get()
    xtax=taxx.get()
    cstxtra=xtracst.get()
    print(ttlprzinpt,ratediscnt,xtax,cstxtra)
 
    sdd=round(int(ratediscnt),3)
    pldd=ttlprzinpt
    p=round(pldd, 4)
    dsctedd=(sdd*pldd)/100
    dscdd=round(dsctedd, 4)
    discount=Label(summaryfrme, text=str(sdd)+"% Discount").place(x=0 ,y=0)
    discount1=Label(summaryfrme,text='Rs. '+str(dscdd))
    discount1.place(x=130 ,y=0)
 
    sbtotlhee=p-dscdd
    subtotl=round(sbtotlhee, 4)
    sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
    sub100=Label(summaryfrme, text='Rs. '+str(subtotl))
    sub100.place(x=130 ,y=21)
 
    taxval91=int(xtax)
    tax1dsplay=(taxval91*subtotl)/100
    taxvalu=round(tax1dsplay,2)
    tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
    tax100=Label(summaryfrme, text='Rs. '+str(taxvalu))
    tax100.place(x=130 ,y=42)
 
    xtracstvaluzz=int(cstxtra)
    xtra1cst=round(xtracstvaluzz, 4)
    cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
    cost100=Label(summaryfrme, text='Rs. '+str(xtra1cst)).place(x=130 ,y=63)
  
    otherttlval=subtotl+taxvalu+xtra1cst
    ot1v1=round(otherttlval, 4)
    order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
    order100=Label(summaryfrme, text='Rs. '+str(ot1v1)).place(x=130 ,y=84)
 
    itemidot=ordrid
    ordttl=ot1v1
    sumedtdisc=dscdd
    sumedttax=taxvalu
    sumedtsub=subtotl
    for record in ordtree.get_children():
      ordtree.delete(record)
    sqlordttl = "UPDATE Orders SET sum_discount=%s,sum_tax=%s,sum_subtotal=%s,Order_total=%s WHERE orderid = %s"
    valqordttl = (sumedtdisc,sumedttax,sumedtsub,ordttl,itemidot, )
    fbcursor.execute(sqlordttl, valqordttl,)
    fbilldb.commit()
    fbcursor.execute('SELECT * FROM orders')

    j = 0
    for i in fbcursor:
     ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    
     ordertotalinput=0
     for line in ordtree.get_children():
      idsave1=ordtree.item(line)['values'][9]
      ordertotalinput += idsave1
    j += 1
    ordtotalrowcol.config(text=ordertotalinput)
   
    messagebox.showinfo('Successfully Updated','Successfully Updated')
    

    for k in vwedttree1.get_children():
      pricetotl=vwedttree1.item(k)['values'][7]

  def edtprintsele():

    printordinv = messagebox.askyesno("Download to Print", "Continue to Download?")
    if printordinv == True:
      # itemid = ordtree.item(ordtree.focus())["values"][1]
      # print(itemid,)
      ordrid=sctxt.get()
      sql = 'SELECT * FROM Orders WHERE orderid=%s'
      val = (ordrid,)
      fbcursor.execute(sql, val)
      koko=fbcursor.fetchone()

      # # fbilldb.commit()
      # selected = ordtree.focus()
      # selected_prdct= ordtree.item(selected)["values"][1]

      dateee=dt.datetime.now()
      for record in ordtree.get_children():
       ordtree.delete(record)
      sqlq = "UPDATE Orders SET printed_on=%s WHERE orderid = %s"
    # 'UPDATE storingproduct SET Productserviceid=%s
      valq = (dateee,ordrid, )
      fbcursor.execute(sqlq, valq,)
      fbilldb.commit()
      fbcursor.execute('SELECT * FROM Orders;')
      ordertotalinput=0
      j = 0
      for i in fbcursor:
       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
       for line in ordtree.get_children():
        idsave1=ordtree.item(line)['values'][9]
       ordertotalinput += idsave1
       j += 1

      ordtotalrowcol.config(text=ordertotalinput)    
      sqlprint='SELECT * FROM storingproduct WHERE orderid=%s'
      valprint=(ordrid,)
      fbcursor.execute(sqlprint,valprint)
      mama=fbcursor.fetchall()
      # ordtree.delete(ordtree.selection()[0])
      # sqlstrngprdct='SELECT * FROM storingproduct WHERE orderid=%s'
      # valstpr = (itemid,)
      # fbcursor.execute(sqlstrngprdct, valstpr)
      # mama=fbcursor.fetchone()
      # fbilldb.commit()
      # print("HELLO",koko)
      # print("HAI",mama)
      # prdisc=discount1.cget("text")
    str_html = ""

    str_html+="""
                    <!doctype>
                    <html>
                        <head>
                            <style>
                                .header {
                                    text-align:center;
                                }
                                .body {
                                    font-size: 1.5rem;
                                    font-weight:normal;
                                }
                                table {
                                    
                                    width:100%
                                }
                                td{
                                    text-align: center;
                                    padding-top: 1%;
                                }
                                .thead .th{
                                    border-top: 1px solid #333;
                                    border-bottom: 1px solid #333;
                                    text-align: center;
                                    font-weight: bold;
                                }
                            </style>
                        </head>


                        <body>

                            <div class="header">
                                <h2>Order Invoice</h2>
                            </div>
                            <br><br><br><br><br><br><br>
                            <div>
                                <h3 >Order Id &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[0])+"""</h3>      

                                <h3 >Order Date &nbsp;&nbsp;"""+str(koko[1])+"""</h3>                             

                                <h3 >Due Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[2])+"""</h3>                             
                            </div>                                                        
                            <br><br>
                            <div style="display:flex;">
                                <div>
                                <h2>Order To</h2><h3>"""+str(koko[3])+"""<br>"""+str(koko[16])+"""</h3>
                                </div>
                                <div>
                                <h2>Ship To</h2><h3>"""+str(koko[17])+"""<br>"""+str(koko[18])+"""</h3>                             
                                </div>                       
                            </div>
                            

                            <div class="body">

                                <table>
                                    <tr class="thead" ; style="background-color:orange">
                                        <td class="th">ID/SKU</td>
                                        <td class="th">Product/Service</td>
                                        <td class="th">Description</td>
                                        <td class="th">Quantity</td>
                                        <td class="th">Unit Price</td>
                                        <td class="th">Price</td>
                                    </tr>

                                    <tbody>"""

    for pri in mama:
     str_html+="""<tr>

                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[4])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[6])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[7])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[22])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(pri[9])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str((pri[22])*(pri[9]))+"""</td>
 
                        </tr>"""

    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th style="border: 1px solid black;border-collapse: collapse;">"""+str(koko[13])+'%  Discount'+"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[24])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Subtotal'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[26])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'TAX1'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[25])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Extra Cost'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[10])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Order Total'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[8])+"""</td>
                        </tr>"""
  
    str_html+="""</tbody>
                    </table>
                </div>
            </body>
        </html>"""

    
    # import pdb;pdb.set_trace()


    html_data = str_html

    file_path = os.getcwd()+'/images/'


    options = {'margin-top': '15.00mm',
                    'margin-right': '15.00mm',
                    'margin-bottom': '15.00mm',
                    'margin-left': '15.00mm',
                    'dpi':300,
                    }

    lst_pdfs = []
    str_file_name ='ORD_'+str(koko[0])+'.pdf'
    filename_ledger =  file_path+'/'+str_file_name
    path_wkthmltopdf = b'C:\Program Files\wkhtmltopdf\\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)
    pdfkit.from_string(html_data,filename_ledger,options=options,configuration=config)
    lst_pdfs.append(filename_ledger)

    pdf_writer = PdfFileWriter()
  # for path in lst_pdfs:
  #     pdf_reader = PdfFileReader(path)
  #     for page in range(pdf_reader.getNumPages()):
  #         pdf_writer.addPage(pdf_reader.getPage(page))

  #         str_file_name = 'Ledger-'+str(datetime.now().strftime('%d-%m-%Y_%H_%M_%S_%f'))+'.pdf'
  #         filename =  file_path+'/'+str_file_name

  #         with open(filename, 'wb') as fh:
  #         pdf_writer.write(fh)            

  #         fs = FileSystemStorage()
  #             # lst_file =[filename]
  #             # lst_encoded_string=[]
  #             encoded_string = ''
  #             # for filename in lst_file:
  #             if fs.exists(filename):
  #                 with fs.open(filename) as pdf:
  #                     encoded_string=str(base64.b64encode(pdf.read()))
  #                     # lst_encoded_string.append(str(base64.b64encode(pdf.read())))


  def send_mail():

    sender_email =  email_from.get()    #"myownworkrelated@gmail.com" 
    sender_password = email_pswrd.get() #"Myownworkrelated1234"

    server = smtplib.SMTP('smtp.gmail.com', 587)
    print("Login successful1")
    server.starttls()
    print("Login successful2")
    server.login(sender_email, sender_password)
    print("Login successful3")
    
    # global email_address, email_subject

    # email_address = StringVar()  
    # email_subject = StringVar()
    # msg = MIMEMultipart()

    # for i in htcodeframe.curselection():
    #   # print("hloo",htcodeframe.get(i))
    #   ya=htcodeframe.get(i)
    #   print(ya,"THIS")

    #   fo = open(ya, "rb")
    #   filecontent = fo.read()
    #   encodedcontentnw = base64.b64encode(filecontent) 

    address_info = email_address.get()
    print(address_info)
    # msg['Subject'] = email_subject.get()
    subject_info = email_subject.get()
    print(subject_info)
    carbcopy_info = carcopyem_address.get()
    print(carbcopy_info)
    email_content=mframe.get('1.0','end-1c')
    print(email_content)

    # message_body = email_body.get()

    server.sendmail(sender_email, address_info,email_content,subject_info)
    server.sendmail(sender_email, carbcopy_info, email_content,subject_info)

    print("message sent")
    
    # subent.delete(0, END)
    # emailtoent.delete(0, END)

  def psfile_image(event):

        for i in htcodeframe.curselection():
          print("hloo",htcodeframe.get(i))
          yawn=htcodeframe.get(i)        
          edit_window_img = Toplevel()
          edit_window_img.title("View Image")
          edit_window_img.geometry("700x500")
          image = Image.open("images/"+yawn)
          resize_image = image.resize((700, 500))
          image = ImageTk.PhotoImage(resize_image)
          psimage = Label(edit_window_img,image=image)
          psimage.photo = image
          psimage.pack()



    # add this
  
    # def upload_file1():
    #  global filename,img, b1
    #  f_types =[('Png files','.png'),('Jpg Files', '.jpg')]
    #  filename = filedialog.askopenfilename(filetypes=f_types)
    #  print(filename, 'name')
    #  #import pdb; pdb.set_trace()
    #  shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
    #  image = Image.open(filename)
    #  resize_image = image.resize((120, 120))
    #  img = ImageTk.PhotoImage(resize_image)
    #  b1 = Label(expenselabelframe,image=img, height=120, width=120)
    #  b1.place(x=450 , y=240)


  ###################DB
  #file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
    #query 
    #val=filename.split('/')[-1]
    
  def edemail_order():
    emitid1 = sctxt.get()
    sql = "select * from Orders where orderid = %s"
    val = (emitid1, )
    fbcursor.execute(sql, val)
    emailnow = fbcursor.fetchone()
    mailDetail=Toplevel()
    mailDetail.title("Order E-Mail")
    mailDetail.geometry("1080x550")
    mailDetail.resizable(False, False)
    def my_SMTP():
      if True:
          em_ser_conbtn.destroy()
          mysmtpservercon=LabelFrame(account_Frame,text="SMTP server connection(ask your ISP for your SMTP settings)", height=165, width=380)
          mysmtpservercon.place(x=610, y=110)
          lbl_hostn=Label(mysmtpservercon, text="Hostname").place(x=5, y=10)
          hostnent=Entry(mysmtpservercon, width=30).place(x=80, y=10)
          lbl_portn=Label(mysmtpservercon, text="Port").place(x=5, y=35)
          portent=Entry(mysmtpservercon, width=30).place(x=80, y=35)
          lbl_usn=Label(mysmtpservercon, text="Username").place(x=5, y=60)
          unament=Entry(mysmtpservercon, width=30).place(x=80, y=60)
          lbl_pasn=Label(mysmtpservercon, text="Password").place(x=5, y=85)
          pwdent=Entry(mysmtpservercon, width=30).place(x=80, y=85)
          ssl_chkvar=IntVar()
          ssl_chkbtn=Checkbutton(mysmtpservercon, variable=ssl_chkvar, text="This server requires a secure connection(SSL)", onvalue=1, offvalue=0)
          ssl_chkbtn.place(x=50, y=110)
          em_ser_conbtn1=Button(account_Frame, text="Test E-mail Server Connection").place(x=610, y=285)
      else:
          pass

 
    style = ttk.Style()
    style.theme_use('default')
    style.configure('TNotebook.Tab', background="#999999", padding=5)
    email_Notebook = ttk.Notebook(mailDetail)
    email_Frame = Frame(email_Notebook, height=500, width=1080)
    account_Frame = Frame(email_Notebook, height=550, width=1080)
    email_Notebook.add(email_Frame, text="E-mail")
    email_Notebook.add(account_Frame, text="Account")
    email_Notebook.place(x=0, y=0)

    messagelbframe=LabelFrame(email_Frame,text="Message", height=500, width=730)
    messagelbframe.place(x=5, y=5)
    global email_address, email_subject, email_from,email_pswrd,carcopyem_address,mframe
    email_address = StringVar() 
    email_subject = StringVar()
    # email_body = StringVar()
    email_from = StringVar()
    email_pswrd = StringVar()
    carcopyem_address = StringVar()
      # content_email = StringVar()
    lbl_emailtoaddr=Label(messagelbframe, text="Email to address").place(x=5, y=5)
    emailtoent=Entry(messagelbframe, width=50,textvariable=email_address)
    emailtoent.place(x=120, y=5)
    emailtoent.delete(0,'end')
    emailtoent.insert(0, emailnow[19])
      #email2 = email_address.get()
      #print(email2)
    sendemail_btn=Button(messagelbframe, text="Send Email", width=10, height=1, command=send_mail).place(x=600, y=10)#,command=addacnt

    lbl_carcopyto=Label(messagelbframe, text="Carbon copy to").place(x=5, y=32)
    carcopyent=Entry(messagelbframe, width=50,textvariable=carcopyem_address)
    carcopyent.place(x=120, y=32)
      # stopemail_btn=Button(messagelbframe, text="Stop sending", width=10, height=1).place(x=600, y=40)
    lbl_subject=Label(messagelbframe, text="Subject").place(x=5, y=59)
    subent=Entry(messagelbframe, width=50, textvariable=email_subject)
    subent.place(x=120, y=59)
    subjectinsrt='ORD_'+str(emailnow[0])
    subent.delete(0,'end')
    subent.insert(0, subjectinsrt)

      
    style = ttk.Style()
    style.theme_use('default')
    style.configure('TNotebook.Tab', background="#999999", width=20, padding=5)
    mess_Notebook = ttk.Notebook(messagelbframe)
    emailmessage_Frame =Frame(mess_Notebook, height=350, width=710)
    htmlsourse_Frame = Frame(mess_Notebook, height=350, width=710)
    mess_Notebook.add(emailmessage_Frame, text="E-mail message")
      # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
      # mess_Notebook.place(x=120, y=59)
    mess_Notebook.add(htmlsourse_Frame, text="Html source code")
    mess_Notebook.place(x=5, y=90)

    btn1=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

      
    btn2=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
    btn3=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
    btn4=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
    btn5=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=undo).place(x=140, y=1)
    btn6=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=redo).place(x=175, y=1)
    btn7=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=bold).place(x=210, y=1)
    btn8=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=italics).place(x=245, y=1)
    btn9=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=underline).place(x=280, y=1)
    btn10=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=left).place(x=315, y=1)
    btn11=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=right).place(x=350, y=1)
    btn12=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=center).place(x=385, y=1)
    btn13=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=hyperlink).place(x=420, y=1)
      
    btn14=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=remove).place(x=455, y=1)


    dropcomp = ttk.Combobox(emailmessage_Frame, width=12, height=3).place(x=500, y=5)
    dropcompo = ttk.Combobox(emailmessage_Frame, width=6, height=3).place(x=600, y=5)
    mframe=scrolledtext.Text(emailmessage_Frame,  undo=True,width=88, bg="white", height=350)
    mframe.place(x=0, y=28)
      # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
      # mess_Notebook.place(x=120, y=59)


    btn1=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

      
    btn2=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
    btn3=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
    btn4=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
    mframe1=Frame(htmlsourse_Frame, height=350, width=710, bg="white")
    mframe1.place(x=0, y=28)
    attachlbframe=LabelFrame(email_Frame,text="Attachment(s)", height=350, width=280)
    attachlbframe.place(x=740, y=5)
    htcodeframe=Listbox(attachlbframe, height=220, width=265, bg="white")
    htcodeframe.place(x=5, y=5)

    lbl_btn_info=Label(attachlbframe, text="Double click on attachment to view").place(x=30, y=230)
    btn17=Button(attachlbframe, width=20, text="Add attachment file....", command=UploadAction).place(x=60, y=260)
    btn18=Button(attachlbframe, width=20, text="Remove attachment").place(x=60, y=295)
    lbl_tt_info=Label(email_Frame, text="You can create predefined invoice, order, estimate\nand payment receipt email templates under Main\nmenu/Settings/E-Mail templates tab")
    lbl_tt_info.place(x=740, y=370)

    ready_frame=Frame(mailDetail, height=20, width=1080, bg="#b3b3b3").place(x=0,y=530)
      
    sendatalbframe=LabelFrame(account_Frame,text="E-Mail(Sender data)",height=270, width=600)
    sendatalbframe.place(x=5, y=5)
    lbl_sendermail=Label(sendatalbframe, text="Company email address").place(x=5, y=10)
    sentent=Entry(sendatalbframe, width=40, textvariable=email_from)
    sentent.place(x=195, y=10)
      #############################################
    lbl_senderpswrd=Label(sendatalbframe, text="Email Password").place(x=5, y=40)
    pswrdent=Entry(sendatalbframe, width=40, textvariable=email_pswrd,show="*")
    pswrdent.place(x=195, y=40)
      #############################################
      # lbl_orcompanyname=Label(sendatalbframe, text="Your name or company name").place(x=5, y=70)
      # nament=Entry(sendatalbframe, width=40).place(x=195, y=70)
      # lbl_reply=Label(sendatalbframe, text="Reply to email address").place(x=5, y=100)
      # replyent=Entry(sendatalbframe, width=40).place(x=195, y=100)
      # lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=140)
      # signent=Entry(sendatalbframe,width=50).place(x=100, y=140,height=75)
      # @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=95)
    signent=Entry(sendatalbframe,width=50).place(x=195, y=70,height=95)
    confirm_chkvar=IntVar()
    confirm_chkbtn=Checkbutton(sendatalbframe, variable=confirm_chkvar, text="Confirmation reading", onvalue=1, offvalue=0)
    confirm_chkbtn.place(x=200, y=215)
    btn18=Button(account_Frame, width=15, text="Save settings",command=savesettings).place(x=25, y=285)

    sendatalbframe=LabelFrame(account_Frame,text="SMTP Server",height=100, width=380)
    sendatalbframe.place(x=610, y=5)
    servar=IntVar()
    SMTP_rbtn=Radiobutton(sendatalbframe, text="Use the Built-In SMTP Server Settings", variable=servar, value=1)
    SMTP_rbtn.place(x=10, y=10)
    MySMTP_rbtn=Radiobutton(sendatalbframe, text="Use My Own SMTP Server Settings(Recommended)", variable=servar, value=2, command=my_SMTP)
    MySMTP_rbtn.place(x=10, y=40)
    em_ser_conbtn=Button(account_Frame, text="Test E-mail Server Connection")
    em_ser_conbtn.place(x=710, y=110)

  def UploadAction(event=None):
    # filename = filedialog.askopenfilename()
    # print('Selected:', filename)
    # name = askopenfilename(filetypes=[('PDF', '*.pdf',)])

    filename = askopenfilename(filetypes=(("png file ",'.png'),("jpg file", ".jpg"), ('PDF', '*.pdf',), ("All files", "*.*"),))
    shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
    htcodeframe.insert(0, filename.split('/')[-1])



      
    
  firFrame=Frame(pop, bg="#f5f3f2", height=60)
  firFrame.pack(side="top", fill=X)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  create = Button(firFrame,compound="top", text="Select\nCustomer",relief=RAISED, image=customer,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=custom)
  create.pack(side="left", pady=3, ipadx=4)


  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  add= Button(firFrame,compound="top", text="Add new\nline item",relief=RAISED, image=photo,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=newline)
  add.pack(side="left", pady=3, ipadx=4)

  dele= Button(firFrame,compound="top", text="Delete line\nitem",relief=RAISED, image=photo2,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=delete2)
  dele.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  # prev= Button(firFrame,compound="top", text="Preview\nOrder",relief=RAISED, image=photo5,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=previewline)
  # prev.pack(side="left", pady=3, ipadx=4)
  prin= Button(firFrame,compound="top", text="Print \nOrder",relief=RAISED, image=photo4,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=edtprintsele)
  prin.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  mail= Button(firFrame,compound="top", text="Email\nOrder",relief=RAISED, image=photo6,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=email_order)
  mail.pack(side="left", pady=3, ipadx=4)

  sms1= Button(firFrame,compound="top", text="Send SMS\nnotification",relief=RAISED, image=photo10,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=sms1)
  sms1.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="left", padx=5)

  calc= Button(firFrame,compound="top", text="Open\nCalculator",relief=RAISED, image=photo9,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=calcu)
  calc.pack(side="left", pady=3, ipadx=4)

  w = Canvas(firFrame, width=1, height=65, bg="#b3b3b3", bd=0)
  w.pack(side="right", padx=5)

  Createorder= Button(firFrame,compound="top", text="Save\nChanges",relief=RAISED, image=tick,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=updateorder)
  Createorder.pack(side="right", pady=3, ipadx=4)

  fir1Frame=Frame(pop, height=180,bg="#f5f3f2")
  fir1Frame.pack(side="top", fill=X)

    
  # fbcursor.execute('SELECT * FROM Productservice;') 
  # j = 0
  # for i in fbcursor:
  #     labelframe1.insert(parent='', index='end', iid=i, text='', values=(' ',i[0],i[4],i[5],i[7],i[8],i[10],))
  #     j += 1


  labelframe1 = LabelFrame(fir1Frame,text="Customers",font=("arial",15))
  labelframe1.place(x=10,y=5,width=640,height=160)


  def treefthcng1():
    itemid = cusventtree1.item(cusventtree1.focus())["values"][0]
    sql = "select * from Customer where customerid = %s"
    val = (itemid, )
    fbcursor.execute(sql, val)
    slctcstr = fbcursor.fetchone()
    cmb1.delete(0,'end')
    cmb1.insert(0, slctcstr[4])
    addrs1.delete('1.0','end')
    addrs1.insert("1.0", slctcstr[5])
    spt1.delete('1.0','end')
    spt1.insert('1.0', slctcstr[6])
    adrs.delete('1.0','end')
    adrs.insert('1.0', slctcstr[7])
    eml.delete('1.0','end')
    eml.insert('1.0', slctcstr[9])
    smsno.delete('1.0','end')
    smsno.insert('1.0', slctcstr[10])
    cuselection.destroy()

##
  sql = "select businessname from Customer"
  fbcursor.execute(sql,)
  pdata = fbcursor.fetchall()
##
  itemid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select * from Orders where orderid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  global osdata
  osdata = fbcursor.fetchone()
##


  order = Label(labelframe1, text="Order to").place(x=10,y=5)
  # cmb1 = ttk.Combobox(labelframe1,values=pdata,width=28)
  cmb1 = Entry(labelframe1,text=pdata,width=31)
  cmb1.place(x=80,y=5)
  cmb1.delete(0,'end')
  cmb1.insert(0, osdata[3])

  address=Label(labelframe1,text="Address").place(x=10,y=30)
  # e2=scrolledtext(labelframe1,width=23).place(x=80,y=30,height=70)
  addrs1 = scrolledtext.Text(labelframe1, undo=True,width=23)
  addrs1.place(x=80,y=30,height=70)
  addrs1.delete('1.0','end')
  addrs1.insert("1.0", osdata[16])
  
  ship=Label(labelframe1,text="Ship to").place(x=342,y=5)
  #e3=Entry(labelframe1,width=30).place(x=402,y=3)
  spt1 = scrolledtext.Text(labelframe1, undo=True,width=23)
  spt1.place(x=402,y=3)
  spt1.delete('1.0','end')
  spt1.insert("1.0", osdata[17])

  address1=Label(labelframe1,text="Address").place(x=340,y=30)
  # e4=Text(labelframe1,width=23).place(x=402,y=30,height=70)
  adrs = scrolledtext.Text(labelframe1, undo=True,width=23)
  adrs.place(x=402,y=30,height=70)
  adrs.delete('1.0','end')
  adrs.insert("1.0", osdata[18])
  
##
  def copying():
    # shpdlt=spt1.get(osdata[17])
    spt1.delete('1.0',END)
    spt1.insert('1.0',osdata[3])
    adrs.delete('1.0',END)
    adrs.insert('1.0',osdata[16])
##
  btnedt=Button(labelframe1,width=3,height=2,compound = LEFT,text=">>",command=copying)
  btnedt.place(x=290, y=48)
  
  labelframe2 = LabelFrame(fir1Frame,text="")
  labelframe2.place(x=10,y=130,width=640,height=42)

  email=Label(labelframe2,text="Email").place(x=10,y=5)
  # e5=Entry(labelframe2,width=30).place(x=80,y=5)
  eml = scrolledtext.Text(labelframe2, undo=True,width=23)
  eml.place(x=80,y=5,height=25)
  eml.delete('1.0','end')
  eml.insert("1.0", osdata[19])


  sms=Label(labelframe2,text="SMS No.").place(x=340,y=5)
  # e6=Entry(labelframe2,width=30).place(x=402,y=5)
  smsno = scrolledtext.Text(labelframe2, undo=True,width=23)
  smsno.place(x=402,y=5,height=25)
  smsno.delete('1.0','end')
  smsno.insert("1.0", osdata[20])
    
  labelframe = LabelFrame(fir1Frame,text="Order",font=("arial",15))
  labelframe.place(x=652,y=5,width=290,height=170)
  order=Label(labelframe,text="Order#").place(x=5,y=5)
  # e1=Entry(labelframe,width=25).place(x=100,y=5,)
  sctxt = Entry(labelframe,width=25)
  sctxt.place(x=100,y=5,height=22)
  sctxt.delete(0,'end')
  sctxt.insert(0, osdata[0])


  orderdate=Label(labelframe,text="Order date").place(x=5,y=33)
  e2=DateEntry(labelframe,width=25)
  e2.place(x=100,y=33)
  # sctxt = scrolledtext.Text(labelframe, undo=True,width=15)
  # sctxt.place(x=150,y=33,height=22)
  e2.delete(0,'end')
  e2.insert(0, osdata[1].strftime('%d/%m/%y'))

  def duecheckord2():
   if checkvarStatus5.get()==0:
     e3['state'] = DISABLED  
   else:
     e3['state'] = NORMAL
  

  checkvarStatus5=IntVar()
  duedate=Checkbutton(labelframe,variable = checkvarStatus5,text="Due date",onvalue =1 ,offvalue = 0, command=duecheckord2).place(x=5,y=62)
  
  e3=DateEntry(labelframe,width=25)
  e3.place(x=100,y=62)
  # sctxt = scrolledtext.Text(labelframe, undo=True,width=15)
  # sctxt.place(x=150,y=63,height=22)
  e3.delete(0,'end')
  e3.insert(0, osdata[2].strftime('%d/%m/%y'))

  terms=Label(labelframe,text="Terms").place(x=5,y=92)
  e4=ttk.Combobox(labelframe, value="",width=25)
  e4.place(x=100,y=92)



  # ref=Label(labelframe,text="Order ref#").place(x=5,y=118)
  # e1=Entry(labelframe,width=27).place(x=100,y=118)

  fir2Frame=Frame(pop, height=150,width=100,bg="#f5f3f2")
  fir2Frame.pack(side="top", fill=X)
  listFrame = Frame(fir2Frame, bg="white", height=140,borderwidth=5,  relief=RIDGE)

  global vwedttree1
  vwedttree1=ttk.Treeview(listFrame, show = "headings")
  vwedttree1.pack(side = 'top')
  vwedttree1["columns"]=["1","2","3","4","5","6","7","8"]
  vwedttree1.column("#0", width=40, anchor="center")
  vwedttree1.column("1", width=80, anchor="center")
  vwedttree1.column("2", width=190, anchor="center")
  vwedttree1.column("3", width=190, anchor="center")
  vwedttree1.column("4", width=80, anchor="center")
  vwedttree1.column("5", width=60, anchor="center")
  vwedttree1.column("6", width=60, anchor="center")
  vwedttree1.column("7", width=60, anchor="center")
  vwedttree1.column("8", width=80, anchor="center")
 
  vwedttree1.heading("#0")
  vwedttree1.heading("1",text="ID/SKU")
  vwedttree1.heading("2",text="Product/Service")
  vwedttree1.heading("3",text="Description")
  vwedttree1.heading("4",text="Unit Price")
  vwedttree1.heading("5",text="Quantity")
  vwedttree1.heading("6",text="Pcs/Weight")
  vwedttree1.heading("7",text="Tax1")
  vwedttree1.heading("8",text="Price")
  pricecol = Label(listFrame,bg="#f5f3f2")
  pricecol.place(x=850,y=200,width=78,height=18)

  ba=osdata[0]
  sql4='SELECT * FROM storingproduct WHERE orderid=%s'
  val10=(ba,)
  fbcursor.execute(sql4,val10)
  totalpriceinput=0
  j = 0
  for i in fbcursor:
    vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))
    # for line in vwedttree1.get_children():
    #   idsave1=vwedttree1.item(line)['values'][7]
    totalpriceinput+=(i[9]*i[22])
    j += 1
  pricecol.config(text=totalpriceinput)
      

  vwedttree1.pack(fill="both", expand=1)
  listFrame.pack(side="top", fill="both", padx=5, pady=3, expand=1)

  # def prdctvwedtupdt():
  #     for line in vwedttree1.get_children():
  #       idsave=vwedttree1.item(line)['values'][0]

  #       sql1= 'SELECT * FROM  Productservice WHERE Productserviceid = %s'
  #       val=(idsave,)
  #       # print(val)
  #       fbcursor.execute(sql1,val,)
  #       child=fbcursor.fetchone()
  #       sql2= 'UPDATE storingproduct SET Productserviceid=%s,orderid=%s,sku=%s,category=%s,name=%s,description=%s,status=%s,unitprice=%s,peices=%s,cost=%s,taxable=%s,priceminuscost=%s,serviceornot=%s,stock=%s,stocklimit=%s,warehouse=%s,privatenote=%s'
  #       val1=(child[0],sctxt,child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16])
  #       fbcursor.execute(sql2,val1,)
  #       fbilldb.commit()
  # updatetree=prdctvwedtupdt()
  

  # fbcursor.execute('SELECT * FROM Productservice;') 
  # j = 0
  # for i in fbcursor:
  #     vwedttree.insert(parent='', index='end', iid=i, text='', values=(i[2],i[4],i[5],i[7],i[13],i[8],i[10],(i[7]*i[13])))
  #     j += 1

  new_value = tk.StringVar()

  def edit_window_box(val):
    
    edit_window = Toplevel(root)
    edit_window.title("Edit the value or cancel")
    edit_window.geometry("400x200")
    label_edit = tk.Label(edit_window , text='Enter value to edit -', 
    font = ("Times", 12)).place(x=68,y=60)
    #create edit box
    edit_box = tk.Entry(edit_window)
    edit_box.insert(0,val)
    edit_box.place(x=200,y=63)
    #auto select edit window 
    edit_window.focus()
    
    def value_assignment(event):
        printing = edit_box.get()
        # print(printing)
        indexcolaa=int(indexcol)-1
        new_value.set(printing)
        print(indexcol)
        #only destroy will not update the value (perhaps event keeps running in background)
        #quit allows event to stop n update value in tree but does not close the window in single click 
        #rather on dbl click shuts down entire app 

        selected0 = vwedttree1.focus()
        valuz1= vwedttree1.item(selected0)["values"]
        idgettingprdt=valuz1[0]
        valuz1[indexcolaa]=printing
        print(valuz1)

        sql2z0= 'UPDATE storingproduct SET sku=%s,name=%s,description=%s,unitprice=%s,peices=%s,taxable=%s WHERE orderid=%s AND Productserviceid=%s'
        val0z1=(valuz1[0],valuz1[1],valuz1[2],valuz1[3],valuz1[5],valuz1[6],osdata[0],idgettingprdt)
        # dnz=(ordrid,sql2z0,)
        print(val0z1)
        fbcursor.execute(sql2z0,val0z1)
        fbilldb.commit()

        qntsql= 'UPDATE storingproduct SET quantity=%s WHERE Productserviceid=%s'
        qntval=(valuz1[4],idgettingprdt)
        fbcursor.execute(qntsql,qntval)
        fbilldb.commit()


        fbcursor.execute("SELECT * FROM storingproduct ORDER BY orderid DESC LIMIT 1")
        bolo = fbcursor.fetchone()[0]
        print(bolo)
        print(osdata[0])

        # if bolo==osdata[0]:        
        for record in vwedttree1.get_children():
         vwedttree1.delete(record)
        m="SELECT * FROM storingproduct  WHERE orderid=%s"
        i=(osdata[0],)
        fbcursor.execute(m,i)
        panrefrdata = fbcursor.fetchall()
        az=0
        countp = 0
        for i in panrefrdata:
          vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))
          az+=(i[9]*i[22])
        countp += 1
        pricecol.config(text=az)
        # else:
        #   pass

        edit_window.quit()
        edit_window.destroy()
    
    edit_window.bind('<Return>', value_assignment )
    


    B1 = tk.Button(edit_window, text=" Okay ")
    B1.bind('<Button-1>',value_assignment)
    B1.place(x=70,y=130)
    
    
    B2 = tk.Button(edit_window, text="Cancel", command = edit_window.destroy).place(x=276,y=130)
    edit_window.mainloop()
    
  #will explain
  #variable to hold col value (col clicked)
  shape1 = tk.IntVar()
  #tracks both col , row on mouse click
  def tree_click_handler(event):
      cur_item = vwedttree1.item(vwedttree1.focus())
      # print(cur_item)
      global indexcol
      indexcol = vwedttree1.identify_column(event.x)[1:]
      rowid = vwedttree1.identify_row(event.y)[1:]
      # print(indexcol)
      # print(rowid)
      #updates list
      shape1.set(indexcol)
      try:
          x,y,w,h = vwedttree1.bbox('I'+rowid,'#'+indexcol)
      except:pass
      #tree.tag_configure("highlight", background="yellow")
      return(indexcol)

  #code linked to event    
  vwedttree1.bind('<ButtonRelease-1>', tree_click_handler)

  #edit a value in a clicked cell
  def edit(event):
      try:
          selected_item = vwedttree1.selection()[0]
          temp = list(vwedttree1.item(selected_item , 'values'))
          tree_click_handler
          col_selected = int(shape1.get())-1
          edit_window_box(temp[col_selected])
          #do not run if edit window is open
          #use edit_window.mainloop() so value assign after window closes
          temp[col_selected] = new_value.get()
          vwedttree1.item(selected_item, values= temp)
          # prstree.insert(prstree.selection()[0])
      except: pass
  #binding allows to edit on screen double click
  vwedttree1.bind('<Double-Button-1>' , edit)


  fir3Frame=Frame(pop,height=200,width=700,bg="#f5f3f2")
  fir3Frame.place(x=0,y=490)

  tabStyle = ttk.Style()
  tabStyle.theme_use('default')
  tabStyle.configure('TNotebook.Tab', background="#999999", width=12, padding=5)
  myNotebook=ttk.Notebook(fir3Frame)
  orderFrame = Frame(myNotebook, height=200, width=800)
  headerFrame = Frame(myNotebook, height=200, width=800)
  commentFrame = Frame(myNotebook, height=200, width=800)
  termsFrame = Frame(myNotebook, height=200, width=800)
  noteFrame = Frame(myNotebook, height=200, width=800)
  documentFrame = Frame(myNotebook, height=200, width=800)
  
  myNotebook.add(orderFrame,compound="left", text="Order")
  myNotebook.add(headerFrame,compound="left",  text="Header/Footer")
  myNotebook.add(commentFrame,compound="left",  text="Comments")
  myNotebook.add(termsFrame,compound="left", text="Terms")
  myNotebook.add(noteFrame,compound="left",  text="Private notes")
  myNotebook.add(documentFrame,compound="left",  text="Documents")
  myNotebook.pack(expand = 1, fill ="both")  

##
  sql = "select extra_cost_name from Orders"
  fbcursor.execute(sql,)
  exdata = fbcursor.fetchall()  
##
  itemid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select * from Orders where orderid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  orbtdata = fbcursor.fetchone()
##

  labelframe1 = LabelFrame(orderFrame,text="",font=("arial",15))
  labelframe1.place(x=1,y=1,width=800,height=170)

  cost1=Label(labelframe1,text="Extra cost name").place(x=2,y=5)
  xtracstnme=ttk.Combobox(labelframe1, value=exdata,width=20)
  xtracstnme.place(x=115,y=5)
  xtracstnme.delete(0,'end')
  xtracstnme.insert(0, orbtdata[9])


  rate=Label(labelframe1,text="Discount rate").place(x=370,y=5)
  dscntrte=Spinbox(labelframe1,width=6,  from_=0, to=100, font="italic 10")
  dscntrte.place(x=460,y=5)
  dscntrte.delete(0,'end')
  dscntrte.insert(0, orbtdata[13])

  
  extra=IntVar()
  cost2=Label(labelframe1,text="Extra cost").place(x=35,y=35)
  xtracst=Entry(labelframe1,width=10,textvariable=extra)
  xtracst.place(x=115,y=35)
  xtracst.delete(0,'end')
  xtracst.insert(0, orbtdata[10])


  tax=Label(labelframe1,text="Tax1").place(x=420,y=35)
  taxx=Entry(labelframe1,width=7)
  taxx.place(x=460,y=35)
  taxx.delete(0,'end')
  taxx.insert(0, orbtdata[14])

##
  sql = "select template from Orders"
  fbcursor.execute(sql,)
  tmpltdata = fbcursor.fetchall()  
##



  template=Label(labelframe1,text="Template").place(x=37,y=70)
  tmplte=ttk.Combobox(labelframe1, value=tmpltdata,width=25)
  tmplte.place(x=115,y=70)
  tmplte.delete(0,'end')
  tmplte.insert(0, orbtdata[11])



  sales=Label(labelframe1,text="Sales Person").place(x=25,y=100)
  salesprsn=Entry(labelframe1,width=18)
  salesprsn.place(x=115,y=100)
  salesprsn.delete(0,'end')
  salesprsn.insert(0, orbtdata[12])


  category=Label(labelframe1,text="Category").place(x=300,y=100)
  ctgry=Entry(labelframe1,width=22)
  ctgry.place(x=370,y=100)
  ctgry.delete(0,'end')
  ctgry.insert(0, orbtdata[15])
  
  statusfrme = LabelFrame(labelframe1,text="Status",font=("arial",15))
  statusfrme.place(x=540,y=0,width=160,height=160)
##
  itemid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select status from Orders where orderid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  drftinvc = fbcursor.fetchone()
##
##
  itemid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select emailed_on from Orders where orderid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  emldon = fbcursor.fetchone()
##
##
  itemid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select printed_on from Orders where orderid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  prntdon = fbcursor.fetchone()
##
  draft=Label(statusfrme, text=drftinvc,font=("arial", 12, "bold"), fg="black")
  draft.place(x=50, y=3)

  on1=Label(statusfrme, text="Emailed on:").place( y=50)
  nev1=Label(statusfrme, text=emldon).place(x=80,y=50)
  on2=Label(statusfrme, text="Printed on:").place( y=90)
  nev2=Label(statusfrme, text=prntdon).place(x=80,y=90)

  text1=Label(headerFrame,text="Title text").place(x=50,y=5)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=5)
  text2=Label(headerFrame,text="Page header text").place(x=2,y=45)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=45)
  text3=Label(headerFrame,text="Footer text").place(x=35,y=85)
  e1=ttk.Combobox(headerFrame, value="",width=60).place(x=125,y=85)

  def edtnotepri():
    vel=edtnote.get('1.0','end-1c')
    edtntsql="UPDATE orders SET private_notes=%s WHERE orderid=%s"
    edtntsval=(vel,osdata[0])
    fbcursor.execute(edtntsql,edtntsval)
    fbilldb.commit()
    notefchz="SELECT * FROM orders WHERE orderid=%s"
    fchz = (osdata[0],)
    fbcursor.execute(notefchz, fchz)
    zoomb=fbcursor.fetchone()   
    notepriv.delete('1.0','end')
    notepriv.insert('1.0', zoomb[27])
  global edtnote
  text=Label(noteFrame,text="Private notes (not shown on invoice/order/estimates)").place(x=10,y=10)
  edtnote=scrolledtext.Text(noteFrame,width=90, undo=True,height=7)
  edtnote.place(x=10,y=32)
  edtnotesave=Button(noteFrame,height=1,width=8,text="Save Note",command=edtnotepri)
  edtnotesave.place(x=600,y=5)
  notefch="SELECT * FROM orders WHERE orderid=%s"
  fch = (osdata[0],)
  fbcursor.execute(notefch, fch)
  ffoo=fbcursor.fetchone()
  print(ffoo)
  edtnote.delete('1.0','end')
  edtnote.insert('1.0', ffoo[27])

  e1=Text(termsFrame,width=100,height=9).place(x=10,y=10)

  e1=Text(commentFrame,width=100,height=9).place(x=10,y=10)
  def ordedupload_file1():
        global filename
        # import shutil
        f_types =[('Png files','.png'),('Jpg Files', '.jpg'),('All Files', '*.*')]
        
        filename = filedialog.askopenfilename(filetypes=f_types)
        shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
        sql = 'INSERT INTO documents(add_document,orderid) VALUES(%s,%s)'
        val = (filename.split('/')[-1],osdata[0])
        fbcursor.execute(sql,val)
        fbilldb.commit()
        sqldoc='SELECT * FROM documents WHERE orderid=%s'
        vall=(osdata[0],)
        fbcursor.execute(sqldoc,vall)
        for record in cusventtreeeddc.get_children():
          cusventtreeeddc.delete(record) 
        j = 0
        for i in fbcursor:
         print(i)
         print(i[2])
         cusventtreeeddc.insert(parent='', index='end', iid=i, text='', values=(i[2]))
        j += 1

  def rmvatt():
    itemidcsv = cusventtreeeddc.item(cusventtreeeddc.focus())["values"][0]
    # print(itemid,)
    # Nullified="NULL"
    sql9 = "DELETE FROM documents WHERE add_document=%s AND orderid=%s"
    val9 = (itemidcsv,osdata[0],)
    fbcursor.execute(sql9, val9)
    fbilldb.commit()
    cusventtreeeddc.delete(cusventtreeeddc.selection()[0])

  btn1=Button(documentFrame,height=2,width=3,text="+",command=ordedupload_file1).place(x=5,y=10)
  btn2=Button(documentFrame,height=2,width=3,text="-",command=rmvatt).place(x=5,y=50)
  # text=Label(documentFrame,text="Attached documents or image files.If you attach large email then email taken long time to send").place(x=50,y=10)
  def edtpsfile_image1(event):
      edit_window_img = Toplevel()
      edit_window_img.title("View Image")
      edit_window_img.geometry("700x500")
      
      
      itemid = cusventtreeeddc.item(cusventtreeeddc.focus())["values"][0]
      sql = "select * from documents where orderid = %s"
      val = (osdata[0], )

      fbcursor.execute(sql, val)
      psdata1 = fbcursor.fetchone() 
      image = Image.open("images/"+psdata1[2])
      resize_image = image.resize((700, 500))
      image = ImageTk.PhotoImage(resize_image)
      psimage = Label(edit_window_img,image=image)
      psimage.photo = image
      psimage.pack()
  cusventtreeeddc=ttk.Treeview(documentFrame, height=5)
  cusventtreeeddc["columns"]=["1"]
  cusventtreeeddc.column("#0", width=20, anchor="center")
  cusventtreeeddc.column("1", width=650, anchor="center")
  # cusventtreeeddc.column("2", width=250, anchor="center")
  # cusventtreeeddc.column("2", width=200, anchor="center")
  cusventtreeeddc.heading("#0",text="", anchor=W)
  # cusventtreeeddc.heading("1",text="Attach to Email")
  cusventtreeeddc.heading("1",text="Filename")
  # cusventtreeeddc.heading("3",text="Filesize")  
  cusventtreeeddc.place(x=50, y=20)
  cusventtreeeddc.bind('<Double-Button-1>' , edtpsfile_image1)
  sqldoc='SELECT * FROM documents Where orderid=%s'
  vall=(osdata[0],)
  fbcursor.execute(sqldoc,vall)
  for record in cusventtreeeddc.get_children():
    cusventtreeeddc.delete(record) 
  j = 0
  for i in fbcursor:
    cusventtreeeddc.insert(parent='', index='end', iid=i, text='', values=(i[2]))
  j += 1
  
  global discount1,sub1,tax1,cost1z,order1
  # ,total1,balance1
  fir4Frame=Frame(pop,height=190,width=210,bg="#f5f3f2")
  fir4Frame.place(x=740,y=520)
  summaryfrme = LabelFrame(fir4Frame,text="Summary",font=("arial",15))
  summaryfrme.place(x=0,y=0,width=200,height=140)
  sl=orbtdata[13]
  s=round(sl, 4)
  pl=totalpriceinput
  p=round(pl, 4)
  dscte=(orbtdata[13]*p)/100
  dsc=round(dscte, 4)
  discount=Label(summaryfrme, text=str(s)+"% Discount").place(x=0 ,y=0)
  discount1=Label(summaryfrme,text='Rs. '+str(dsc))
  discount1.place(x=130 ,y=0)
  sbtotlhehe=p-dsc
  sbtotl=round(sbtotlhehe, 4)
  sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
  sub1=Label(summaryfrme, text='Rs. '+str(sbtotl)).place(x=130 ,y=21)
  taxval9=orbtdata[14]
  tax1dsply=(taxval9*sbtotl)/100
  taxval=round(tax1dsply,2)
  tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
  tax1=Label(summaryfrme, text='Rs. '+str(taxval)).place(x=130 ,y=42)
  xtracstvalzz=orbtdata[10]
  xtracst10=round(xtracstvalzz, 4)
  cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
  cost1z=Label(summaryfrme, text='Rs. '+str(orbtdata[10])).place(x=130 ,y=63)

  othrttlval=sbtotl+taxval+xtracstvalzz
  ov=round(othrttlval, 4)
  order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
  order1=Label(summaryfrme, text='Rs. '+str(ov)).place(x=130 ,y=84)
  # # totalpaid='10/-'
  # total=Label(summaryfrme, text="Total paid").place(x=0 ,y=105)
  # total1=Label(summaryfrme, text="0.0/-").place(x=130 ,y=105)
  # # paidtotal=str(totalpaid)
  # # balanz=othrttlval-paidtotal
  # balance=Label(summaryfrme, text="Balance").place(x=0 ,y=126)
  # balance1=Label(summaryfrme, text='Rs. '+str(ov)).place(x=130 ,y=126)

  fir5Frame=Frame(pop,height=38,width=210)
  fir5Frame.place(x=735,y=485)

  def recalcultn():
   global discount1,dscdd
  #  totalpriceinput.config(text="")
   ttlprzinpt=pricecol.cget("text")
   ratediscnt=dscntrte.get()
   xtax=taxx.get()
   cstxtra=xtracst.get()
   print(ttlprzinpt,ratediscnt,xtax,cstxtra)

   sdd=round(int(ratediscnt),3)
   pldd=ttlprzinpt
   p=round(pldd, 4)
   dsctedd=(sdd*pldd)/100
   dscdd=round(dsctedd, 4)
   discount=Label(summaryfrme, text=str(sdd)+"% Discount").place(x=0 ,y=0)
   discount1=Label(summaryfrme,text='Rs. '+str(dscdd))
   discount1.place(x=130 ,y=0)

   sbtotlhee=p-dscdd
   subtotl=round(sbtotlhee, 4)
   sub=Label(summaryfrme, text="Subtotal").place(x=0 ,y=21)
   sub100=Label(summaryfrme, text='Rs. '+str(subtotl))
   sub100.place(x=130 ,y=21)

   taxval91=int(xtax)
   tax1dsplay=(taxval91*subtotl)/100
   taxvalu=round(tax1dsplay,2)
   tax=Label(summaryfrme, text="Tax1").place(x=0 ,y=42)
   tax100=Label(summaryfrme, text='Rs. '+str(taxvalu))
   tax100.place(x=130 ,y=42)

   xtracstvaluzz=int(cstxtra)
   xtra1cst=round(xtracstvaluzz, 4)
   cost=Label(summaryfrme, text="Extra cost").place(x=0 ,y=63)
   cost100=Label(summaryfrme, text='Rs. '+str(xtra1cst)).place(x=130 ,y=63)
 
   otherttlval=subtotl+taxvalu+xtra1cst
   ot1v1=round(otherttlval, 4)
   order=Label(summaryfrme, text="Order total").place(x=0 ,y=84)
   order100=Label(summaryfrme, text='Rs. '+str(ot1v1)).place(x=130 ,y=84)

   itemidot = ordtree.item(ordtree.focus())["values"][1]
   ordttl=ot1v1
   for record in ordtree.get_children():
     ordtree.delete(record)
   sqlordttl = "UPDATE Orders SET sum_discount=%s,sum_tax=%s,sum_subtotal=%s,Order_total=%s WHERE orderid = %s"
   valqordttl = (dscdd,taxvalu,subtotl,ordttl,itemidot, )
   fbcursor.execute(sqlordttl, valqordttl,)
   fbilldb.commit()
   fbcursor.execute('SELECT * FROM orders')   
   j = 0
   for i in fbcursor:
    ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
   j += 1
      
  recalbtn=Button(fir5Frame, compound="left", text="Show Summary",command=recalcultn)
  recalbtn.place(x=75, y=0,height=35)
  # btnup=Button(fir5Frame, compound="left", text="Line Up").place(x=150, y=0)







   
  # ttk.Treeview.place(x=230, y=160)
# # -------------------------subtree for total------------------------
#   tv1 = ttk.Treeview(ws, style="mystyle.Treeview", height=5)
#   # ttk.Style().configure("Treeview", background="#ebebe0", foreground="#ebebe0")
#   tv1.insert(parent='', index=0, iid=0, values=("vineet", "e11", 1000000.00))
#   tv1.insert(parent='', index=1, iid=1, values=("anil", "e12", 120000.00))
#   tv1.insert(parent='', index=2, iid=2, values=("ankit", "e13", 41000.00))
#   tv1.insert(parent='', index=3, iid=3, values=("Shanti", "e14", 22000.00))
#   tv1.place(x=900, y=549)
 

  
  # scrollbar = Scrollbar(ws)
  # scrollbar.place(x=1339, y=3, height=725)
  # scrollbar.config( command=tree.yview )

  # ws.mainloop()
################################################################################################################################################  
# def printsele():
#       try:
#         # import pdb;pdb.set_trace()
#         # credit = round(float(data[0]['dblDebit'].split(' ')[0]),2)
#         # debit = round(float(data[-2]['dblDebit'].split(' ')[0]))

#         # for dat in data:
#         #     if len(dat) == 2:
#         #         continue
#         #     elif dat['datDate']=='Opening Balance':
#         #         continue
#         #     elif dat['datDate']=='Closing Balance':
#         #         continue
#         #     elif dat['dblDebit']!=0:
#         #         debit -= dat['dblDebit']
#         #     elif dat['dblCredit']!=0:
#         #         credit -= dat['dblCredit']
#         str_html = ""

#         str_html+="""
#                 <!doctype>
#                 <html>
#                     <head>
#                         <style>
#                             .header {
#                                 text-align:center;
#                             }
#                             .body {
#                                 font-size: 1.5rem;
#                                 font-weight:normal;
#                             }
#                             table {
#                                 width:100%
#                             }
#                             td{
#                                 text-align: center;
#                                 padding-top: 1%;
#                             }
#                             .thead .th{
#                                 border-top: 1px solid #333;
#                                 border-bottom: 1px solid #333;
#                                 text-align: center;
#                                 font-weight: bold;
#                             }
#                         </style>
#                     </head>


#                     <body>

#                         <div class="header">
#                             <h2>Ledger Statements</h2>
#                             <h3>"""+str(data[1]['strName'])+"""</h3>
#                             <h5>"""+data[-1]['datFrom']+""" to """+data[-1]['datTo']+"""</h5>
#                         </div>
                        
#                         <br>

#                         <div class="body">
#                             <table>
#                                 <tr class="thead">
#                                     <td class="th">Date</td>
#                                     <td class="th">Name</td>
#                                     <td class="th">Debit</td>
#                                     <td class="th">Credit</td>
#                                 </tr>

#                                 <tbody>"""
#         for dat in data:
#             if len(dat) == 2:
#                 continue
#             elif dat['datDate']=='Opening Balance':
#                 str_html+="""<tr>
#                         <td></td>
#                         <td>"""+dat['datDate']+"""</td>
#                         <td>"""+str(dat['dblDebit'])+"""</td>
#                         <td></td>
#                     </tr>"""
#             elif dat['datDate']=='Closing Balance':
#                 str_html+="""<tr>
#                         <td></td>
#                         <td>"""+dat['datDate']+"""</td>
#                         <td>"""+str(dat['dblDebit'])+"""</td>
#                         <td></td>
#                     </tr>"""
#             else:
#                 str_html+="""<tr>
#                         <td>"""+dat['datDate']+"""</td>
#                         <td>"""+dat['strName']+"""</td>
#                         <td>"""+str(dat['dblDebit'])+"""</td>
#                         <td>"""+str(dat['dblCredit'])+"""</td>
#                     </tr>"""

#         str_html+="""</tbody>
#                 </table>
#             </div>
#         </body>
#     </html>"""

#         return str_html

#     except Exception as e:
#         raise
  # def property1():
  #   propert=Toplevel()
  #   propert.title("Microsoft Print To PDF Advanced Document Settings")
  #   propert.geometry("670x500+240+150")

  #   def property2():
  #     propert1=Toplevel()
  #     propert1.title("Microsoft Print To PDF Advanced Document Settings")
  #     propert1.geometry("670x500+240+150")

  #     name=Label(propert1, text="Microsoft Print To PDF Advanced Document Settings").place(x=10, y=5)
  #     paper=Label(propert1, text="Paper/Output").place(x=30, y=35)
  #     size=Label(propert1, text="Paper size").place(x=55, y=65)
  #     n = StringVar()
  #     search = ttk.Combobox(propert1, width = 15, textvariable = n )
  #     search['values'] = ('letter')
  #     search.place(x=150,y=65)
  #     search.current(0)
  #     copy=Label(propert1, text="Copy count:").place(x=55, y=95)

  #     okbtn=Button(propert1,compound = LEFT,image=tick , text="Ok", width=60).place(x=460, y=450)
  #     canbtn=Button(propert1,compound = LEFT,image=cancel, text="Cancel", width=60).place(x=570, y=450)
      
      


  #   style = ttk.Style()
  #   style.theme_use('default')
  #   style.configure('TNotebook.Tab', background="#999999", padding=5)
  #   property_Notebook = ttk.Notebook(propert)
  #   property_Frame = Frame(property_Notebook, height=500, width=670)
  #   property_Notebook.add(property_Frame, text="Layout")
  #   property_Notebook.place(x=0, y=0)

  #   name=Label(property_Frame, text="Orientation:").place(x=10, y=5)
  #   n = StringVar()
  #   search = ttk.Combobox(property_Frame, width = 23, textvariable = n )
  #   search['values'] = ('Portrait')
  #   search.place(x=10,y=25)
  #   search.current(0)

  #   text=Text(property_Frame,width=50).place(x=250, y=5,height=350)

  #   btn=Button(property_Frame, text="Advanced",command=property2).place(x=550, y=380)
  #   btn=Button(property_Frame,compound = LEFT,image=tick  ,text="OK", width=60,).place(x=430, y=420)
  #   btn=Button(property_Frame,compound = LEFT,image=cancel , text="Cancel", width=60,).place(x=550, y=420)     


    
  # if(False):
  #     messagebox.showwarning("FBilling Revelution 2020", "Customer is required, Please select customer for this invoice\nbefore printing")
  # elif(False):
  #     messagebox.showinfo("FBilling Revelution 2020", "Print job has been completed.")
  # else:
  #     print1=Toplevel()
  #     print1.title("Print")
  #     print1.geometry("670x400+240+150")
      
  #     printerframe=LabelFrame(print1, text="Printer", height=80, width=650)
  #     printerframe.place(x=7, y=5)      
  #     name=Label(printerframe, text="Name:").place(x=10, y=5)
  #     e1= ttk.Combobox(printerframe, width=40).place(x=70, y=5)
  #     where=Label(printerframe, text="Where:").place(x=10, y=30)
  #     printocheckvar=IntVar()
  #     printochkbtn=Checkbutton(printerframe,text="Print to file",variable=printocheckvar,onvalue=1,offvalue=0,height=1,width=10)
  #     printochkbtn.place(x=450, y=30)
  #     btn=Button(printerframe, text="Properties", width=10,command=property1).place(x=540, y=5)

  #     pageslblframe=LabelFrame(print1, text="Pages", height=140, width=320)
  #     pageslblframe.place(x=10, y=90)
  #     radvar=IntVar()
  #     radioall=Radiobutton(pageslblframe, text="All", variable=radvar, value="1").place(x=10, y=5)
  #     radiocpage=Radiobutton(pageslblframe, text="Current Page", variable=radvar, value="2").place(x=10, y=25)
  #     radiopages=Radiobutton(pageslblframe, text="Pages: ", variable=radvar, value="3").place(x=10, y=45)
  #     pagecountentry = Entry(pageslblframe, width=23).place(x=80, y=47)
  #     pageinfolabl=Label(pageslblframe, text="Enter page numbers and/or page ranges\nseperated by commas. For example:1,3,5-12")
  #     pageinfolabl.place(x=5, y=75)

  #     copylblframe=LabelFrame(print1, text="Copies", height=140, width=320)
  #     copylblframe.place(x=335, y=90)
  #     nolabl=Label(copylblframe, text="Number of copies").place(x=5, y=5)      
  #     noentry = Entry(copylblframe, width=18).place(x=130, y=5)      
  #     one=Frame(copylblframe, width=30, height=40, bg="black").place(x=20, y=40)     
  #     two=Frame(copylblframe, width=30, height=40, bg="grey").place(x=15, y=45)     
  #     three=Frame(copylblframe, width=30, height=40, bg="white").place(x=10, y=50)      
  #     four=Frame(copylblframe, width=30, height=40, bg="black").place(x=80, y=40)      
  #     fiv=Frame(copylblframe, width=30, height=40, bg="grey").place(x=75, y=45)      
  #     six=Frame(copylblframe, width=30, height=40, bg="white").place(x=70, y=50)      
  #     collatecheckvar=IntVar()
  #     collatechkbtn=Checkbutton(copylblframe,text="Collate",variable=collatecheckvar,onvalue=1,offvalue=0,height=1,width=10)
  #     collatechkbtn.place(x=130, y=70)

  #     othrlblframe=LabelFrame(print1, text="Other", height=120, width=320)
  #     othrlblframe.place(x=10, y=235)
  #     printlb=Label(othrlblframe, text="Print").place(x=5, y=0)
  #     dropprint = ttk.Combobox(othrlblframe, width=23).place(x=80, y=0)
  #     orderlb=Label(othrlblframe, text="Order").place(x=5, y=25)
  #     dropord = ttk.Combobox(othrlblframe, width=23).place(x=80, y=25)
  #     duplexlb=Label(othrlblframe, text="Duplex").place(x=5, y=50)
  #     droplex = ttk.Combobox(othrlblframe, width=23).place(x=80, y=50)

  #     prmodelblframe=LabelFrame(print1, text="Print mode", height=120, width=320)
  #     prmodelblframe.place(x=335, y=235)
  #     dropscal = ttk.Combobox(prmodelblframe, width=30).place(x=5, y=5)
  #     poslb=Label(prmodelblframe, text="Print on sheet").place(x=5, y=35)
  #     droppos = ttk.Combobox(prmodelblframe, width=10).place(x=155, y=35)

  #     okbtn=Button(print1,compound = LEFT,image=tick , text="Ok", width=60).place(x=460, y=370)
  #     canbtn=Button(print1,compound = LEFT,image=cancel, text="Cancel", width=60).place(x=570, y=370)
      


#email

def send_mail(file=None):

  sender_email = email_from.get()    
  sender_password =email_pswrd.get() 

  server = smtplib.SMTP('smtp.gmail.com', 587)
  print("Login successful1")
  server.starttls()
  print("Login successful2")
  server.login(sender_email, sender_password)
  print("Login successful3")
  
  # global email_address, email_subject

  # email_address = StringVar()  
  # email_subject = StringVar()
  # msg = MIMEMultipart()

  # for i in htcodeframe.curselection():
  #   # print("hloo",htcodeframe.get(i))
  #   ya=htcodeframe.get(i)
  #   print(ya,"THIS")

  #   fo = open(ya, "rb")
  #   filecontent = fo.read()
  #   encodedcontentnw = base64.b64encode(filecontent) 


  # address_info = email_address.get()
  # print(address_info)
  # # msg['Subject'] = email_subject.get()
  # subject_info = email_subject.get()
  # print(subject_info)
  carbcopy_info = carcopyem_address.get()
  print(carbcopy_info)
  # email_content=mframe.get('1.0','end-1c')
  # print(email_content)
  
  
  msg = MIMEMultipart()
  msg['Subject'] = email_subject.get() 
  mail_content  = mframe.get('1.0','end-1c') 
  msg['From'] = email_from.get()
  msg['To'] = email_address.get()
  # msg.attach(MIMEText(file("text.txt").read()))

  # msg.attach(MIMEImage(open('images/'+filenamez.split('/')[-1],"rb").read()))
  # for i in htcodeframe.curselection():
  
  gettingimg=lstfrm.get()
  lst_data = gettingimg[1:-1].split(',')
  print(lst_data,"happy")
  # print(gettingimg)
  
  msg.attach(MIMEText(mail_content, 'plain'))
  
  for i in lst_data:
    if len(i.strip()[1:-1])>1:
    # print(i[0],"IMAGE")
      with open('images/'+ i.strip()[1:-1], "rb") as attachment:
          # MIME attachment is a binary file for that content type "application/octet-stream" is used
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
      # encode into base64 
        encoders.encode_base64(part) 
  
        part.add_header('Content-Disposition', "attachment; filename= %s" % 'images/'+ i.strip()[1:-1]) 

      # attach the instance 'part' to instance 'message' 
        msg.attach(part)
    # message_body = email_body.get()

  server.sendmail(email_from.get(),email_address.get(),msg.as_string())
  server.sendmail(email_from.get(), carbcopy_info,msg.as_string())


  dateeem=dt.datetime.now()
  emitemid = ordtree.item(ordtree.focus())["values"][1]
  for record in ordtree.get_children():
    ordtree.delete(record)
  sqlq = "UPDATE Orders SET emailed_on=%s WHERE orderid = %s"
  valq = (dateeem,emitemid, )
  fbcursor.execute(sqlq, valq,)
  fbilldb.commit()
  fbcursor.execute('SELECT * FROM Orders;')
  ordertotalinput=0
  j = 0
  for i in fbcursor:
   ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
   for line in ordtree.get_children():
    idsave1=ordtree.item(line)['values'][9]
   ordertotalinput += idsave1
   j += 1
  ordtotalrowcol.config(text=ordertotalinput)

  print("message sent")
  
  # subent.delete(0, END)
  # emailtoent.delete(0, END)

def empsfile_image(event):
       global yawn
       for i in htcodeframe.curselection():
        print("hloo",htcodeframe.get(i))
        yawn=htcodeframe.get(i)        
        edit_window_img = Toplevel()
        edit_window_img.title("View Image")
        edit_window_img.geometry("700x500")
        image = Image.open("images/"+yawn)
        resize_image = image.resize((700, 500))
        image = ImageTk.PhotoImage(resize_image)
        psimage = Label(edit_window_img,image=image)
        psimage.photo = image
        psimage.pack()

def UploadAction(event=None):
  global filenamez
  # filename = filedialog.askopenfilename()
  # print('Selected:', filename)
  # name = askopenfilename(filetypes=[('PDF', '*.pdf',)])

  filenamez = askopenfilename(filetypes=(("png file ",'.png'),("jpg file", ".jpg"), ('PDF', '*.pdf',), ("All files", "*.*"),))
  shutil.copyfile(filenamez, os.getcwd()+'/images/'+filenamez.split('/')[-1])
  htcodeframe.insert(0, filenamez.split('/')[-1])


   # add this
 
  # def upload_file1():
  #  global filename,img, b1
  #  f_types =[('Png files','.png'),('Jpg Files', '.jpg')]
  #  filename = filedialog.askopenfilename(filetypes=f_types)
  #  print(filename, 'name')
  #  #import pdb; pdb.set_trace()
  #  shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
  #  image = Image.open(filename)
  #  resize_image = image.resize((120, 120))
  #  img = ImageTk.PhotoImage(resize_image)
  #  b1 = Label(expenselabelframe,image=img, height=120, width=120)
  #  b1.place(x=450 , y=240)


###################DB
 #file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
  #query 
  #val=filename.split('/')[-1]
    
def email_order():
 try:
  emitid = ordtree.item(ordtree.focus())["values"][1]
  sql = "select * from Orders where orderid = %s"
  val = (emitid, )
  fbcursor.execute(sql, val)
  emailnow = fbcursor.fetchone()
  mailDetail=Toplevel()
  mailDetail.title("Order E-Mail")
  mailDetail.geometry("1080x550")
  mailDetail.resizable(False, False)
  # def my_SMTP():
  #     if True:
  #         em_ser_conbtn.destroy()
  #         mysmtpservercon=LabelFrame(account_Frame,text="SMTP server connection(ask your ISP for your SMTP settings)", height=165, width=380)
  #         mysmtpservercon.place(x=610, y=110)
  #         lbl_hostn=Label(mysmtpservercon, text="Hostname").place(x=5, y=10)
  #         hostnent=Entry(mysmtpservercon, width=30).place(x=80, y=10)
  #         lbl_portn=Label(mysmtpservercon, text="Port").place(x=5, y=35)
  #         portent=Entry(mysmtpservercon, width=30).place(x=80, y=35)
  #         lbl_usn=Label(mysmtpservercon, text="Username").place(x=5, y=60)
  #         unament=Entry(mysmtpservercon, width=30).place(x=80, y=60)
  #         lbl_pasn=Label(mysmtpservercon, text="Password").place(x=5, y=85)
  #         pwdent=Entry(mysmtpservercon, width=30).place(x=80, y=85)
  #         ssl_chkvar=IntVar()
  #         ssl_chkbtn=Checkbutton(mysmtpservercon, variable=ssl_chkvar, text="This server requires a secure connection(SSL)", onvalue=1, offvalue=0)
  #         ssl_chkbtn.place(x=50, y=110)
  #         em_ser_conbtn1=Button(account_Frame, text="Test E-mail Server Connection").place(x=610, y=285)
  #     else:
  #         pass

 
  style = ttk.Style()
  style.theme_use('default')
  style.configure('TNotebook.Tab', background="#999999", padding=5)
  email_Notebook = ttk.Notebook(mailDetail)
  email_Frame = Frame(email_Notebook, height=500, width=1080)
  account_Frame = Frame(email_Notebook, height=550, width=1080)
  email_Notebook.add(email_Frame, text="E-mail")
  email_Notebook.add(account_Frame, text="Account")
  email_Notebook.place(x=0, y=0)

  messagelbframe=LabelFrame(email_Frame,text="Message", height=500, width=730)
  messagelbframe.place(x=5, y=5)
  global email_address, email_subject, email_from,email_pswrd,carcopyem_address,mframe,htcodeframe,lstfrm,langs
  email_address = StringVar() 
  email_subject = StringVar()
  # email_body = StringVar()
  email_from = StringVar()
  email_pswrd = StringVar()
  carcopyem_address = StringVar()
  # content_email = StringVar()
  lbl_emailtoaddr=Label(messagelbframe, text="Email to address").place(x=5, y=5)
  emailtoent=Entry(messagelbframe, width=50,textvariable=email_address)
  emailtoent.place(x=120, y=5)
  emailtoent.delete(0,'end')
  emailtoent.insert(0, emailnow[19])
  #email2 = email_address.get()
  #print(email2)
  sendemail_btn=Button(messagelbframe, text="Send Email", width=10, height=1, command=send_mail).place(x=600, y=10)#,command=addacnt

  lbl_carcopyto=Label(messagelbframe, text="Carbon copy to").place(x=5, y=32)
  carcopyent=Entry(messagelbframe, width=50,textvariable=carcopyem_address)
  carcopyent.place(x=120, y=32)
  # stopemail_btn=Button(messagelbframe, text="Stop sending", width=10, height=1).place(x=600, y=40)
  lbl_subject=Label(messagelbframe, text="Subject").place(x=5, y=59)
  subent=Entry(messagelbframe, width=50, textvariable=email_subject)
  subent.place(x=120, y=59)
  subjectinsrt='ORD_'+str(emailnow[0])
  subent.delete(0,'end')
  subent.insert(0, subjectinsrt)

  
  style = ttk.Style()
  style.theme_use('default')
  style.configure('TNotebook.Tab', background="#999999", width=20, padding=5)
  mess_Notebook = ttk.Notebook(messagelbframe)
  emailmessage_Frame =Frame(mess_Notebook, height=350, width=710)
  htmlsourse_Frame = Frame(mess_Notebook, height=350, width=710)
  mess_Notebook.add(emailmessage_Frame, text="E-mail message")
  # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
  # mess_Notebook.place(x=120, y=59)
  mess_Notebook.add(htmlsourse_Frame, )#text="Html source code")
  mess_Notebook.place(x=5, y=90)

  # btn1=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

  
  # btn2=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
  # btn3=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
  # btn4=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
  # btn5=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=undo).place(x=140, y=1)
  # btn6=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=redo).place(x=175, y=1)
  # btn7=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=bold).place(x=210, y=1)
  # btn8=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=italics).place(x=245, y=1)
  # btn9=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=underline).place(x=280, y=1)
  # btn10=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=left).place(x=315, y=1)
  # btn11=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=right).place(x=350, y=1)
  # btn12=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=center).place(x=385, y=1)
  # btn13=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=hyperlink).place(x=420, y=1)
  
  # btn14=Button(emailmessage_Frame,width=31,height=23,compound = LEFT,image=remove).place(x=455, y=1)


  # dropcomp = ttk.Combobox(emailmessage_Frame, width=12, height=3).place(x=500, y=5)
  # dropcompo = ttk.Combobox(emailmessage_Frame, width=6, height=3).place(x=600, y=5)
  mframe=scrolledtext.Text(emailmessage_Frame,  undo=True,width=88, bg="white", height=22)
  mframe.place(x=0, y=10)
  # mess_Notebook=Entry(emailmessage_Frame, width=710,height=350, textvariable=email_subject)
  # mess_Notebook.place(x=120, y=59)


  # btn1=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=selectall).place(x=0, y=1)

  
  # btn2=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=cut).place(x=36, y=1)
  # btn3=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=copy).place(x=73, y=1)
  # btn4=Button(htmlsourse_Frame,width=31,height=23,compound = LEFT,image=paste).place(x=105, y=1)
  # mframe1=Frame(htmlsourse_Frame, height=350, width=710, bg="white")
  # mframe1.place(x=0, y=28)
  attachlbframe=LabelFrame(email_Frame,text="Attachment(s)", height=350, width=280)
  attachlbframe.place(x=740, y=5)
  # langs=[]
  lstfrm=StringVar()  
  htcodeframe=Listbox(attachlbframe, height=13, width=43,listvariable=lstfrm, bg="white")
  htcodeframe.place(x=5, y=5)
  htcodeframe.bind('<Double-Button-1>' , empsfile_image)

  def deslist():
      laa=htcodeframe.curselection()
      print("hloo",htcodeframe.get(laa))
      yawn=htcodeframe.get(laa)        
      htcodeframe.delete(ACTIVE)

  lbl_btn_info=Label(attachlbframe, text="Double click on attachment to view").place(x=30, y=230)
  btn17=Button(attachlbframe, width=20, text="Add attachment file...", command=UploadAction).place(x=60, y=260)
  btn18=Button(attachlbframe, width=20, text="Remove attachment",command=deslist).place(x=60, y=295)
  lbl_tt_info=Label(email_Frame, text="You can create predefined invoice, order, estimate\nand payment receipt email templates under Main\nmenu/Settings/E-Mail templates tab")
  lbl_tt_info.place(x=740, y=370)

  ready_frame=Frame(mailDetail, height=20, width=1080, bg="#b3b3b3").place(x=0,y=530)
  
  sendatalbframe=LabelFrame(account_Frame,text="E-Mail(Sender data)",height=140, width=600)
  sendatalbframe.place(x=5, y=5)
  lbl_sendermail=Label(sendatalbframe, text="Company email address").place(x=5, y=10)
  sentent=Entry(sendatalbframe, width=40, textvariable=email_from)
  sentent.place(x=195, y=10)
  #############################################
  lbl_senderpswrd=Label(sendatalbframe, text="Email Password").place(x=5, y=40)
  pswrdent=Entry(sendatalbframe, width=40, textvariable=email_pswrd,show="*")
  pswrdent.place(x=195, y=40)
  #############################################
  # lbl_orcompanyname=Label(sendatalbframe, text="Your name or company name").place(x=5, y=70)
  # nament=Entry(sendatalbframe, width=40).place(x=195, y=70)
  # lbl_reply=Label(sendatalbframe, text="Reply to email address").place(x=5, y=100)
  # replyent=Entry(sendatalbframe, width=40).place(x=195, y=100)
  # lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=140)
  # signent=Entry(sendatalbframe,width=50).place(x=100, y=140,height=75)
  # @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
  # lbl_sign=Label(sendatalbframe, text="Signature").place(x=5, y=95)
  # signent=Entry(sendatalbframe,width=50).place(x=195, y=70,height=95)
  # confirm_chkvar=IntVar()
  # confirm_chkbtn=Checkbutton(sendatalbframe, variable=confirm_chkvar, text="Confirmation reading", onvalue=1, offvalue=0)
  # confirm_chkbtn.place(x=200, y=215)
  # btn18=Button(account_Frame, width=15, text="Save settings",command=savesettings).place(x=25, y=285)

  # sendatalbframe=LabelFrame(account_Frame,text="SMTP Server",height=100, width=380)
  # sendatalbframe.place(x=610, y=5)
  # servar=IntVar()
  # SMTP_rbtn=Radiobutton(sendatalbframe, text="Use the Built-In SMTP Server Settings", variable=servar, value=1)
  # SMTP_rbtn.place(x=10, y=10)
  # MySMTP_rbtn=Radiobutton(sendatalbframe, text="Use My Own SMTP Server Settings(Recommended)", variable=servar, value=2, command=my_SMTP)
  # MySMTP_rbtn.place(x=10, y=40)
  # em_ser_conbtn=Button(account_Frame, text="Test E-mail Server Connection")
  # em_ser_conbtn.place(x=710, y=110)
 except:
   try:
     mailDetail.destroy()
   except:
     pass
     messagebox.showerror('F-Billing Revolution', 'You have to select a record.') 


#sms notification order
  
def sms():
  send_SMS=Toplevel()
  send_SMS.geometry("700x480+240+150")
  send_SMS.title("Send SMS notification")

  style = ttk.Style()
  style.theme_use('default')
  style.configure('TNotebook.Tab', background="#999999", padding=5)
  sms_Notebook = ttk.Notebook(send_SMS)
  SMS_Notification = Frame(sms_Notebook, height=470, width=700)
  SMS_Service_Account = Frame(sms_Notebook, height=470, width=700)
  sms_Notebook.add(SMS_Notification, text="SMS Notification")
  sms_Notebook.add(SMS_Service_Account, text="SMS Service Account")
  sms_Notebook.place(x=0, y=0)

  numlbel=Label(SMS_Notification, text="SMS number or comma seperated SMS number list(Please start each SMS number with the country code)")
  numlbel.place(x=10, y=10)
  numentry=Entry(SMS_Notification, width=92).place(x=10, y=30)
  stexbel=Label(SMS_Notification, text="SMS Text").place(x=10, y=60)
  stex=Entry(SMS_Notification, width=40).place(x=10, y=85,height=120)
  
  dclbel=Label(SMS_Notification, text="Double click to insert into text")
  dclbel.place(x=410, y=60)
  dcl=Entry(SMS_Notification, width=30)
  dcl.place(x=400, y=85,height=200)
  
  smstype=LabelFrame(SMS_Notification, text="SMS message type", width=377, height=60)
  smstype.place(x=10, y=223)
  snuvar=IntVar()
  normal_rbtn=Radiobutton(smstype, text="Normal SMS(160 chars)", variable=snuvar, value=1)
  normal_rbtn.place(x=5, y=5)
  unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)", variable=snuvar, value=2)
  unicode_rbtn.place(x=190, y=5)
  tiplbf=LabelFrame(SMS_Notification, text="Tips", width=680, height=120)
  tiplbf.place(x=10, y=290)
  tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="Always start the SMS nymber with the country code. Do not use the + sign at the beginning(example\nUS number:8455807546). Do not use any special characters in your normal SMS text. Please use the\nstndard SMS characters or the English alphabet and numbers only. Otherwise the SMS will be\nunreadable or undeliverable. If you need to enter international characters, accents,email address, or\nspecial characters to the SMS text field then choose the Unicode SMS format.")
  tiplabl.place(x=5, y=5)

  btn1=Button(SMS_Notification, width=20, text="Send SMS notification").place(x=10, y=420)
  btn2=Button(SMS_Notification, width=25, text="Confirm SMS cost before sending").place(x=280, y=420)
  btn3=Button(SMS_Notification, width=15, text="Cancel").place(x=550, y=420)
  

  smstype=LabelFrame(SMS_Service_Account, text="Select the notification service provider", width=670, height=65)
  smstype.place(x=10, y=5)
  snumvar=IntVar()
  normal_rbtn=Radiobutton(smstype,text="BULKSMS(www.bulksms.com)",variable=snumvar,value=1,)
  normal_rbtn.place(x=5, y=5)
  unicode_rbtn=Radiobutton(smstype, text="Unicode SMS(70 chars)-Recommended", variable=snumvar, value=2)
  unicode_rbtn.place(x=290, y=5)

  sms1type=LabelFrame(SMS_Service_Account, text="Your BULKSMS.COM Account", width=670, height=100)
  sms1type.place(x=10, y=80)
  name=Label(sms1type, text="Username").place(x=10, y=5)
  na=Entry(sms1type, width=20).place(x=100, y=5)
  password=Label(sms1type, text="Password").place(x=10, y=45)
  pas=Entry(sms1type, width=20).place(x=100, y=45)
  combo=Label(sms1type, text="Route").place(x=400, y=5)
  n = StringVar()
  combo1 = ttk.Combobox(sms1type, width = 20, textvariable = n ).place(x=450,y=5)
  btn1=Button(sms1type, width=10, text="Save settings").place(x=550, y=45)

  
  tiplbf=LabelFrame(SMS_Service_Account, text="Terms of service", width=680, height=250)
  tiplbf.place(x=10, y=190)
  tiplabl=Label(tiplbf,justify=LEFT,fg="red",  text="The SMS notification service is not free.This service costs you creadit.You must have your own account\nat BULKSMS.COM and you need to have sufficient creadit and an active internet connection to use\nthis feature.Please review all fields in this form for accuracy")
  tiplabl.place(x=0, y=5)
  tiplabl1=Label(tiplbf,justify=LEFT,fg="black",  text="visit www.bulksms.com website to create your own account.please make sure the BULKSMS .COM\n service works well in your country before you busy creadit")
  tiplabl1.place(x=0, y=60)
  tiplabl2=Label(tiplbf,justify=LEFT,fg="black",  text="Our SMS notification tool comes without any warranty.our software only forwards your SMS message\nthe BULKSMS API server .The BULKSMS API server will try to sent SMS message your recipient")
  tiplabl2.place(x=0, y=100)
  tiplabl3=Label(tiplbf,justify=LEFT,fg="red",  text="Please note that you access and use the SMS notification tool your own risk.F-Billing software is not\nresponsible for any type of loss or damage or undelivered SMS massage which you may as a result\nof accessing and using the SMS notification service.")
  tiplabl3.place(x=0, y=140)
  checkvar1=IntVar()
  chkbtn1=Checkbutton(tiplbf,text="I have read and agree to the terms of service above",variable=checkvar1,onvalue=1,offvalue=0).place(x=70, y=200)  



#print preview order
def printpreview():
  messagebox.showerror("F-Billing Revolution","Customer is required,please select customer for this order before printing.")



#convert to invoice
def convertinv():
  global convert
  convert=messagebox.askyesno("Make invoice from Orders", "Are you sure to make invoice from this Orders ")
  if convert == True:
    convertid = ordtree.item(ordtree.focus())["values"][1]
    # print (convertid)
    # sql = 'INSERT INTO invoice WHERE invoiceid=%s'
    
    # for line in ordtree.get_children():
    # idsave1 = ordtree.item(ordtree.focus())["values"][1]
    # sql1= 'SELECT * FROM  storingproduct WHERE orderid = %s'
    # val=(idsave1,)
    # print(val)
    # fbcursor.execute(sql1,val,)
    # child=fbcursor.fetchone()
    # print(child)
    # sql2= 'INSERT INTO storingproduct(Productserviceid,orderid,sku,category,name,description,status,unitprice,peices,cost,taxable,priceminuscost,serviceornot,stock,stocklimit,warehouse,privatenote) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
    # val1=(child[0],idsave1,child[2],child[3],child[4],child[5],child[6],child[7],child[8],child[9],child[10],child[11],child[12],child[13],child[14],child[15],child[16])
    # fbcursor.execute(sql2,val1,)
    # fbilldb.commit()
    
    orddta= "SELECT orderid,businessname,businessaddress,shipname,shipaddress,cpemail,cpmobileforsms FROM Orders WHERE orderid=%s"
    val=(convertid,)
    fbcursor.execute(orddta,val,)
    abc=fbcursor.fetchone()
    print (abc)
  
    sql='INSERT INTO invoice(orderid,businessname,businessaddress,shipname,shipaddress,cpemail,cpmobileforsms) VALUES(%s,%s,%s,%s,%s,%s,%s)' #adding values into db
    val1=(abc[0],abc[1],abc[2],abc[3],abc[4],abc[5],abc[6])
    fbcursor.execute(sql,val1,)
    fbilldb.commit()

# .....>>>>> looking for invoiced/draft <<<<<.....

    itemid = ordtree.item(ordtree.focus())["values"][1]
    invoz="Invoiced"
    for record in ordtree.get_children():
      ordtree.delete(record)
    sqlq = "UPDATE Orders SET status=%s WHERE orderid = %s"
    # 'UPDATE storingproduct SET Productserviceid=%s
    valq = (invoz,itemid, )
    fbcursor.execute(sqlq, valq,)
    fbilldb.commit()

    fbcursor.execute('SELECT * FROM orders')   
    j = 0
    for i in fbcursor:
     ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    j += 1


    # messagebox.showinfo("Make invoice from Estimate", "Invoice Creation was Successfull.")
  # else:
  #     messagebox.destroy()
  

#delete orders  
# def dele():  
#   messagebox.askyesno("Delete order", "Are you sure to delete this order? All products will be placed back into stock")

#delete orders  
def dele():
  delmess = messagebox.askyesno("Delete Order", "Are you sure to delete this Order?")
  if delmess == True:
    itemid = ordtree.item(ordtree.focus())["values"][1]
    # print(itemid,)
    sql = 'DELETE FROM Orders WHERE orderid=%s'
    val = (itemid,)
    fbcursor.execute(sql, val)
    fbilldb.commit()

    ordtree.delete(ordtree.selection()[0])
    sqlstrngprdct='DELETE FROM storingproduct WHERE orderid=%s'
    valstpr = (itemid,)
    fbcursor.execute(sqlstrngprdct, valstpr)
    fbilldb.commit()

    sqldoc='DELETE FROM documents WHERE orderid=%s'
    valdoc = (itemid,)
    fbcursor.execute(sqldoc, valdoc)
    fbilldb.commit()

    notepriv.delete('1.0','end')
    for record in ordtree.get_children():
      ordtree.delete(record)
    for record in treeblw.get_children():
      treeblw.delete(record)
    sql = "SELECT * FROM Orders"
    fbcursor.execute(sql,)
    
    a=0
    j = 0
    for i in fbcursor:
      ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
      for line in ordtree.get_children():
          idsave=ordtree.item(line)['values'][9]
      a+=idsave
    j += 1
    ordtotalrowcol.config(text=a)
  else:
    pass

# #delete p/s in view/edit  
# def dele():
#   delmess = messagebox.askyesno("Delete Productservice", "Are you sure to delete this Productservice?")
#   if delmess == True:
#     itemid = prstree.item(prstree.focus())["values"][1]
#     print(itemid,)
#     # sql = 'DELETE FROM Productservice WHERE Productserviceid=%s'
#     val = (itemid,)
#     fbcursor.execute(val)
#     fbilldb.commit()
#     prstree.delete(prstree.selection()[0])
#   else:
#     pass


# def orderto():
#     sql = 'SELECT businessname FROM Customer'
#     Ordto=fbcursor.execute(sql,)
#     print(Ordto)
#     fbilldb.commit()
#     return(Ordto)

def ordsearch():
   query = searchord.get()
   selections = []
   for child in ordtree.get_children():
      if query in ordtree.item(child)['values']:
          print(ordtree.item(child)['values'])
          selections.append(child)
   ordtree.selection_set(selections)
def searchclose():
   top.destroy()
#search in orders  
def search():  
    global searchord,top
    top = Toplevel()     
    top.title("Find Text")   
    top.geometry("400x200")
    findwhat1=Label(top,text="Find Customer Name :",pady=5,padx=10).place(x=25,y=20)
    searchord = StringVar()
    findwhat = Entry(top, width = 40, text = searchord ).place(x=90,y=65)
   
    # findin1=Label(top,text="Find in:",pady=5,padx=10).place(x=5,y=47)
    # n = StringVar()
    # findIN = ttk.Combobox(top, width = 30, textvariable = n )
    # findIN['values'] = ('Product/Service id', ' Category', ' Active',' name',' stock',' location', ' image',' <<All>>')                       
    # findIN.place(x=90,y=54)
    # findIN.current(0)

    findButton = Button(top, text ="Find",width=10,command=ordsearch)
    findButton.place(x=70,y=130)
    closeButton = Button(top,text ="Close",width=10,command=searchclose).place(x=276,y=130)
    
    # match1=Label(top,text="Match:",pady=5,padx=10).place(x=5,y=74)
    # n = StringVar()
    # match = ttk.Combobox(top, width = 23, textvariable = n )   
    # match['values'] = ('From Any part',' Whole Field',' From the beginning of the field')                                   
    # match.place(x=90,y=83)
    # match.current(0)

    # search1=Label(top,text="Search:",pady=5,padx=10).place(x=5,y=102)
    # n = StringVar()
    # search = ttk.Combobox(top, width = 23, textvariable = n )
    # search['values'] = ('All', 'up',' Down')
    # search.place(x=90,y=112)
    # search.current(0)
    # checkvarStatus4=IntVar()  
    # Button4 = Checkbutton(top,variable = checkvarStatus4,text="Match Case",onvalue =0 ,offvalue = 1,height=3,width = 15)
    # Button4.place(x=90,y=141)
    # checkvarStatus5=IntVar()   
    # Button5 = Checkbutton(top,variable = checkvarStatus5,text="Match Format",onvalue =0 ,offvalue = 1,height=3,width = 15)
    # Button5.place(x=300,y=141)


# global printsele1
def printsele():
    printordinv = messagebox.askyesno("Download to Print", "Continue to Download?")
    if printordinv == True:
      itemid = ordtree.item(ordtree.focus())["values"][1]
      # print(itemid,)
      sql = 'SELECT * FROM Orders WHERE orderid=%s'
      val = (itemid,)
      fbcursor.execute(sql, val)
      koko=fbcursor.fetchone()


      # itemid = ordtree.item(ordtree.focus())["values"][1]
      dateee=dt.datetime.now()
      for record in ordtree.get_children():
       ordtree.delete(record)
      sqlq = "UPDATE Orders SET printed_on=%s WHERE orderid = %s"
    # 'UPDATE storingproduct SET Productserviceid=%s
      valq = (dateee,itemid, )
      fbcursor.execute(sqlq, valq,)
      fbilldb.commit()
      fbcursor.execute('SELECT * FROM Orders;')
      ordertotalinput=0
      j = 0
      for i in fbcursor:
       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
       for line in ordtree.get_children():
        idsave1=ordtree.item(line)['values'][9]
       ordertotalinput += idsave1
       j += 1
      ordtotalrowcol.config(text=ordertotalinput)


      # # fbilldb.commit()
      # selected = ordtree.focus()
      # selected_prdct= ordtree.item(selected)["values"][1]
    
      sqlprint='SELECT * FROM storingproduct WHERE orderid=%s'
      valprint=(itemid,)
      fbcursor.execute(sqlprint,valprint)
      mama=fbcursor.fetchall()
      # ordtree.delete(ordtree.selection()[0])
      # sqlstrngprdct='SELECT * FROM storingproduct WHERE orderid=%s'
      # valstpr = (itemid,)
      # fbcursor.execute(sqlstrngprdct, valstpr)
      # mama=fbcursor.fetchone()
      # fbilldb.commit()
      # print("HELLO",koko)
      # print("HAI",mama)
      # prdisc=discount1.cget("text")
    str_html = ""

    str_html+="""
                    <!doctype>
                    <html>
                        <head>
                            <style>
                                .header {
                                    text-align:center;
                                }
                                .body {
                                    font-size: 1.5rem;
                                    font-weight:normal;
                                }
                                table {
                                    
                                    width:100%
                                }
                                td{
                                    text-align: center;
                                    padding-top: 1%;
                                }
                                .thead .th{
                                    border-top: 1px solid #333;
                                    border-bottom: 1px solid #333;
                                    text-align: center;
                                    font-weight: bold;
                                }
                            </style>
                        </head>


                        <body>

                            <div class="header">
                                <h2>Order Invoice</h2>
                            </div>
                            <br><br><br><br><br><br><br>
                            <div>
                                <h3 >Order Id &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[0])+"""</h3>      

                                <h3 >Order Date &nbsp;&nbsp;"""+str(koko[1])+"""</h3>                             

                                <h3 >Due Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"""+str(koko[2])+"""</h3>                             
                            </div>                                                        
                            <br><br>
                            <div style="display:flex;">
                                <div>
                                <h2>Order To</h2><h3>"""+str(koko[3])+"""<br>"""+str(koko[16])+"""</h3>
                                </div>
                                <div>
                                <h2>Ship To</h2><h3>"""+str(koko[17])+"""<br>"""+str(koko[18])+"""</h3>                             
                                </div>                       
                            </div>
                            

                            <div class="body">

                                <table>
                                    <tr class="thead" ; style="background-color:orange">
                                        <td class="th">ID/SKU</td>
                                        <td class="th">Product/Service</td>
                                        <td class="th">Description</td>
                                        <td class="th">Quantity</td>
                                        <td class="th">Unit Price</td>
                                        <td class="th">Price</td>
                                    </tr>

                                    <tbody>"""

    for pri in mama:
     str_html+="""<tr>

                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[4])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[6])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[7])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+str(pri[22])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(pri[9])+"""</td>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str((pri[22])*(pri[9]))+"""</td>
 
                        </tr>"""

    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th style="border: 1px solid black;border-collapse: collapse;">"""+str(koko[13])+'%  Discount'+"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[24])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Subtotal'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[26])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'TAX1'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[25])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th  style="border: 1px solid black;border-collapse: collapse;">"""'Extra Cost'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[10])+"""</td>
                        </tr>"""
    str_html+="""<tr>
                       
                            <td></td>
                            <td></td>
                            <td></td>
                            <td></td>
                            <th style="border: 1px solid black;border-collapse: collapse;">"""'Order Total'"""</th>
                            <td style="border: 1px solid black;border-collapse: collapse;">"""+'Rs. '+str(koko[8])+"""</td>
                        </tr>"""
  
    str_html+="""</tbody>
                    </table>
                </div>
            </body>
        </html>"""

    
    # import pdb;pdb.set_trace()


    html_data = str_html

    # file_path = filedialog.asksaveasfile(initialdir=os.getcwd,title="Save File",filetypes=[('PDF', '*.pdf',)])
    file_path=os.getcwd()+'/images/'

    options = {'margin-top': '15.00mm',
                    'margin-right': '15.00mm',
                    'margin-bottom': '15.00mm',
                    'margin-left': '15.00mm',
                    'dpi':300,
                    }

    lst_pdfs = []
    str_file_name ='ORD_'+str(koko[0])+'.pdf'
    filename_ledger =  file_path+'/'+str_file_name
    path_wkthmltopdf = b'C:\Program Files\wkhtmltopdf\\bin\wkhtmltopdf.exe'
    config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)
    pdfkit.from_string(html_data,filename_ledger,options=options,configuration=config)
    lst_pdfs.append(filename_ledger)

    pdf_writer = PdfFileWriter()



mainFrameord=Frame(tab2, relief=GROOVE, bg="#f8f8f2")
mainFrameord.pack(side="top", fill=BOTH)

midFrame=Frame(mainFrameord, bg="#f5f3f2", height=60)
midFrame.pack(side="top", fill=X)

w = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=(5, 2))
w = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=(0, 5))

invoiceLabel = Button(midFrame,compound="top", text="Create new\nOrder",relief=RAISED, image=photo,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,command=create)
invoiceLabel.pack(side="left", pady=3, ipadx=4)

orderLabel = Button(midFrame,compound="top", text="View/Edit\nOrders",relief=RAISED, image=photo1,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=edit)
orderLabel.pack(side="left")

estimateLabel = Button(midFrame,compound="top", text="Delete\nSelected",relief=RAISED, image=photo2,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=dele)
estimateLabel.pack(side="left")

w = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=5)

recurLabel = Button(midFrame,compound="top", text="Convert to\nInvoice",relief=RAISED, image=photo3,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=convertinv)
recurLabel.pack(side="left")

w = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=5)

# previewLabel = Button(midFrame,compound="top", text="Print\nPreview",relief=RAISED, image=photo5,bg="#f8f8f2", fg="black", height=55, bd=1, width=55, activebackground="red",command=printpreview)
# previewLabel.pack(side="left")

purchaseLabel = Button(midFrame,compound="top", text="Print\nSelected",relief=RAISED, image=photo4,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=printsele)
purchaseLabel.pack(side="left")

w = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=5)

expenseLabel = Button(midFrame,compound="top", text="E-mail\nOrder",relief=RAISED, image=photo6,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=email_order)
expenseLabel.pack(side="left")

smsLabel = Button(midFrame,compound="top", text="Send SMS\nnotification",relief=RAISED, image=photo10,bg="#f8f8f2", fg="black", height=55, bd=1, width=62,command=sms)
smsLabel.pack(side="left")

w = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
w.pack(side="left", padx=5)

productLabel = Button(midFrame,compound="top", text="Search\nOrders",relief=RAISED, image=photo7,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=search)
productLabel.pack(side="left")


# def date_range(start, end):
#     delta = end - start  # as timedelta
#     days = [start + timedelta(days=i) for i in range(delta.days + 1)]
#     return days

lbframe = LabelFrame(midFrame, height=60, width=200, bg="#f8f8f2")
lbframe.pack(side="left", padx=10, pady=0)
lbl_invdt = Label(lbframe, text="Order date from : ", bg="#f8f8f2")
lbl_invdt.grid(row=0, column=0, pady=5, padx=(5, 0))
lbl_invdtt = Label(lbframe, text="Order date to  :  ", bg="#f8f8f2")
lbl_invdtt.grid(row=1, column=0, pady=5, padx=(5, 0))


def date_range(): # Start and stop dates for range
  var1=start_dte.get_date().strftime('%y/%m/%d')
  var2=stop_dte.get_date().strftime('%y/%m/%d')
  for record in ordtree.get_children():
    ordtree.delete(record)

  sqldate='SELECT * FROM Orders WHERE order_date BETWEEN %s AND %s'
  valuz=(var1,var2,)
  fbcursor.execute(sqldate,valuz)
  filterdate=fbcursor.fetchall()
  dt=0
  countp = 0
  for i in filterdate:
    ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    for line in ordtree.get_children():
      idsavedt=ordtree.item(line)['values'][9]
    dt+=idsavedt
  countp += 1
  ordtotalrowcol.config(text=dt)
  for record in prstree.get_children():
    prstree.delete(record)
  notepriv.delete('1.0','end')
  for record in treeblw.get_children():
    treeblw.delete(record)
  
start_dte = DateEntry(lbframe, width=15)
start_dte.grid(row=0, column=1)
stop_dte = DateEntry(lbframe, width=15)
stop_dte.grid(row=1, column=1)
checkvar1 = IntVar()
chkbtn1 = Button(lbframe, text = "Apply\nFilter",height = 2, width = 8,command=date_range,bg="#f8f8f2")
chkbtn1.grid(row=0, column=2, rowspan=2, padx=(5,5))

def refreshbttn():
  for record in ordtree.get_children():
    ordtree.delete(record)
  fbcursor.execute("select * from Orders")
  pandsdata = fbcursor.fetchall()
  rfrs=0
  countp = 0
  for i in pandsdata:
    ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    for line in ordtree.get_children():
     idsaverefrsh=ordtree.item(line)['values'][9]
    rfrs+=idsaverefrsh
  countp += 1
  ordtotalrowcol.config(text=rfrs)
  dropordr.delete(0,END)
  notepriv.delete('1.0','end')
  for record in prstree.get_children():
    prstree.delete(record)
  for record in treeblw.get_children():
    treeblw.delete(record)


productLabel = Button(midFrame,compound="top", text="Refresh\nOrders list",relief=RAISED, image=photo8,fg="black", height=55, bd=1, width=55,command=refreshbttn)
productLabel.pack(side="left")

# w = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
# w.pack(side="left", padx=5)

# global showhide
# showhide=BooleanVar()
# productLabel = Button(midFrame,compound="top", text="Hide totals\nSum",relief=RAISED, image=photo9,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=lambda: label_hide_show('Hide totals\nSum'))
# productLabel.pack(side="left")
# productLabel = Button(midFrame,compound="top", text="Show totals\nSum",relief=RAISED, image=photo9,bg="#f8f8f2", fg="black", height=55, bd=1, width=55,command=lambda: label_show_show('Show totals\nSum'))
# productLabel.pack(side="left")

invoilabel = Label(mainFrameord, text="Orders(All)", font=("arial", 18), bg="#f8f8f2")
invoilabel.pack(side="left", padx=(20,0))

# def ftr(event):
#   filt = dropordr.get()
#   for record in ordtree.get_children():
#     ordtree.delete(record) 
#   sqlmo = "SELECT * from Orders"
#   # val = (filt, )
#   fbcursor.execute(sqlmo, )
#   records = fbcursor.fetchall()
  
#   count=0
#   for i in records:
#     if True:
#       ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
#     else:
#       pass
#   count += 1

# sqlcat = "SELECT businessname FROM Orders"
# fbcursor.execute(sqlcat,)
# rec = fbcursor.fetchall()
# dropordr = ttk.Combobox(mainFrameord,)
# dropordr['values'] = rec
# dropordr.pack(side="right", padx=(0,10))

# dropordr.bind("<<ComboboxSelected>>", ftr)

def ftr(event):
  filt = dropordr.get()
  for record in ordtree.get_children():
    ordtree.delete(record)
    
  sql = "select * from orders where category = %s"
  val = (filt, )
  fbcursor.execute(sql, val)
  records = fbcursor.fetchall()
  catftr=0
  count=0
  for i in records:
    if True:
      ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
      for line in ordtree.get_children():
       idsavecatf=ordtree.item(line)['values'][9]
      catftr+=idsavecatf
    else:
      pass
  count += 1
  ordtotalrowcol.config(text=catftr)
  for record in prstree.get_children():
    prstree.delete(record)
  notepriv.delete('1.0','end')
  for record in treeblw.get_children():
    treeblw.delete(record)

sql = "SELECT DISTINCT category FROM orders"
fbcursor.execute(sql,)
rec = fbcursor.fetchall()

dropordr = ttk.Combobox(mainFrameord, value=" ")
dropordr.pack(side="right", padx=(0,10))
dropordr['values'] = rec
dropordr.bind("<<ComboboxSelected>>", ftr)
ordllabel = Label(mainFrameord, text="Category Filter ", font=("arial", 15), bg="#f8f8f2")
ordllabel.pack(side="right", padx=(0,10))

# class MyApp:
#   def __init__(self, parent):
    
#     self.myParent = parent 

#     self.myContainer1 = Frame(parent) 
#     self.myContainer1.pack()
    
#     self.top_frame = Frame(self.myContainer1) 
#     self.top_frame.pack(side=TOP,
#       fill=BOTH, 
#       expand=YES,
#       )  

#     self.left_frame = Frame(self.top_frame, background="white",
#       borderwidth=5,  relief=RIDGE,
#       height=250, 
#       width=2000, 
#       )
#     self.left_frame.pack(side=LEFT,
#       fill=BOTH, 
#       expand=YES,
#       )

s = ttk.Style()
s.configure('Treeview.Heading', background='white', foreground='dark blue', State='DISABLE')


ordtree=ttk.Treeview(tab2,selectmode='browse')
ordtree.pack(side = 'top')
ordtree["columns"]=("1","2","3","4","5","6","7","8","9","10")
ordtree["show"]='headings'
ordtree["height"]='15'
ordtree.heading(1)
ordtree.heading(2, text="Order#")
ordtree.heading(3, text="Order date")
ordtree.heading(4, text="Due date")
ordtree.heading(5, text="Customer Name")
ordtree.heading(6, text="Status")
ordtree.heading(7, text="Emailed on")
ordtree.heading(8, text="Printed on")
ordtree.heading(9, text="SMS on")
ordtree.heading(10, text="Order Total")   
ordtree.column(1, width = 50, anchor="center")
ordtree.column(2, width = 140, anchor="center")
ordtree.column(3, width = 140, anchor="center")
ordtree.column(4, width = 140, anchor="center")
ordtree.column(5, width = 210, anchor="center")
ordtree.column(6, width = 130, anchor="center")
ordtree.column(7, width = 150, anchor="center")
ordtree.column(8, width = 130, anchor="center")
ordtree.column(9, width = 130, anchor="center")
ordtree.column(10, width = 130, anchor="center")
ordtotalrowcol = Label(tab2,bg="#f5f3f2")
ordtotalrowcol.place(x=1260,y=400,width=80,height=18)
hidden=True
def label_hide_show(text):
    global hidden
    if hidden:
      
        # show=0
        # j = 0
        # for i in fbcursor:
        #   ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
        #   for line in ordtree.get_children():
        #     idsave=ordtree.item(line)['values'][9]
        #   show+=idsave
        # j += 1
        ordtotalrowcol.config(text=text)
        # ordtotalrowcol.config(text=ordtotalrowcol)
        # ordtotalrowcol.place(x=1260,y=400,width=80,height=18)
        ordtotalrowcol.pack(side = "top", fill = tk.BOTH)
        # ordtotalrowcol.pack()
        hidden = False
    else:
        ordtotalrowcol.pack_forget()
        hidden=True

# def label_show_show(text):
#       ordtotalrowcol.pack()

# #btn1 = tk.Button(root, text="Get Started", height=3,width=26,bg="White", fg="Black", command=lambda: label_hide_show('Get started'))
def prdctpicker(event):

        selected = ordtree.focus()
        selected_prdct= ordtree.item(selected)["values"][1]
        for record in prstree.get_children():
         prstree.delete(record)
        sqlprdct='SELECT * FROM storingproduct WHERE orderid=%s'
        valprdct=(selected_prdct,)
        fbcursor.execute(sqlprdct,valprdct)
        prdctsltn=fbcursor.fetchall()
        j = 0
        for i in prdctsltn:
         prstree.insert(parent='', index='end', iid=i, text='', values=(' ',i[21], i[6], i[7],i[9],i[22],i[12],(i[9]*i[22])))
        j += 1

        notefchz="SELECT * FROM orders WHERE orderid=%s"
        fchz = (selected_prdct,)
        fbcursor.execute(notefchz, fchz)
        zoomb=fbcursor.fetchone()   
        notepriv.delete('1.0','end')
        notepriv.insert('1.0', zoomb[27])

        sqldoc='SELECT * FROM documents Where orderid=%s'
        vall=(selected_prdct,)
        fbcursor.execute(sqldoc,vall)
        for record in treeblw.get_children():
            treeblw.delete(record) 
        j = 0
        for i in fbcursor:
         treeblw.insert(parent='', index='end', iid=i, text='', values=(' ',i[2]))
        j += 1

ordtree.bind('<Double-Button-1>' , prdctpicker)

  # ba=osdata[0]
  # sql4='SELECT * FROM storingproduct WHERE orderid=%s'
  # val10=(ba,)
  # fbcursor.execute(sql4,val10)
  # totalpriceinput=0
  # j = 0
  # for i in fbcursor:
  #   vwedttree1.insert(parent='', index='end', iid=i, text='', values=(i[4],i[6],i[7],i[9],i[22],i[10],i[12],(i[9]*i[22])))
  #   # for line in vwedttree1.get_children():
  #   #   idsave1=vwedttree1.item(line)['values'][7]
  #   totalpriceinput+=(i[9]*i[22])
  #   j += 1
  # pricecol.config(text=totalpriceinput)

fbcursor.execute('SELECT * FROM Orders;')
ordertotalinput=0
j = 0
for i in fbcursor:
    ordtree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0], i[1], i[2], i[3], i[4],i[5], i[6], i[7], i[8], i[9], i[10]))
    for line in ordtree.get_children():
     idsave1=ordtree.item(line)['values'][9]
    ordertotalinput += idsave1
    j += 1
ordtotalrowcol.config(text=ordertotalinput)

scrollbar = Scrollbar(tab2)
scrollbar.place(x=1016+300+25, y=120, height=280+20)

# inverticalbar=ttk.Scrollbar()
# inverticalbar.place(x=995+350,y=143,height=300+20)
tabControl = ttk.Notebook(tab2)
tab1 = ttk.Frame(tabControl)
tab2 = ttk.Frame(tabControl)
tab3 =  ttk.Frame(tabControl)
tab4 = ttk.Frame(tabControl)
tabControl.add(tab1,image=invoices,compound = LEFT, text ='Order Items',)
tabControl.add(tab2,image=orders,compound = LEFT, text ='Private Notes')
tabControl.add(tab3,image=estimates,compound = LEFT, text ='SMS Log')
tabControl.add(tab4,image=estimates,compound = LEFT, text ='Documents')
tabControl.pack(expand = 1, fill ="both")

prstree = ttk.Treeview(tab1, columns = (1,2,3,4,5,6,7,8,), height = 15, show = "headings")
prstree.pack(side = 'top')
prstree.heading(1)
prstree.heading(2, text="Product/Service ID",)
prstree.heading(3, text="Name")
prstree.heading(4, text="Description")
prstree.heading(5, text="Price")
prstree.heading(6, text="QTY")
prstree.heading(7, text="Tax1")
prstree.heading(8, text="Line Total")   
prstree.column(1, width = 50, anchor="center")
prstree.column(2, width = 270, anchor="center") 
prstree.column(3, width = 270, anchor="center")
prstree.column(4, width = 300, anchor="center")
prstree.column(5, width = 130, anchor="center")
prstree.column(6, width = 100, anchor="center")
prstree.column(7, width = 100, anchor="center")
prstree.column(8, width = 150, anchor="center")

# fbcursor.execute('SELECT * FROM Productservice;') 
# j = 0
# for i in fbcursor:
#     prstree.insert(parent='', index='end', iid=i, text='', values=(' ',i[0],i[4],i[5],i[7],i[8],i[10],(i[7]*i[8])))
#     j += 1
def mainnotepri():
  selectedmain = ordtree.focus()
  selected_note= ordtree.item(selectedmain)["values"][1]
  vel=notepriv.get('1.0','end-1c')
  edtntsql="UPDATE orders SET private_notes=%s WHERE orderid=%s"
  edtntsval=(vel,selected_note)
  fbcursor.execute(edtntsql,edtntsval)
  fbilldb.commit()
  notefchz="SELECT * FROM orders WHERE orderid=%s"
  fchz0 = (selected_note,)
  fbcursor.execute(notefchz, fchz0)
  zoom=fbcursor.fetchone()   
  notepriv.delete('1.0','end')
  notepriv.insert('1.0', zoom[27])
  edtnote.delete('1.0','end')
  edtnote.insert('1.0', zoom[27])
notepriv=scrolledtext.Text(tab2, width=165,height=11)
notepriv.place(x=10, y=25)     
noteprivButton = Button(tab2, text ="Save Note",width=10,command=mainnotepri)
noteprivButton.place(x=1080,y=1)

notesms=Text(tab3, width=2200,height=10).place(x=10, y=10)
treeblw = ttk.Treeview(tab4, columns = (1,2), height = 10, show = "headings")
treeblw.pack(side = 'top')
treeblw.heading(1)
# treeblw.heading(2, text="Attach to Email",)
treeblw.heading(2, text="Filename")
treeblw.column(1, width = 50, anchor="center")
treeblw.column(2, width = 1300, anchor="center")
# treeblw.column(3, width = 1000, anchor="center")
# expverticalbar=ttk.Scrollbar()
# expverticalbar.place(x=995+350,y=520,height=195)
scrollbar = Scrollbar()
scrollbar.place(x=992+350,y=519,height=180)



def calcu():
  subprocess.Popen('C:\\Windows\\System32\\calc.exe')
  



def ord_cust_reg():# Storing values into db (user)contp,cemail,ctel,cfax,cmob,scontp,scemail,sctel,scfax,ccountry,ccity
  customerid = custid.get()
  businessname = bname.get()
  businessaddress = baddress.get()
  category= cat.get()
  shipname= sname.get()
  shipaddress = saddress.get()
  contactperson= contp.get()
  cpemail = cemail.get()
  cptelno = ctel.get()
  cpfax = cfax.get()
  cpmobileforsms = cmob.get()
  shipcontactperson= scontp.get()
  shipcpemail= scemail.get()
  shipcptelno= sctel.get()
  shipcpfax = scfax.get()
  country= ccountry.get()
  city = ccity.get()
  notes= cnotes.get()
  status= checkvar1.get()
  taxexempt= checkvar2.get()

  sql='INSERT INTO Customer (customerid,businessname,businessaddress,category,status,shipname,shipaddress,contactperson,cpemail,cptelno,cpfax,cpmobileforsms,shipcontactperson,shipcpemail,shipcptelno,shipcpfax,taxexempt,country,city,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)' #adding values into db
  val=(customerid,businessname,businessaddress,category,status,shipname,shipaddress,contactperson,cpemail,cptelno,cpfax,cpmobileforsms,shipcontactperson,shipcpemail,shipcptelno,shipcpfax,taxexempt,country,city,notes)
  fbcursor.execute(sql,val)
  fbilldb.commit()
  messagebox.showinfo('Registration successfull','Registration successfull')

def ord_prdt_reg():# Storing values into db (user)
  sku = codeentry.get()
  category = country.get()
  name = nameentry.get()
  description = desentry.get()
  unitprice = unitentry.get()
  peices = pcsentry.get()
  cost = costentry.get()
  priceminuscost = priceentry.get()
  stock = stockentry.get()
  stocklimit = lowentry.get()
  warehouse = wareentry.get()
  privatenote = txt.get("1.0",'end-1c')
  status = checkvarStatus.get()
  taxable = checkvarStatus.get()
  serviceornot = checkvarStatus.get()
  sql='INSERT INTO Productservice (sku,category,name,description,unitprice,peices,cost,priceminuscost,stock,stocklimit,warehouse,privatenote,status,taxable,serviceornot) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)' #adding values into db
  val=(sku,category,name,description,unitprice,peices,cost,priceminuscost,stock,stocklimit,warehouse,privatenote,status,taxable,serviceornot)
  fbcursor.execute(sql,val)
  fbilldb.commit()
  messagebox.showinfo('Registration successfull','Registration successfull')

  # ##################Refresh MySQL data to Treeview#####################
  # def refresh(self):
  #     self.table.delete(*self.table.get_children())

  #     cursor = fbilldb.cursor()
  #     cursor.execute("select * from requested order by done")
  #     for row in cursor:
  #         self.table.insert('','end', values = (row[8], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[9], row[10], row[12]))




def addacnt():  
   messagebox.showinfo("F-Billing Revolution 2022", "No sender email address.\nPlease fill Your company email address textfield under the Account tab.")

def savesettings():  
   messagebox.showinfo("F-Billing Revolution 2022", "Your E-mail configuration settings has been saved.")




######################################################## END OF ORDER MODULE ###############################################################



############################## productservice Module ########################################


######### ADD New Product ##############

def adda_product():
    top = Toplevel()  
    
    top.title("Add a new Product/Service")
    p2 = PhotoImage(file = 'images/fbicon.png')
    top.iconphoto(False, p1)
    top.geometry("600x550+390+125")
    
    
    tabControl = ttk.Notebook(top)
    s = ttk.Style()
    s.theme_use('default')
    s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)


    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)
    
    tabControl.add(tab1,compound = LEFT, text ='Product/Service')
    tabControl.add(tab2,compound = LEFT, text ='Product Image')
    
    tabControl.pack(expand = 1, fill ="both")
    
    innerFrame = Frame(tab1,bg="#f5f3f2", relief=GROOVE, height=490)
    innerFrame.pack(side="top",fill=BOTH)

    Customerlabelframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=475)
    Customerlabelframe.pack(side="top",fill=BOTH,padx=10)
    
    global filename
    filename = ""
    
    def upload_file():
      global filename,img, b2
      f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
      filename = filedialog.askopenfilename(filetypes=f_types)
      shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
      image = Image.open(filename)
      resize_image = image.resize((350, 350))
      img = ImageTk.PhotoImage(resize_image)
      b2 = Button(imageFrame,image=img)
      b2.place(x=130, y=80)
    
    def addproducts():
      global img , filename 
      sku = codeentry.get()
      status = checkvarStatus.get()
      catgory = n.get()
      name = nameentry.get()
      description = desentry.get()
      unitprice = uval.get()
      peices = pcsentry.get()
      cost = costval.get()
      price_cost = priceval.get()
      taxable = checkvarStatus2.get()
      nostockcontrol = checkvarStatus3.get()
      stock = stockval.get()
      lowstock = lowval.get()
      warehouse = wareentry.get()
      pnotes = sctxt.get("1.0",'end-1c')
      entries = [sku,name, unitprice, cost]
      entri = []
      for i in entries:
        if i == '':
          entri.append(i)
      if len(entri) == 0:
        sql = 'select * from Productservice where sku = %s or name = %s'
        val  = (sku, name)
        fbcursor.execute(sql, val)
        fbcursor.fetchall()
        row_count = fbcursor.rowcount
        if row_count == 0:
          if filename == "":
            sql = 'insert into Productservice(sku, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes)
            fbcursor.execute(sql, val)
            fbilldb.commit()
          else:
            file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            sql = 'insert into Productservice(sku, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, image, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, filename.split('/')[-1], pnotes)
            fbcursor.execute(sql, val)
            fbilldb.commit()
        else:
          messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
        top.destroy()
        messagebox.showinfo("F-Billing Revolution", "Product added successfully.")
        for record in treeproducts.get_children():
          treeproducts.delete(record)
        fbcursor.execute("select *  from Productservice")
        pandsdata = fbcursor.fetchall()
        countp = 0
        for i in pandsdata:
          if i[6] == '1':
            acti = 'Active'
          else:
            acti = 'Inactive' 
          if i[13] > i[14]:
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
            countp += 1
          elif (i[12] =="0") == (i[13] <= i[14]):
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
            countp += 1
          elif i[12] == '1':
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
            countp += 1
          else:
            pass
          
       
      else:
        messagebox.showinfo("Alert", "Fields name and SKU should not be empty.\nFill out required fields and try again")
        top.destroy()
 
    fbcursor.execute("SELECT * FROM Productservice ORDER BY sku DESC LIMIT 1")
    skuin = fbcursor.fetchone()
    
    
    code1=Label(Customerlabelframe,text="Code or SKU* :",fg="blue",pady=10,padx=10)
    code1.place(x=20,y=0)
    codeentry = Entry(Customerlabelframe,width=35)
    codeentry.place(x=110,y=8)
    # if not skuin == None:
    #   fk=skuin[2]+1
    # else:
    #   fk=1
    # codeentry.insert(0, fk)

    checkvarStatus=IntVar()
    status1=Label(Customerlabelframe,text="Status:")
    status1.place(x=380,y=8)
    Button1 = Checkbutton(Customerlabelframe, 
                      variable = checkvarStatus,text="Active",compound="right",
                      onvalue =1,
                      offvalue = 0,
                      width = 10)

    Button1.place(x=420,y=5)

    category1=Label(Customerlabelframe,text="Category:",pady=5,padx=10)
    category1.place(x=20,y=40)
    n = StringVar() 
    catgory = ttk.Combobox(Customerlabelframe, width = 40, textvariable = n ) 
    catgory.place(x=110,y=45)
    catgory.insert(0, 'Default')


    name1=Label(Customerlabelframe,text="Name* :",fg="blue",pady=5,padx=10)
    name1.place(x=20,y=70)
    nameentry = Entry(Customerlabelframe,width=70)
    nameentry.place(x=110,y=75)

    des1=Label(Customerlabelframe,text="Description :",pady=5,padx=10)
    des1.place(x=20,y=100)
    desentry = Entry(Customerlabelframe,width=70)
    desentry.place(x=110,y=105)

    uval = IntVar()
    unit1=Label(Customerlabelframe,text="Unit Price:",fg="blue",pady=5,padx=10)
    unit1.place(x=20,y=130)
    unitentry = Entry(Customerlabelframe,width=20,textvariable=uval)
    unitentry.place(x=110,y=135)
    

    # pcsval = IntVar()
    pcs1=Label(Customerlabelframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
    pcs1.place(x=320,y=130)
    pcsentry = Entry(Customerlabelframe,width=20)
    pcsentry.place(x=410,y=135)

    costval = IntVar()
    cost1=Label(Customerlabelframe,text="Cost:",pady=5,padx=10)
    cost1.place(x=20,y=160)
    
    
    costentry = Entry(Customerlabelframe,width=20,textvariable=costval)
    costentry.place(x=110,y=165)
    # unit = int(uval.get())
    # cost = int(costentry.get())
    # prccst = unit-cost
    def set_label(name, index, mode):
      priceval.set(uval.get() - costval.get())
    priceval = IntVar()
    price1=Label(Customerlabelframe,text="(Price-Cost):",pady=5,padx=10)
    price1.place(x=20,y=190)
    priceentry = Entry(Customerlabelframe,width=20,textvariable=priceval,state=DISABLED,disabledbackground="white",disabledforeground="black")
    priceentry.place(x=110,y=195)
    
    uval.trace('w', set_label)
    costval.trace('w', set_label)

    checkvarStatus2=IntVar()
   
    Button2 = Checkbutton(Customerlabelframe,variable = checkvarStatus2, 
                      text="Taxable Tax1rate",compound="right",
                      onvalue =1 ,
                      offvalue = 0,
                      height=2,
                      width = 12)

    Button2.place(x=415,y=153)

    def switch():
      if checkvarStatus3.get():
        stockentry["state"] = DISABLED
        lowentry["state"] = DISABLED
        wareentry["state"] = DISABLED
      else:
        stockentry["state"] = NORMAL
        lowentry["state"] = NORMAL
        wareentry["state"] = NORMAL
    checkvarStatus3=BooleanVar()
    Button3 = Checkbutton(Customerlabelframe,variable = checkvarStatus3,command=switch, 
                      text="This is a service(no stock control)", 
                      onvalue =1 ,
                      offvalue = 0,
                      height=3)

    Button3.place(x=40,y=220)


    stockval = IntVar()
    stock1=Label(Customerlabelframe,text="Stock:",pady=5,padx=10)
    stock1.place(x=90,y=260)
    stockentry = Entry(Customerlabelframe,width=15,textvariable=stockval)
    stockentry.place(x=140,y=265)

    lowval = IntVar()
    low1=Label(Customerlabelframe,text="Low Stock Warning Limit:",pady=5,padx=10)
    low1.place(x=280,y=260)
    lowentry = Entry(Customerlabelframe,width=15,textvariable=lowval)
    lowentry.place(x=435,y=265)

   
    ware1=Label(Customerlabelframe,text="Warehouse:",pady=5,padx=10)
    ware1.place(x=60,y=290)
    wareentry = Entry(Customerlabelframe,width=64)
    wareentry.place(x=140,y=295)

    # pnoteval = StringVar()
    text1=Label(Customerlabelframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
    text1.place(x=20,y=320)
    sctxt = scrolledtext.ScrolledText(Customerlabelframe, undo=True,width=62,height=4)
    sctxt.place(x=32,y=358)
    
    okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=addproducts)
    okButton.pack(side=LEFT, padx=(10, 0), pady=(5, 10))
    
    def closetab():
      top.destroy()

    cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=closetab)
    cancelButton.pack(side=RIGHT, padx=(0, 10), pady=(5, 10))

    imageFrame = Frame(tab2, relief=GROOVE,height=580)
    imageFrame.pack(side="top",fill=BOTH)

    
      
    browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
    browseimg.place(x=30,y=35)
      
    browsebutton=Button(imageFrame,text = 'Browse',command=upload_file)
    browsebutton.place(x=485,y=30,height=30,width=50)

    removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT, command=lambda: b2.destroy())
    removeButton.place(x=410,y=460)

    top.mainloop()

def fileimport_product():

    top=Toplevel()
    top.title("Import items list from Excel(XLS)File")
    top.geometry("785x520+280+100")
    importframe=Frame(top)
    importframe.place(x=0,y=0,height=700,width=785)
    impolbl=Label(importframe,text="Import source Excel(xlsx) File:").place(x=8,y=30)
    impoentry=Entry(importframe,bg="white")
    impoentry.place(x=8,y=50,width=280, height=25)
    previewlbl=Label(importframe,text="Source File preview").place(x=8,y=77)
   
    ###### LISTBOX #####################
    scrollbarx = Scrollbar(importframe, orient=HORIZONTAL)
    scrollbary = Scrollbar(importframe, orient=VERTICAL)
    imptree = ttk.Treeview(importframe, columns=("PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"), height=400,     selectmode="extended", yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
    scrollbary.config(command=imptree.yview)
    scrollbary.place(x=354,y=100,height=325)
    scrollbarx.config(command=imptree.xview)
    scrollbarx.place(x=0,y=410, width=356)
    imptree.heading('PRODUCT SERVICE ID', text="PRODUCT SERVICE ID", anchor=W)
    imptree.heading('CODE OR SKU', text="CODE OR SKU", anchor=W)
    imptree.heading('NAME', text="NAME", anchor=W)
    imptree.heading('CATEGORY', text="CATEGORY", anchor=W)
    imptree.heading('NAME', text="NAME", anchor=W)
    imptree.heading('DESCRIPTION', text="DESCRIPTION", anchor=W)
    imptree.heading('QTY UNIT', text="QTY UNIT", anchor=W)
    imptree.heading('COST', text="COST", anchor=W)
    imptree.heading('PRICE', text="PRICE", anchor=W)
    imptree.heading('TAX1', text="TAX1", anchor=W)
    imptree.heading('STOCK', text="STOCK", anchor=W)
    imptree.heading('LOW STOCK', text="LOW STOCK", anchor=W)
    imptree.heading('LOCATION', text="LOCATION", anchor=W)
    imptree.heading('ACTIVE', text="ACTIVE", anchor=W)
    imptree.heading('SERVICE', text="SERVICE", anchor=W)
    

    imptree.column('#0', stretch=NO, minwidth=0, width=0)
    imptree.column('#1', stretch=NO, minwidth=0, width=120)
    imptree.column('#2', stretch=NO, minwidth=0, width=100)
    imptree.column('#3', stretch=NO, minwidth=0, width=100)
    imptree.column('#4', stretch=NO, minwidth=0, width=100)
    imptree.column('#5', stretch=NO, minwidth=0, width=100)
    imptree.column('#6', stretch=NO, minwidth=0, width=100)
    imptree.column('#7', stretch=NO, minwidth=0, width=100)
    imptree.column('#8', stretch=NO, minwidth=0, width=100)
    imptree.column('#9', stretch=NO, minwidth=0, width=100)
    imptree.column('#10', stretch=NO, minwidth=0, width=100)
    imptree.column('#12', stretch=NO, minwidth=0, width=100)
    imptree.column('#13', stretch=NO, minwidth=0, width=100)
    imptree.column('#14', stretch=NO, minwidth=0, width=100)

 

    imptree.place(x=0,y=100,height=315,width=356)
    # langs = ()

    # langs_var = StringVar(value=langs)

    # listbox = Listbox(
    #     importframe,
    #     listvariable=langs_var,
    #     width=60,
    #     height=6,
    #     selectmode='extended')

    # listbox.place(x=8,y=100,height=320)
    
    # link a scrollbar to a list
    # scrollbar = Scrollbar(
    #     importframe,
    #     orient='vertical',
    #     command=listbox.yview
    # )
    
    # listbox['yscrollcommand'] = scrollbar.set
    # scrollbar.place(x=354,y=100,height=319)


    # scrollbar12 = Scrollbar(
    #     importframe,
    #     orient='horizontal',
    #     command=listbox.xview 
    # )
    # listbox['xscrollcommand'] = scrollbar.set
    # scrollbar12.place(x=0,y=402, width=354)

    
    

    lb1=Label(importframe,text="Select import source XLs file first after build column associations").place(x=8,y=480)

    

    def export_product_1():
      global Productserviceid,name12,category12,description,peices,cost12,priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot,name
      name = askopenfilename(filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
      # df = pd.read_csv(name)
      # for i in df:
      #   listbox.insert(END, df)
      with open(name) as f:
        reader = csv.DictReader(f, delimiter=',')
        print(reader)
        for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
          Productserviceid = row['PRODUCT SERVICE ID']
          sku = row['CODE OR SKU']          
          name12 = row['NAME']
          category12 = row['CATEGORY']
          
          description = row['DESCRIPTION']
          peices = row['QTY UNIT']
          cost12 = row['COST']
          priceminuscost = row['PRICE']
          taxable = row['TAX1']
          stock12 = row['STOCK']
          stocklimit = row['LOW STOCK']
          warehouse = row['LOCATION']
          status = row['ACTIVE']
          serviceornot = row['SERVICE']
        
          imptree.insert("", 0, values=(Productserviceid,sku,name12,category12,description,peices,cost12,priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot))
          

        
          # print(row)
          # listbox.insert(END, row)
      impoentry.delete(0, 'end')
      impoentry.insert(0, name)
      
    def nxtscreen():
      def save_pro_import():
        with open(name) as f:
          reader = csv.DictReader(f, delimiter=',')
          for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
            Productserviceid = int(row['PRODUCT SERVICE ID'])
            sku = int(row['CODE OR SKU'])  
            name12 = row['NAME']
            category12 = row['CATEGORY']
            description = row['DESCRIPTION']
            peices = int(row['QTY UNIT'])
            cost12 = int(row['COST'])
            unitprice = int(row['PRICE'])
            taxable = int(row['TAX1'])
            stock12 = int(row['STOCK'])
            stocklimit = int(row['LOW STOCK'])
            warehouse = row['LOCATION']
            status = int(row['ACTIVE'])
            serviceornot = int(row['SERVICE'])
            min = int(unitprice) - int(cost12)

            sql = 'select * from Productservice where Productserviceid = %s or name = %s or sku=%s'
            val  = (Productserviceid, name12,sku)
            fbcursor.execute(sql, val)
            fbcursor.fetchall()
            row_count = fbcursor.rowcount
            if row_count == 0:
              sql = 'insert into Productservice(Productserviceid,sku,name,category,description,peices,cost,unitprice,taxable,stock,stocklimit,warehouse,status,serviceornot,priceminuscost) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)'
              val = (Productserviceid,sku,name12,category12,description,peices,cost12,unitprice,taxable,stock12,stocklimit,warehouse,status,serviceornot,min)
              fbcursor.execute(sql, val)
              fbilldb.commit()
              topp.destroy()
              for record in treeproducts.get_children():
                treeproducts.delete(record)
              fbcursor.execute("select *  from Productservice")
              pandsdata = fbcursor.fetchall()
              countp = 0
              for i in pandsdata:
                if i[6] == '1':
                  acti = 'Active'
                else:
                  acti = 'Inactive' 
                if i[13] > i[14]:
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
                  countp += 1
                elif (i[12] =="0") == (i[13] <= i[14]):
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
                  countp += 1
                elif i[12] == '1':
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
                  countp += 1
                else:
                  pass
            else:
              messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
            
            

            # if len(entri) == 0:
            #   sql = 'select * from Productservice where sku = %s or name = %s'
            #   val  = (sku, name)
            #   fbcursor.execute(sql, val)
            #   fbcursor.fetchall()
            #   row_count = fbcursor.rowcount
            #   if row_count == 0:
            #     if filename == "":
                  
      
        

     
      topp=Toplevel()
      topp.title("Import items list from Excel(XLS)File")
      topp.geometry("785x520+280+100")
      scrollbarx = Scrollbar(topp, orient=HORIZONTAL)
      scrollbary = Scrollbar(topp, orient=VERTICAL)
      nxttree = ttk.Treeview(topp, columns=("PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"),height=400,     selectmode="extended", yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
      scrollbary.config(command=nxttree.yview)
      scrollbary.place(x=768,y=0,height=490)
      scrollbarx.config(command=nxttree.xview)
      scrollbarx.place(x=0,y=470,width=763)
      nxttree.heading('CODE OR SKU', text="CODE OR SKU", anchor=W)
      nxttree.heading('PRODUCT SERVICE ID', text="PRODUCT SERVICE ID", anchor=W)
      nxttree.heading('NAME', text="NAME", anchor=W)
      nxttree.heading('CATEGORY', text="CATEGORY", anchor=W)
      nxttree.heading('NAME', text="NAME", anchor=W)
      nxttree.heading('DESCRIPTION', text="DESCRIPTION", anchor=W)
      nxttree.heading('QTY UNIT', text="QTY UNIT", anchor=W)
      nxttree.heading('COST', text="COST", anchor=W)
      nxttree.heading('PRICE', text="PRICE", anchor=W)
      nxttree.heading('TAX1', text="TAX1", anchor=W)
      nxttree.heading('STOCK', text="STOCK", anchor=W)
      nxttree.heading('LOW STOCK', text="LOW STOCK", anchor=W)
      nxttree.heading('LOCATION', text="LOCATION", anchor=W)
      nxttree.heading('ACTIVE', text="ACTIVE", anchor=W)
      nxttree.heading('SERVICE', text="SERVICE", anchor=W)
  
      nxttree.column('#0', stretch=NO, minwidth=0, width=0)
      nxttree.column('#1', stretch=NO, minwidth=0, width=120)
      nxttree.column('#2', stretch=NO, minwidth=0, width=100)
      nxttree.column('#3', stretch=NO, minwidth=0, width=100)
      nxttree.column('#4', stretch=NO, minwidth=0, width=100)
      nxttree.column('#5', stretch=NO, minwidth=0, width=100)
      nxttree.column('#6', stretch=NO, minwidth=0, width=100)
      nxttree.column('#7', stretch=NO, minwidth=0, width=100)
      nxttree.column('#8', stretch=NO, minwidth=0, width=100)
      nxttree.column('#9', stretch=NO, minwidth=0, width=100)
      nxttree.column('#10', stretch=NO, minwidth=0, width=100)
      nxttree.column('#12', stretch=NO, minwidth=0, width=100)
      nxttree.column('#13', stretch=NO, minwidth=0, width=100)
      nxttree.column('#14', stretch=NO, minwidth=0, width=100)
    
      with open(name) as f:
        reader = csv.DictReader(f, delimiter=',')
        for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
          Productserviceid = row['PRODUCT SERVICE ID']
          sku = row['CODE OR SKU']
          name12 = row['NAME']
          category12 = row['CATEGORY']
          
          description = row['DESCRIPTION']
          peices = row['QTY UNIT']
          cost12 = row['COST']
          priceminuscost = row['PRICE']
          taxable = row['TAX1']
          stock12 = row['STOCK']
          stocklimit = row['LOW STOCK']
          warehouse = row['LOCATION']
          status = row['ACTIVE']
          serviceornot = row['SERVICE']
      

      
          nxttree.insert("", 0, values=(Productserviceid,sku,name12,category12,description,peices,cost12,  priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot))
       
    
      nxttree.place(x=0,y=0,height=470,width=770)
      back = Button(topp,text="back",command=lambda:topp.destroy())
      back.place(x=5,y=492)
      Finish = Button(topp,text="Finish",command=save_pro_import)
      Finish.place(x=740,y=492)
     
      


      

      # name = askopenfilename(filetypes=[('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
      # impoentry.delete(0, 'end')
      # impoentry.insert(0, name)
      # wb = Workbook()
      # wb = load_workbook(name)
      # ws = wb.active
      # column_a = ws['A']
      # column_b = ws['B']
      # paslist = []
      # for cell in column_b:
      #     paslist.append(str(cell.value))
      
      # pslist = paslist
      # nameentry.insert(0, pslist[0])
      # category.insert(0, pslist[1])
      # desc.insert(0, pslist[2])
      # quantity.insert(0, pslist[3])
      # cost.insert(0, pslist[4])
      # price.insert(0, pslist[5])
      # tax.insert(0, pslist[6])
      # tax2.insert(0, pslist[7])
      # stock2.insert(0, pslist[8])
      # lowstock.insert(0, pslist[9])
      # location.insert(0, pslist[10])
      # active.insert(0, pslist[11])
      # service.insert(0, pslist[12])
      
      # paslist.reverse()
      # listbox.delete(0, 'end')
      # for i in paslist:
      #   listbox.insert(0, f'{i}\n')
      
    
    importbutton=Button(top,command=export_product_1,text = 'Browse',compound=LEFT)
    importbutton.place(x=290,y=48,height=25,width=80)

    lb1=Label(importframe,text="Please associate datafields with data columns").place(x=460,y=30)

    id=Label(importframe,text="PRODUCT SERVICE ID = ",pady=5,padx=10,fg="blue")
    id.place(x=430,y=90)
    idd = StringVar() 
    idds = ttk.Combobox(importframe, width = 30, textvariable = idd) 
    idds.place(x=570,y=96) 

    name1=Label(importframe,text="NAME = ",pady=5,padx=10,fg="blue")
    name1.place(x=450,y=115)
    namevar = StringVar() 
    nameentry = ttk.Combobox(importframe, width = 30, textvariable = namevar ) 
    nameentry.place(x=570,y=121)

   
    category1=Label(importframe,text="CATEGORY = ",pady=5,padx=10,fg="blue")
    category1.place(x=450,y=140)
    categoryvar = StringVar() 
    category = ttk.Combobox(importframe, width = 30, textvariable = categoryvar ) 
    category.place(x=570,y=146)

    desc1=Label(importframe,text="DESCRIPTION = ",pady=5,padx=10)
    desc1.place(x=450,y=165)
    
    descvar = StringVar() 
    desc = ttk.Combobox(importframe, width = 30, textvariable = descvar ) 
    desc.place(x=570,y=171)

    quan1=Label(importframe,text="QUANTITY= ",pady=5,padx=10)
    quan1.place(x=450,y= 190)
    
    quanvar = StringVar() 
    quantity = ttk.Combobox(importframe, width = 30, textvariable = categoryvar ) 
    quantity.place(x=570,y=196)

    cost1=Label(importframe,text="COST = ",pady=5,padx=10)
    cost1.place(x=450,y=215)
    
    costvar = StringVar() 
    cost = ttk.Combobox(importframe, width = 30, textvariable = costvar ) 
    cost.place(x=570,y=221)

    price1=Label(importframe,text="PRICE = ",pady=5,padx=10,fg="blue")
    price1.place(x=450,y=240)
    
    pricevar = StringVar() 
    price = ttk.Combobox(importframe, width = 30, textvariable = costvar ) 
    price.place(x=570,y=246)


    taxx1=Label(importframe,text="TAX1 = ",pady=5,padx=10,)
    taxx1.place(x=450,y=265)
    taxvar = StringVar() 
    tax = ttk.Combobox(importframe, width = 30, textvariable = taxvar ) 
    tax.place(x=570,y=271)

    # taxx2=Label(importframe,text="TAX2 = ",pady=5,padx=10)
    # taxx2.place(x=450,y=265)
    # tax2var = StringVar() 
    # tax2 = Entry(importframe, width = 30, textvariable = tax2var ) 
    # tax2.place(x=570,y=271)

    stock2=Label(importframe,text="STOCK = ",pady=5,padx=10)
    stock2.place(x=450,y=290)
    
    stock2var = StringVar() 
    stock2 = ttk.Combobox(importframe, width = 30, textvariable = stock2var ) 
    stock2.place(x=570,y=296)


    lowstock2=Label(importframe,text="LOW STOCK = ",pady=5,padx=10)
    lowstock2.place(x=450,y=315)
    
    lowstock2var = StringVar() 
    lowstock = ttk.Combobox(importframe, width = 30, textvariable = lowstock2var ) 
    lowstock.place(x=570,y=321)


    location2=Label(importframe,text="LOCATION = ",pady=5,padx=10)
    location2.place(x=450,y=340)
    Location2var = StringVar() 
    location = ttk.Combobox(importframe, width = 30, textvariable = Location2var ) 
    location.place(x=570,y=346)

    active2=Label(importframe,text="ACTIVE = ",pady=5,padx=10)
    active2.place(x=450,y=365)
    
    active2var = StringVar() 
    active = ttk.Combobox(importframe, width = 30, textvariable = active2var ) 
    active.place(x=570,y=371)

    service2=Label(importframe,text="SERVICE = ",pady=5,padx=10)
    service2.place(x=450,y=390)
    
    service2var = StringVar() 
    service = ttk.Combobox(importframe, width = 30, textvariable = service2var ) 
    service.place(x=570,y=396)

    b = Button(importframe,text = "Clear associations").place(x=600,y=470)  
    n = Button(importframe, text ="Next",command=nxtscreen).place(x=710,y=470)
  
    
    top.mainloop()
    
#########EXPORT PRODUCT#######################################################################################

def export_product():
  cols = ["PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"] # Your column headings here
  path = filedialog.asksaveasfilename(initialdir=os.getcwd,title="Save File",filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
  excel_name = 'newfile.xlsx'
  lst = []
  with open(path, "w", newline='') as myfile:
      csvwriter = csv.writer(myfile, delimiter=',')
      sql = 'select Productserviceid,sku,name,category,description,peices,cost,unitprice,taxable,stock,stocklimit,warehouse,status,serviceornot from Productservice'
      fbcursor.execute(sql)
      pandsdata = fbcursor.fetchall()
      print (pandsdata)
      for row_id in pandsdata:
          row = row_id
          lst.append(row)
      lst = list(map(list,lst))
      lst.insert(0,cols)
      for row in lst:
          csvwriter.writerow(row)

  writer = pd.ExcelWriter(excel_name)
  df = pd.read_csv(path)
  df.to_excel(writer,'sheetname')
  writer.save()
    # name = filedialog.asksaveasfilename(initialdir=os.getcwd,title="Save File",filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
    # with open(name,mode='w',newline='') as myfile:
    #   exp_writer = csv.writer(myfile,delimiter='\t')
    #   for row_id in treeproducts.get_children():
    #       row = treeproducts.item(row_id)['values']
    #       exp_writer.writerow(row)
        
    #   # sql = 'select * from Productservice'
    #   # fbcursor.execute(sql)
    #   # pandsdata = fbcursor.fetchall()
    #   # for i in pandsdata:
    #   #   exp_writer.writerow(i)
    # messagebox.showinfo("saved")

######################## VIEW/EDIT PRODUCT #######################################################################

  
def edit_product():  
    try:
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      
      global filename
      filename = ""
      
      def upload_file():
        global filename,img, b2
        f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
        filename = filedialog.askopenfilename(filetypes=f_types)
        shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
        image = Image.open(filename)
        resize_image = image.resize((350, 350))
        img = ImageTk.PhotoImage(resize_image)
        b2 = Button(imageFrame,image=img)
        b2.place(x=130, y=80)
      
      def updateproducts():
        global img , filename 
        sku = codeentry.get()
        status = checkvarStatus.get()
        catgory = n.get()
        name = nameentry.get()
        description = desentry.get()
        unitprice = uval.get()
        peices = pcsentry.get()
        cost = costval.get()
        price_cost = priceval.get()
        taxable = checkvarStatus2.get()
        nostockcontrol = checkvarStatus3.get()
        stock = stockval.get()
        lowstock = lowval.get()
        warehouse = wareentry.get()
        pnotes = sctxt.get("1.0", 'end-1c')
        entries = [sku, name, unitprice, cost]
        entri = []
        for i in entries:
          if i == '':
            entri.append(i)
        if len(entri) == 0:
          if filename == "":
            sql = "update Productservice set sku=%s, category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, privatenote=%s where Productserviceid = %s"
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes, itemid)
            fbcursor.execute(sql, val)
            fbilldb.commit()
          else:
            file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            sql = "update Productservice set category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, image=%s, privatenote=%s where Productserviceid = %s"
            val = (catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse,filename.split('/')[-1], pnotes, itemid)
            fbcursor.execute(sql, val)
            fbilldb.commit()
            messagebox.showinfo("F-Billing Revolution", "Product updated successfully.")
            for record in treeproducts.get_children():
              treeproducts.delete(record)
            fbcursor.execute("select *  from Productservice")
            pandsdata = fbcursor.fetchall()
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4]  , i[7], i[13], i[15],i[17]),tags=('red',))
                countp += 1
              elif i[12] == '1':
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
                countp += 1
              else:
                pass
          top.destroy()
        else:
          messagebox.showinfo("F-Billing Revolution", "Fields name or SKU entered is already in database.")
          top.destroy()
        for record in treeproducts.get_children():
              treeproducts.delete(record)
        fbcursor.execute("select *  from Productservice")
        pandsdata = fbcursor.fetchall()
        countp = 0
        for i in pandsdata:
          if i[6] == '1':
            acti = 'Active'
          else:
            acti = 'Inactive' 
          if i[13] > i[14]:
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
            countp += 1
          elif (i[12] =="0") == (i[13] <= i[14]):
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('red',))
            countp += 1
          elif i[12] == '1':
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
            countp += 1
          else:
            pass
         
      sql = "select * from Productservice where Productserviceid = %s"
      val = (itemid, )
      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone()
      
      
      top = Toplevel()  
      top.title("Edit Product/Service details")
      p3 = PhotoImage(file = 'images/fbicon.png')
      top.iconphoto(False, p1)
      top.geometry("600x550+390+125")
      tabControl = ttk.Notebook(top)
      s = ttk.Style()
      s.theme_use('default')
      s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)

      taba = ttk.Frame(tabControl)
      tabb = ttk.Frame(tabControl)
      
      tabControl.add(taba,compound = LEFT, text ='Product/Service')
      tabControl.add(tabb,compound = LEFT, text ='Product Image')
      
      tabControl.pack(expand = 1, fill ="both")
      
      innerFrame = Frame(taba,bg="#f5f3f2", relief=GROOVE)
      innerFrame.pack(side="top",fill=BOTH)

      updateframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
      updateframe.pack(side="top",fill=BOTH,padx=10)

      code1=Label(updateframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
      code1.place(x=20,y=0)
      codeentry = Entry(updateframe,width=35)
      codeentry.place(x=110,y=8)
      codeentry.insert(0, psdata[2])

      checkvarStatus=IntVar()
      status1=Label(updateframe,text="Status:")
      status1.place(x=380,y=8)
      Button1 = Checkbutton(updateframe, 
                        variable = checkvarStatus,text="Active",compound="right",
                        onvalue =1,
                        offvalue =0,
                        width = 10)
      Button1.place(x=420,y=5)
      sta = psdata[6]
      if sta == '1':
        Button1.select()
      else:
        Button1.deselect()



      category1=Label(updateframe,text="Category:",pady=5,padx=10)
      category1.place(x=20,y=40)
      n = StringVar() 
      category = Entry(updateframe,width=70,textvariable=n) 
      category.place(x=110,y=45)
      category.insert(0, psdata[3])


      name1=Label(updateframe,text="Name :",fg="blue",pady=5,padx=10)
      name1.place(x=20,y=70)
      nameentry = Entry(updateframe,width=70)
      nameentry.place(x=110,y=75)
      nameentry.insert(0, psdata[4])

      des1=Label(updateframe,text="Description :",pady=5,padx=10)
      des1.place(x=20,y=100)
      desentry = Entry(updateframe,width=70)
      desentry.place(x=110,y=105)
      desentry.insert(0, psdata[5])

      def set_label(name, index, mode):
        priceval.set(uval.get() - costval.get())
      
      unit1=Label(updateframe,text="Unit Price:",fg="blue",pady=5,padx=10)
      unit1.place(x=20,y=130)
      
      uval = IntVar()
      unitentry = Entry(updateframe,width=20,textvariable=uval)
      unitentry.place(x=110,y=135)
      unitentry.delete(0,'end')
      unitentry.insert(0, psdata[7])
      

      pcsval = IntVar()
      pcs1=Label(updateframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
      pcs1.place(x=320,y=130)
      pcsentry = Entry(updateframe,width=20,textvariable=pcsval)
      pcsentry.place(x=410,y=135)
      pcsentry.delete(0,'end')
      pcsentry.insert(0, psdata[8])
      

      costval = IntVar()
      cost1=Label(updateframe,text="Cost:",pady=5,padx=10)
      cost1.place(x=20,y=160)
      costentry = Entry(updateframe,width=20,textvariable=costval)
      costentry.place(x=110,y=165)
      costentry.delete(0,'end')
      costentry.insert(0, psdata[9])
      

      priceval = IntVar()
      price1=Label(updateframe,text="(Price-Cost):",pady=5,padx=10)
      price1.place(x=20,y=190)
      priceentry = Entry(updateframe,width=20,textvariable=priceval)
      priceentry.place(x=110,y=195)
      priceentry.delete(0,'end')
      priceentry.insert(0, psdata[11])

      uval.trace('w', set_label)
      costval.trace('w', set_label)
      

      checkvarStatus2=IntVar()
    
      Button2 = Checkbutton(updateframe,variable = checkvarStatus2, 
                        text="Taxable Tax1rate",compound="right",
                        onvalue =1 ,
                        offvalue =0,
                        height=2,
                        width = 12)

      Button2.place(x=415,y=153)
      tax = psdata[10]
      if tax == '1':
        Button2.select()
      else:
        Button2.deselect()
      

      

      checkvarStatus3=IntVar()
      def switch():
        if checkvarStatus3.get():
          stockentry["state"] = DISABLED
          lowentry["state"] = DISABLED
          wareentry["state"] = DISABLED
        else:
          stockentry["state"] = NORMAL
          lowentry["state"] = NORMAL
          wareentry["state"] = NORMAL
    
      Button3 = Checkbutton(updateframe,variable = checkvarStatus3,command=switch, 
                        text="No stock Control", 
                        onvalue =1 ,
                        offvalue = 0,
                        height=3,
                        width = 15)

      Button3.place(x=40,y=220)

      


      stockval = IntVar(updateframe)
      stock1=Label(updateframe,text="Stock:",pady=5,padx=10)
      stock1.place(x=90,y=260)
      stockentry = Entry(updateframe,width=15,textvariable=stockval)
      stockentry.place(x=140,y=265)
      stockentry.delete(0,'end')
      stockentry.insert(0, psdata[13])
      

      lowval = IntVar(updateframe)
      low1=Label(updateframe,text="Low Stock Warning Limit:",pady=5,padx=10)
      low1.place(x=280,y=260)
      lowentry = Entry(updateframe,width=15,textvariable=lowval)
      lowentry.place(x=435,y=265)
      lowentry.delete(0,'end')
      lowentry.insert(0, psdata[14])
      

    
      ware1=Label(updateframe,text="Warehouse:",pady=5,padx=10)
      ware1.place(x=60,y=290)
      wareentry = Entry(updateframe,width=64)
      wareentry.place(x=140,y=295)
      wareentry.insert(0, psdata[15])

      scr = psdata[12]
      if scr == '1':
        Button3.select()
        stockentry["state"] = DISABLED
        lowentry["state"] = DISABLED
        wareentry["state"] = DISABLED
      else:
        Button3.deselect()
        stockentry["state"] = NORMAL
        lowentry["state"] = NORMAL
        wareentry["state"] = NORMAL
      
      

      

      text1=Label(updateframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
      text1.place(x=20,y=320)
      sctxt = scrolledtext.ScrolledText(updateframe, undo=True,width=62,height=4)
      sctxt.place(x=32,y=358)
      try:
        sctxt.insert("1.0", psdata[16])
      except:
        pass

      okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=updateproducts)
      okButton.pack(side=LEFT, padx=(10, 0))

      cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=lambda : top.destroy())
      cancelButton.pack(side=RIGHT, padx=(0, 10))
      
      
      imageFrame = Frame(tabb, relief=GROOVE,height=580)
      imageFrame.pack(side="top",fill=BOTH)

      browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
      browseimg.place(x=15,y=35)

      browsebutton=Button(imageFrame,text = 'Browse', command=upload_file)
      browsebutton.place(x=470,y=30,height=30,width=50)

      try:
        image = Image.open("images/"+psdata[17])
        resize_image = image.resize((350, 350))
        image = ImageTk.PhotoImage(resize_image)
        b2 = Label(imageFrame,image=image,width=350,height=350)
        b2.photo = image
        b2.place(x=130, y=80)
        print(image)
      except:
        pass

      removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT,command=lambda: b2.destroy())
      removeButton.place(x=410,y=460)
    except:
      try:
        top.destroy()
      except:
        pass
      messagebox.showerror('F-Billing Revolution', 'Select a record to edit.')
  


######################## DELETE PRODUCT #######################################################################


def delete_product():
    delmess = messagebox.askyesno("Delete product/service", "Are you sure to delete this product/service?")
    if delmess == True:
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      sql = "delete from Productservice where Productserviceid = %s"
      val = (itemid, )
      fbcursor.execute(sql, val)
      fbilldb.commit()
      treeproducts.delete(treeproducts.selection()[0])
      # messagebox.showinfo("F-Billing Revolution", "Record deleted successfully.")
    else:
      pass


######################## SEARCH PRODUCT  #######################################################################

def search_pro():
  query = searchvar.get()
  selections = []
  for child in treeproducts.get_children():
      if query in treeproducts.item(child)['values']:
          print(treeproducts.item(child)['values'])
          selections.append(child)
  treeproducts.selection_set(selections)
  

def search_product():
    global searchvar, searchtop
    searchtop = Toplevel()  
    searchtop.title("Find Text")
    searchtop.geometry("520x200+390+250")
    
    findwhat1=Label(searchtop,text="Find What:",pady=5,padx=10)
    findwhat1.place(x=5,y=20)
    searchvar = StringVar() 
    findwhat = ttk.Combobox(searchtop, width = 50, textvariable = searchvar)
    findwhat.place(x=80,y=25) 
    

    findButton = Button(searchtop, text ="Find next",width=10, command=search_pro)
    findButton.place(x=420,y=20)
    
    findin1=Label(searchtop,text="Find in:",pady=5,padx=10)
    findin1.place(x=5,y=47)
    n = StringVar() 
    findIN = ttk.Combobox(searchtop, width = 37, textvariable = n ) 
    # Adding combobox drop down list 
    findIN['values'] = ('Product/Service id',  
                              'Category', 
                              'Active', 
                              'name', 
                              'stock', 
                              'location', 
                              'image', 
                              '<<All>>') 
      
    findIN.place(x=80,y=54) 
    findIN.current(0)

    closeButton = Button(searchtop, text ="Close",width=10, command=lambda : searchtop.destroy())
    closeButton.place(x=420,y=50)

    match1=Label(searchtop,text="Match:",pady=5,padx=10)
    match1.place(x=5,y=74)
    n = StringVar() 
    match = ttk.Combobox(searchtop, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    match['values'] = ('From Any part',' Whole Field',  
                              ' From the beginning of the field')
      
    match.place(x=80,y=83) 
    match.current(0)

    search1=Label(searchtop,text="Search:",pady=5,padx=10)
    search1.place(x=5,y=102)
    n = StringVar() 
    search = ttk.Combobox(searchtop, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    search['values'] = ('All', 'up', 
                              'Down')
      
    search.place(x=80,y=112) 
    search.current(0)


    checkvarStatus4=IntVar()
   
    Button4 = Checkbutton(searchtop,variable = checkvarStatus4, 
                      text="Match Case", 
                      onvalue =0 ,
                      offvalue = 1,
                      height=3,
                      width = 15)

    Button4.place(x=60,y=141)

    checkvarStatus5=IntVar()
   
    Button5 = Checkbutton(searchtop,variable = checkvarStatus5, 
                      text="Match Format", 
                      onvalue =0 ,
                      offvalue = 1,
                      height=3,
                      width = 15)

    Button5.place(x=270,y=141)

    searchtop.mainloop()


######################## REFRESH PRODUCT  #######################################################################

def refresh_pro_s():
  for record in treeproducts.get_children():
    treeproducts.delete(record)
  fbcursor.execute("select *  from Productservice")
  pandsdata = fbcursor.fetchall()
  countp = 0
  for i in pandsdata:
    if i[6] == '1':
      acti = 'Active'
    else:
      acti = 'Inactive' 
    if i[13] > i[14]:
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
      countp += 1
    elif (i[12] =="0") == (i[13] <= i[14]):
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
      countp += 1
    elif i[12] == '1':
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
      countp += 1
    else:
      pass

######################## View Image Product services
# #######################################################################
def psfile_image(event):
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      sql = "select * from Productservice where Productserviceid = %s"
      val = (itemid, )

      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone() 
      if psdata[17] is None:
        pass
      else:
        edit_window_img = Toplevel()
        edit_window_img.title("View Image")
        edit_window_img.geometry("700x500")
        image = Image.open("images/"+psdata[17])
        resize_image = image.resize((700, 500))
        image = ImageTk.PhotoImage(resize_image)
        psimage = Label(edit_window_img,image=image)
        psimage.photo = image
        psimage.pack()
  
######################## FRONT PAGE OF PRODUCT SERVICE SECTION #######################################################################

    
mainFrameps=Frame(tab8, relief=GROOVE, bg="#f8f8f2")
mainFrameps.pack(side="top", fill=BOTH)

midFrame=Frame(mainFrameps, bg="#f5f3f2", height=60)
midFrame.pack(side="top", fill=X)

lFrame=Frame(tab8, bg="#f8f8f2", height=600)
lFrame.pack(side="top", fill=X)



pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=(5, 2))
pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=(0, 5))

productIcon = ImageTk.PhotoImage(Image.open("images/plus.png"))
productLabel = Button(midFrame,compound="top", text="Add new\nProduct",relief=RAISED,command=adda_product, image=productIcon, font=("arial", 8),bg="#f5f3f2", fg="black", height=55, bd=1, width=55)
productLabel.pack(side="left", pady=3, ipadx=4)

proeditIcon = ImageTk.PhotoImage(Image.open("images/edit.png"))
proeditLabel = Button(midFrame,compound="top", text="Edit/View\nProduct",relief=RAISED, image=proeditIcon,command=edit_product, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proeditLabel.pack(side="left")

prodeleteIcon = ImageTk.PhotoImage(Image.open("images/delete.png"))
prodeleteLabel = Button(midFrame,compound="top", text="Delete\nSelected",relief=RAISED, command=delete_product,image=prodeleteIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
prodeleteLabel.pack(side="left")

pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=5)

prosearchIcon = ImageTk.PhotoImage(Image.open("images/research.png"))
prosearchLabel = Button(midFrame,compound="top", text="Search in\nproducts",relief=RAISED,command=search_product, image=prosearchIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55, activebackground="red")
prosearchLabel.pack(side="left")

proimportIcon = ImageTk.PhotoImage(Image.open("images/import.png"))
proimportLabel = Button(midFrame,compound="top", text="Import\nProducts",relief=RAISED,command=fileimport_product, image=proimportIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proimportLabel.pack(side="left")

pn = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=5)

proexportIcon = ImageTk.PhotoImage(Image.open("images/export-file.png"))
proexportLabel = Button(midFrame,compound="top",command=export_product, text="Export\nProducts",relief=RAISED, image=proexportIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proexportLabel.pack(side="left")


pn = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
pn.pack(side="left")

prorefreshIcon = ImageTk.PhotoImage(Image.open("images/refresh.png"))
productrefreshLabel = Button(midFrame,compound="top", text="Refresh\nProduct List",relief=RAISED, image=prorefreshIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55, command=refresh_pro_s)
productrefreshLabel.pack(side="left")

prolabel = Label(mainFrameps, text="Products/Services", font=("arial", 18), bg="#f8f8f2")
prolabel.pack(side="left", padx=(20,0))

pr_label = Label(mainFrameps, text="Category", font=("arial", 16), bg="#f8f8f2")
pr_label.place(x=1099,y=70)

def pro_fil(event):
  pro_f = dro.get()
  for record in treeproducts.get_children():
    treeproducts.delete(record)

  sql = "select * from Productservice where category = %s"
  val = (pro_f,)
  fbcursor.execute(sql, val)
  product_fil = fbcursor.fetchall()
  
  countp = 0
  for i in product_fil:
    if i[6] == '1':
      acti = 'Active'
    else:
      acti = 'Inactive' 
    if i[13] > i[14]:
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
      countp += 1
    elif (i[12] =="0") == (i[13] <= i[14]):
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('red',))
      countp += 1
    elif i[12] == '1':
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
      countp += 1
    else:
      pass

sql = "SELECT DISTINCT category FROM Productservice"
fbcursor.execute(sql,)
lic = fbcursor.fetchall()
dro = ttk.Combobox(mainFrameps,)
dro.pack(side="right", padx=(0,10))
dro['values'] = lic
dro.bind("<<ComboboxSelected>>", pro_fil)

pro_label = Label(mainFrameps, text="Right click on datagrid row for more options.",  bg="#f8f8f2")
pro_label.pack(side="right", padx=(0,260))


sql = 'select * from Productservice'
fbcursor.execute(sql)
pandsdata = fbcursor.fetchall()

treeproducts=ttk.Treeview(tab8,selectmode='browse')
treeproducts.place(x=8,y=100,height=580)
vertical_bar=ttk.Scrollbar(tab8,orient="vertical")
vertical_bar.place(x=1083,y=101,height=580)

treeproducts["columns"]=("1","2","3","4","5","6","7","8","9")
treeproducts["show"]='headings'
treeproducts.column("1",width=0,anchor='c', stretch=False)
treeproducts.column("2",width=160,anchor='c')
treeproducts.column("3",width=190,anchor='c')
treeproducts.column("4",width=120,anchor='c')
treeproducts.column("5",width=120,anchor='c')
treeproducts.column("6",width=120,anchor='c')
treeproducts.column("7",width=130,anchor='c')
treeproducts.column("8",width=120,anchor='c')
treeproducts.column("9",width=112,anchor='c')
treeproducts.heading("1",text="")
treeproducts.heading("2",text="Product/Service ID")
treeproducts.heading("3",text="Category")
treeproducts.heading("4",text="Status")
treeproducts.heading("5",text="Name")
treeproducts.heading("6",text="Price")
treeproducts.heading("7",text="Stock")
treeproducts.heading("8",text="Location/warehouse")
treeproducts.heading("9",text="Image")
treeproducts.bind('<Double-Button-1>' , psfile_image)
countp = 0
treeproducts.tag_configure('green', foreground='green')
treeproducts.tag_configure('red', foreground='red')
treeproducts.tag_configure('blue', foreground='blue')
for i in pandsdata:
  if i[6] == '1':
    acti = 'Active'
  else:
    acti = 'Inactive' 
  if i[13] > i[14]:
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
    countp += 1
  elif (i[12] =="0") == (i[13] <= i[14]):
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
    countp += 1
  elif i[12] == '1':
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
    countp += 1
  else:
    pass



treeproducts.place(height=580, width=1070, x=10, y=101)

######side_Listbox##############

treeps=ttk.Treeview(tab8,selectmode='browse')
treeps.place(height=580,width=254,
                      x=1099,y=101
                      )
treeps["columns"]=("1")
treeps["show"]='headings'
treeps.column("1",width=254,anchor='c')
treeps.heading("1",text="View filter by category")

def items_selected(event):
  selected_indices = listbox.curselection()
  selected_filter = ",".join([listbox.get(i) for i in selected_indices])

  sql = 'select * from Productservice'
  fbcursor.execute(sql)
  pandsdata = fbcursor.fetchall()
  psql = "select * from Productservice where serviceornot=%s"
  val = ('0', )
  fbcursor.execute(psql, val)
  pdata = fbcursor.fetchall()

  ssql = "select * from Productservice where serviceornot=%s"
  val = ('1', )
  fbcursor.execute(ssql, val)
  sdata = fbcursor.fetchall()

  # pssql = "select * from Productservice where category=%s"
  # psval = (selected_filter, )
  # fbcursor.execute(pssql, psval)
  # pssdata = fbcursor.fetchall()
  if selected_filter == "View all records":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in pandsdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  elif selected_filter == "View all products":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in pdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  elif selected_filter == "View all services":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in sdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  # elif selected_filter == True:
  #     for record in treeproducts.get_children():
  #       treeproducts.delete(record)
  #     countp = 0
  #     for i in pssdata:
  #       treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i    [2], i[4], i[7], i[13], i[15]))
  #       countp += 1
# def cat_pro(event):
#   selected_indices = listbox.curselection()
#   selected_filt = ",".join([listbox.get(i) for i in selected_indices])
#   pssql = "select * from Productservice where category=%s"
#   psval = (selected_filt, )
#   fbcursor.execute(pssql, psval)
#   pssdata = fbcursor.fetchall()
#   print(pssdata)
#   for record in treeproducts.get_children():
#     treeproducts.delete(record)
#   countp = 0
#   for i in pssdata:
#     treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i[2], i[4], i[7], i[13], i[15]))
#     countp += 1


# sql = "SELECT DISTINCT category FROM Productservice WHERE NOT category = %s AND NOT category = %s"
# val = ('service','product',)
# fbcursor.execute(sql,val,)
# lic = fbcursor.fetchall()
# print(lic)
listbox = Listbox(tab8,height = 8,  
                      width = 29,  
                      bg = "white",
                      activestyle = 'dotbox',  
                      fg = "black",
                      highlightbackground="white")  
listbox.insert(0, "View all records")
listbox.insert(1, "View all products")
listbox.insert(2, "View all services")
# for nc in lic:
#     listbox.insert(END, nc)

listbox.place(x=1099,y=118,height=564,width=255)

listbox.bind('<<ListboxSelect>>', items_selected)
# listbox.bind('<<ListboxSelect-1>>', cat_pro)

stockok = Label(tab8,text="Green: Stock is Ok",foreground="green",background="white").place(x =1110,y =580)
stocko = Label(tab8,text="Red: Limit <= Low Stock Limit",foreground="red",background="white").place(x =1110,y =600)
stock = Label(tab8,text="Blue: Service,no Stock Control",foreground="blue",background="white").place(x =1110,y =620)





#################################  END OF PRODUCT_SERVICE  ######################################








###########################    EXPENSES MODULE   ##############################

def add_expense():
    global expamountval,expdate,vn,cn,expdescriptionentry,expstaffentry,checkvarStatus4,cus,rebi,id_sku1,rebill_amoun,exptxt,expenselabelframe,rebill,imge,other
    window = Toplevel()  
    
    window.title("Add new Expense")
    p2 = PhotoImage(file = 'images/fbicon.png')
    window.iconphoto(False, p1)
 
    window.geometry("618x449+380+167")

    innerexpFrame = Frame(window, relief=GROOVE)
    innerexpFrame.pack(side="top",fill=BOTH)

    expenselabelframe = LabelFrame(innerexpFrame,text="Expense Cost",width=580,height=400)
    expenselabelframe.pack(side="top",fill=BOTH,padx=10)


    expamountval = IntVar(expenselabelframe, value='00')
    expamount=Label(expenselabelframe,text="Expense amount:",pady=10,padx=10)
    expamount.place(x=12,y=0)
    expamountentry = Entry(expenselabelframe,width=15,textvariable=expamountval)
    expamountentry.place(x=130,y=10)

    lbl_date=Label(expenselabelframe,text=" Date :",fg='black')
    lbl_date.place(x=380,y=10)
    
    expdate=DateEntry(expenselabelframe)
    expdate.place(x=450,y=12)

    sql = "select businessname from Customer where customertype =%s or customertype =%s"
    val = ('vendor','both(client,vendor)')
    fbcursor.execute(sql,val)
    pdata = fbcursor.fetchall()

    vendor1=Label(expenselabelframe,text="Vendor:",pady=5,padx=10)
    vendor1.place(x=20,y=40)
    vn = StringVar() 
    vendor = ttk.Combobox(expenselabelframe, width = 27, textvariable = vn ) 
      
    # Adding combobox drop down list 
    vendor['values'] = pdata
      
    vendor.place(x=130,y=45) 
    # vendor.current(0)

    categoryexp1=Label(expenselabelframe,text="Category:",pady=5,padx=10)
    categoryexp1.place(x=330,y=40)
    cn = StringVar() 
    categorydrop = ttk.Combobox(expenselabelframe, width = 22, textvariable = cn ) 
      
    # Adding combobox drop down list 
    categorydrop['values'] = ('Default' ) 
      
    categorydrop.place(x=400,y=45) 
    categorydrop.current(0)

    

    expdescription=Label(expenselabelframe,text="Description:",pady=10,padx=10)
    expdescription.place(x=12,y=70)
    expdescriptionentry = Entry(expenselabelframe,width=70)
    expdescriptionentry.place(x=130,y=81)

    expstafftval = StringVar(expenselabelframe, value='Administrator')
    expstaff=Label(expenselabelframe,text="Staff member:",pady=10,padx=10)
    expstaff.place(x=12,y=108)
    expstaffentry = Entry(expenselabelframe,width=30,textvariable=expstafftval)
    expstaffentry.place(x=130,y=118)

    checkvarStatus4=BooleanVar()
   
    Button4 = Checkbutton(expenselabelframe,variable = checkvarStatus4, 
                      text="Taxable Tax1 rate", 
                      onvalue ='Yes' ,
                      offvalue = 'No',
                      height=3,
                      width = 15)


    Button4.place(x=400,y=120)

    sql = "select businessname from Customer"
    fbcursor.execute(sql,)
    cusdata = fbcursor.fetchall()
    print(cusdata)

    def toggle():
      if other.get():
        ent.place(x=45,y=180)
        button51.place(x=250, y=160)
      else:
        ent.place_forget()
        button51.place_forget()
    other = BooleanVar()
    button5 = Checkbutton(expenselabelframe, text="Assign to customer (optional)", variable=other, 
    command=toggle)
    button5.place(x=40, y=160)
    cus = StringVar()
    ent=ttk.Combobox(expenselabelframe,width=30,textvariable=cus,values=cusdata)

    ent.delete(0,'end')
    def toggle():
      if rebill.get():
        id_skulabel.place(x=375,y=160)
        id_skuentry.place(x=420,y=160)
        rebill_label.place(x=335,y=180)
        rebill_entry.place(x=420, y=180)
      else:
        id_skulabel.place_forget()
        id_skuentry.place_forget()
        rebill_label.place_forget()
        rebill_entry.place_forget()
    rebill = BooleanVar()
    rebi = StringVar()
    button51 = Checkbutton(expenselabelframe, text="Rebillable" ,variable=rebill, command=toggle,onvalue ='Yes' ,offvalue = 'NO')
    
    
    id_sku1 = IntVar()
    id_skulabel=Label(expenselabelframe,text="id_sku:")
    id_skuentry = Entry(expenselabelframe,width=15,textvariable=id_sku1)
   

    rebill_amoun = IntVar()
    rebill_label=Label(expenselabelframe,text="Rebill amount:")
    rebill_entry = Entry(expenselabelframe,width=15,textvariable=rebill_amoun)
    


    
    
    def toggle():
      if imge.get():
        browseimg.place(x=40,y=220)
        browsebutton.place(x=350,y=220,height=30,width=50)
        
      else:
        browseimg.place_forget()
        browsebutton.place_forget()
      
    imge = BooleanVar()
    Button6 = Checkbutton(expenselabelframe, text = "Attach receipt image(optional,image will be stored to the database)",command=toggle,variable=imge)
    Button6.place(x=40, y=200)
    browseimg=Label(expenselabelframe,text="(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
    browsebutton=Button(expenselabelframe,text = 'Browse',command=upload_file)


    exptext1=Label(expenselabelframe,text="Notes",pady=5,padx=10)
    exptext1.place(x=12,y=246)
    exptxt = scrolledtext.ScrolledText(expenselabelframe, undo=True,width=50,height=5)
    exptxt.place(x=22,y=280)

    expokButton = Button(window, text ="Ok",image=tick,width=70,compound = LEFT,command=insert_expenses)
    expokButton.place(x=280,y=415)

    window.mainloop()

def upload_file():
      import shutil
      global filename,img, b2
      f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
      filename = filedialog.askopenfilename(filetypes=f_types)
      print(filename, 'name')
      #import pdb; pdb.set_trace()
      shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
      image = Image.open(filename)
      resize_image = image.resize((120, 120))
      img = ImageTk.PhotoImage(resize_image)
      b2 = Label(expenselabelframe,image=img, height=120, width=120)
      b2.place(x=450, y=240)
      
      




def insert_expenses():# Storing values into db (user)
  global img , filename 
  expense_amount = expamountval.get()
  date = expdate.get_date()
  vendor = vn.get()
  catagory = cn.get()
  description = expdescriptionentry.get()
  staff_members = expstaffentry.get()
  taxable = checkvarStatus4.get()
  customer = cus.get()
  id_sku = id_sku1.get()
  notes = exptxt.get('1.0', 'end-1c')
  rebill_amount = rebill_amoun.get()
  rebillab = rebill.get()
  recipt = imge.get()
  assign_cus = other.get()

  
  

  # file=open(filename,'rb').read() # filename from upload_file()
  # file = base64.b64encode(file)
  
  # sql='INSERT INTO Expenses (expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,rebill_amount,rebillable,receipt,assign_customer) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)' #adding values into db
  # val=(expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,rebill_amount,rebillab,recipt,assign_cus)
  # fbcursor.execute(sql,val)
  # fbilldb.commit()

  
  shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
  sql='INSERT INTO Expenses (expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,rebill_amount,image,rebillable,receipt,assign_customer) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)' #adding values into db
  val=(expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,  rebill_amount,filename.split('/')[-1],rebillab,recipt,assign_cus)
  fbcursor.execute(sql,val)
  fbilldb.commit()
  

  for record in exp_tree.get_children():
    exp_tree.delete(record)
  count=0
  fbcursor.execute('SELECT * FROM Expenses;')
  for i in fbcursor:
    if True:
      if i[13] == '1':
        e = 'Yes'
      else:
        e = 'No'
      exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i[7], e , i[14], i[11],i[16],i[3]))
      pass
  count += 1
#     fbcursor.execute('SELECT * FROM Expenses;')
#   j = 0
#   for i in fbcursor:
#     exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i  [7], i[13], i[14], i[11],i[16],i[3]))
#   j += 1
  messagebox.showinfo('Registration successfull','Registration successfull')

  

########################VIEW/EDIT EXPENSE#######################################################################



def edit_expense():
    global expamountval,expdate,vn,cn,expdescriptionentry,expstafftval,checkvarStatus4,cus,rebi,id_sku1,rebill_amoun,exptxt,expenselabelframe,recimage
    try:
      itemid = exp_tree.item(exp_tree.focus())["values"][0]
      sql = "select * from Expenses where expensesid = %s"
      val = (itemid, )

      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone()

      def update_expenses():# Storing values into db (user)
        global img , filename 
        itemid = exp_tree.item(exp_tree.focus())["values"][0]
        expense_amount = expamountval.get()
        date = expdate.get_date()
        vendor = vn.get()
        catagory = cn.get()
        description = expdescriptionentry.get()
        staff_members = expstafftval.get()
        taxable = checkvarStatus4.get()
        customer = cus.get()
        id_sku = id_sku1.get()
        notes = exptxt.get('1.0', 'end-1c')
        rebill_amount = rebill_amoun.get()
        rebillabe = rebill.get()
        assign_cus = other.get()
        recepit = imge.get()


        
        itemid1 = exp_tree.item(exp_tree.focus())["values"][0]
        sq = 'select image from Expenses where expensesid = %s'
        va =(itemid1,)
        fbcursor.execute(sq,va)
        up = fbcursor.fetchone()
        print(up,recimage)
        # file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
        if up:
          sql='UPDATE Expenses set expense_amount=%s,date=%s,vendor=%s,catagory=%s,description=%s,    staff_members=%s,taxable=%s,customer=%s,id_sku=%s,notes=%s,rebill_amount=%s,rebillable=%s,assign_customer=%s,receipt=%s where expensesid=%s'
          val=(expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,
          rebill_amount,rebillabe,assign_cus,recepit,itemid)
          fbcursor.execute(sql,val)
        else:
          pass
        try:
          file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
          sql='UPDATE Expenses set expense_amount=%s,date=%s,vendor=%s,catagory=%s,description=%s,    staff_members=%s,taxable=%s,customer=%s,id_sku=%s,notes=%s,rebill_amount=%s,image=%s,rebillable=%s,assign_customer=%s,receipt=%s where expensesid=%s'
          val=(expense_amount,date,vendor,catagory,description,staff_members,taxable,customer,id_sku,notes,
          rebill_amount,filename.split('/')[-1],rebillabe,assign_cus,recepit,itemid)
          fbcursor.execute(sql,val)
        except:
          pass

        fbilldb.commit()
        for record in exp_tree.get_children():
            exp_tree.delete(record)
        count=0
        fbcursor.execute('SELECT * FROM Expenses;')
        for i in fbcursor:
            if True:
              if i[13] == '1':
                e = 'Yes'
              else:
                e = 'No'
              exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i[7], e , i[14], i[11],i[16],i[3]))
            else:
                pass
        count += 1
        messagebox.showinfo('Update Successfull','Update Successfull')
      
      
      window1 = Toplevel()  
      
      window1.title("Edit Expense")
      p2 = PhotoImage(file = 'images/fbicon.png')
      # recimage= PhotoImage(file= 'images/'+psdata[11])
     
      # image = Image.open(recimage)
      # resize_image = image.resize((120, 120))
      # imga = ImageTk.PhotoImage(resize_image)
      window1.iconphoto(False, p1)
   
      window1.geometry("618x449+380+167")
  
      innerexpFrame = Frame(window1, relief=GROOVE)
      innerexpFrame.pack(side="top",fill=BOTH)
  
      expenselabelframe = LabelFrame(innerexpFrame,text="Expense Cost",width=580,height=400)
      expenselabelframe.pack(side="top",fill=BOTH,padx=10)
  
  
      expamountval = IntVar(expenselabelframe, value='$.00')
      expamount=Label(expenselabelframe,text="Expense amount:",pady=10,padx=10)
      expamount.place(x=12,y=0)
      expamountentry = Entry(expenselabelframe,width=15,textvariable=expamountval)
      expamountentry.place(x=130,y=10)
      expamountentry.delete(0,'end')
      expamountentry.insert(0, psdata[3])
  
      lbl_date=Label(expenselabelframe,text=" Date :",fg='black')
      lbl_date.place(x=380,y=10)
      
      
      expdate=DateEntry(expenselabelframe)
      expdate.place(x=450,y=12)
      expdate.delete(0,'end')
      expdate.insert(0, psdata[4])

      sql = "select businessname from Customer where customertype =%s or customertype =%s"
      val = ('vendor','both(client,vendor)')
      fbcursor.execute(sql,val)
      pdat = fbcursor.fetchall()
      vendor1=Label(expenselabelframe,text="Vendor:",pady=5,padx=10)
      vendor1.place(x=20,y=40)
      vn = StringVar() 
      vendor = ttk.Combobox(expenselabelframe, width = 27, textvariable = vn ) 
     
        
      # Adding combobox drop down list 
      vendor['values'] = pdat
        
      vendor.place(x=130,y=45) 
      vendor.delete(0,'end')
      vendor.insert(0, psdata[5]) 
  
      categoryexp1=Label(expenselabelframe,text="Category:",pady=5,padx=10)
      categoryexp1.place(x=330,y=40)
      cn = StringVar() 
      categorydrop = ttk.Combobox(expenselabelframe, width = 22, textvariable = cn ) 
      categorydrop.delete(0,'end')
      categorydrop.insert(0, psdata[6])
   
        
      # Adding combobox drop down list 
      categorydrop['values'] = ('Default') 
        
      categorydrop.place(x=400,y=45)
    
  
      
  
      expdescription=Label(expenselabelframe,text="Description:",pady=10,padx=10)
      expdescription.place(x=12,y=70)
      expdescriptionentry = Entry(expenselabelframe,width=70)
      expdescriptionentry.place(x=130,y=81)
      expdescriptionentry.delete(0,'end')
      expdescriptionentry.insert(0, psdata[7])
  
      expstafftval = StringVar(expenselabelframe, value='Administrator')
      expstaff=Label(expenselabelframe,text="Staff member:",pady=10,padx=10)
      expstaff.place(x=12,y=108)
      expstaffentry = Entry(expenselabelframe,width=30,textvariable=expstafftval)
      expstaffentry.place(x=130,y=118)
      expstaffentry.delete(0,'end')
      expstaffentry.insert(0, psdata[8])


      
      

  
      checkvarStatus4=BooleanVar()
     
      Button4 = Checkbutton(expenselabelframe,variable = checkvarStatus4, 
                        text="Taxable Tax1 rate", 
                        onvalue ='1',
                        offvalue = '0',
                        height=3,
                        width = 15)
  
      Button4.place(x=400,y=120)
      # Button4.bind("<Button-1>", getBool)
      
      ps = psdata[9]
      print(ps)
      if ps == '1':
       Button4.select()
      else:
        Button4.deselect()
          
      
 
  
      sql = "select businessname from Customer"
      fbcursor.execute(sql,)
      cusdta = fbcursor.fetchall()
      
  
      def toggle():
        if other.get():
          ent.place(x=45,y=180)
          button51.place(x=250, y=160)
        else:
          ent.place_forget()
          button51.place_forget()
      other = BooleanVar()
      button5 = Checkbutton(expenselabelframe, text="Assign to customer (optional)", variable=other, 
      command=toggle)
      button5.place(x=40, y=160)
      cus = StringVar()
      ent=ttk.Combobox(expenselabelframe,width=30,textvariable=cus)
      ent['values'] = cusdta
      ent.delete(0,'end')
      ent.insert(0, psdata[10])

      




  
      # def va():
      #   id_skulabel.place(x=375,y=160)
      #   id_skuentry.place(x=420,y=160)
      #   rebill_label.place(x=335,y=180)
      #   rebill_entry.place(x=420, y=180)

      def toggle():
        if rebill.get():
          id_skulabel.place(x=375,y=160)
          id_skuentry.place(x=420,y=160)
          rebill_label.place(x=335,y=180)
          rebill_entry.place(x=420, y=180)
        else:
          id_skulabel.place_forget()
          id_skuentry.place_forget()
          rebill_label.place_forget()
          rebill_entry.place_forget()
      rebill = BooleanVar()
      rebi = IntVar
      button51 = Checkbutton(expenselabelframe, text="Rebillable" ,variable=rebill, command=toggle)
      
      cns = psdata[17]
      if cns == '1':
        button5.select()
        ent.place(x=45,y=180)
        button51.place(x=250, y=160)
      else:
        button5.deselect()
        
      
      
      id_sku1 = IntVar(expenselabelframe, value='-Expense-')
      id_skulabel=Label(expenselabelframe,text="id_sku:")
      id_skuentry = Entry(expenselabelframe,width=15,textvariable=id_sku1)
      id_skuentry.delete(0,'end')
      id_skuentry.insert(0, psdata[15])
  
      rebill_amoun = IntVar(expenselabelframe, value='$.00')
      rebill_label=Label(expenselabelframe,text="Rebill amount:")
      rebill_entry = Entry(expenselabelframe,width=15,textvariable=rebill_amoun)
      rebill_entry.delete(0,'end')
      rebill_entry.delete(0,'end')
      rebill_entry.insert(0, psdata[16])

      reb = psdata[13]
      print(ps)
      if reb == '1':
        button51.select()
        id_skulabel.place(x=375,y=160)
        id_skuentry.place(x=420,y=160)
        rebill_label.place(x=335,y=180)
        rebill_entry.place(x=420, y=180)
      else:
        button51.deselect()
  
  
      
      
      def toggle():
        if imge.get():
          browseimg.place(x=40,y=220)
          browsebutton.place(x=350,y=220,height=30,width=50)
          b2.place(x=450, y=240)

        else:
          browseimg.place_forget()
          browsebutton.place_forget()
          b2.place_forget()
        
      imge = BooleanVar()
      Button6 = Checkbutton(expenselabelframe, text = "Attach receipt image(optional,image will be stored   to the database)",command=toggle,variable=imge)
      Button6.place(x=40, y=200)
      browseimg=Label(expenselabelframe,text="(recommended image type:JPG,size 480x320 pixels) ",  bg='#f5f3f2')
      browsebutton=Button(expenselabelframe,text = 'Browse', command=upload_file1)
     
      try:
        image = Image.open("images/"+psdata[11])
        resize_image = image.resize((120, 120))
        recimage = ImageTk.PhotoImage(resize_image)
        b2 = Button(expenselabelframe,image=recimage, height=120, width=120,)
        b2.photo = recimage
        print(image)
      except:
        pass
      
      
      rec = psdata[18]
      print(rec)
      if rec == '1':
        Button6.select()
        browseimg.place(x=40,y=220)
        browsebutton.place(x=350,y=220,height=30,width=50)
        b2.place(x=450, y=240)
      else:
        Button6.deselect()

  
      exptext1=Label(expenselabelframe,text="Notes",pady=5,padx=10)
      exptext1.place(x=12,y=246)
      exptxt = scrolledtext.ScrolledText(expenselabelframe, undo=True,width=50,height=5)
      exptxt.place(x=22,y=280)
      exptxt.delete('1.0','end')
      exptxt.insert('1.0', psdata[12])

      expokButton = Button(window1, text ="Ok",image=tick,width=70,compound = LEFT,command=update_expenses)
      expokButton.place(x=280,y=415)
    except:
        try:
            window1.destroy()
        except: 
            pass
        messagebox.showerror('F-Billing Revolution', 'Select a record to edit.')
    window1.mainloop()

def upload_file1():
   global filename,img, b1
   f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
   filename = filedialog.askopenfilename(filetypes=f_types)
   print(filename, 'name')
   #import pdb; pdb.set_trace()
   shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
   image = Image.open(filename)
   resize_image = image.resize((120, 120))
   img = ImageTk.PhotoImage(resize_image)
   b1 = Label(expenselabelframe,image=img, height=120, width=120)
   b1.place(x=450, y=240)
      
def file_image(event):
  itemid = exp_tree.item(exp_tree.focus())["values"][0]
  sql = "select * from Expenses where expensesid = %s"
  val = (itemid, )
  fbcursor.execute(sql, val)
  psda = fbcursor.fetchone() 
  if psda[11] is None:
    pass
  else:
    edit_window = Toplevel()
    edit_window.title("Edit the value or cancel")
    edit_window.geometry("700x500")
    
      
      
    image = Image.open("images/"+psda[11])
    resize_image = image.resize((700, 500))
    eximage = ImageTk.PhotoImage(resize_image)
    b2 = Button(edit_window,image=eximage)
    b2.photo = eximage
    b2.pack()
  

######################## DELETE EXPENSE #######################################################################


def delete_expense():
    
    delmess = messagebox.askyesno("Delete Expense", "Are you sure to delete this Expense?")
    if delmess == True:
      itemid = exp_tree.item(exp_tree.focus())["values"][0]
      print(itemid)
      sql = 'DELETE FROM Expenses WHERE expensesid=%s'
      val = (itemid,)
      fbcursor.execute(sql, val)
      fbilldb.commit()
      #selrow = exp_tree.selection()[0]
      exp_tree.delete(exp_tree.selection()[0])
    else:
      pass
  

######################## SEARCH EXPENSE ######################################################################
def close_expenses():
  top.destroy()

def search_exp():
  query = searchvar.get()
  selections = []
  for child in exp_tree.get_children():
      if query in exp_tree.item(child)['values']:
          print(exp_tree.item(child)['values'])
          selections.append(child)
  exp_tree.selection_set(selections)
  
  
  

def search_expense():
    global top,searchvar
    top = Toplevel()  
    
    top.title("Find Text")
    
    
    top.geometry("520x200+390+250")
    findwhat1=Label(top,text="Find What:",pady=5,padx=10)
    findwhat1.place(x=5,y=20)
    searchvar = StringVar() 
    findwhat = ttk.Combobox(top, width = 50, textvariable = searchvar ) 
      
    # Adding combobox drop down list 
    
    findwhat.place(x=80,y=25) 
    

    findButton = Button(top, text ="Find next",width=10, command=search_exp)
    findButton.place(x=420,y=20)

    findin1=Label(top,text="Find in:",pady=5,padx=10)
    findin1.place(x=5,y=47)
    n = StringVar() 
    findIN = ttk.Combobox(top, width = 37, textvariable = n ) 
      
    # Adding combobox drop down list 
    findIN['values'] = ('Client',  
                              ' Date', 
                              ' Category', 
                              ' Vendor', 
                              ' Staff Member', 
                              ' Description', 
                              ' Rebillable',
                              'Invoiced',
                              'Image',
                              'Rebill Amount',
                              'Amount',
                        
                              ' <<All>>') 
      
    findIN.place(x=80,y=54) 
    findIN.current(0)

    closeButton = Button(top, text ="Close",width=10,command=close_expenses)
    closeButton.place(x=420,y=50)

    match1=Label(top,text="Match:",pady=5,padx=10)
    match1.place(x=5,y=74)
    n = StringVar() 
    match = ttk.Combobox(top, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    match['values'] = ('From Any part of the field',' Whole Field',  
                              ' From the beginning of the field')
      
    match.place(x=80,y=83) 
    match.current(0)

    search1=Label(top,text="Search:",pady=5,padx=10)
    search1.place(x=5,y=102)
    n = StringVar() 
    search = ttk.Combobox(top, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    search['values'] = ('All', 'up', 
                              ' Down')
      
    search.place(x=80,y=112) 
    search.current(0)


    checkvarStatus4=IntVar()
   
    Button4 = Checkbutton(top,variable = checkvarStatus4, 
                      text="Match Case", 
                      onvalue =1,
                      offvalue = 0,
                      height=3,
                      width = 15)

    Button4.place(x=60,y=141)

    checkvarStatus5=IntVar()
   
    Button5 = Checkbutton(top,variable = checkvarStatus5, 
                      text="Match Format", 
                      onvalue =0 ,
                      offvalue = 1,
                      height=3,
                      width = 15)

    Button5.place(x=270,y=141)







  

######################## FRONT PAGE OF EXPENSE MODULE #######################################################################

    
expframe = Frame(tab6,relief=GROOVE,bg="#f8f8f2")
expframe.pack(side="top", fill=BOTH)

expmidFrame=Frame(expframe, height=60,bg="#f5f3f2")
expmidFrame.pack(side="top", fill=X)

e = Canvas(expmidFrame, width=1, height=65, bg="#f8f8f2", bd=0)
e.pack(side="left", padx=(5, 2))
e = Canvas(expmidFrame, width=1, height=65, bg="#f8f8f2", bd=0)
e.pack(side="left", padx=(0, 5))

expenseIcon = ImageTk.PhotoImage(Image.open("images/plus.png"))
expenseLabel = Button(expmidFrame,compound="top", text="Create new\nExpense",relief=RAISED, command=add_expense, image=expenseIcon,bg="#f5f3f2", fg="black", height=55, bd=1, width=55,)
expenseLabel.pack(side="left", pady=3, ipadx=4)

expeditIcon = ImageTk.PhotoImage(Image.open("images/edit.png"))
expeditLabel = Button(expmidFrame,compound="top", text="Edit/View\nExpense",relief=RAISED,  image=expeditIcon,command=edit_expense,bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
expeditLabel.pack(side="left")

expdeleteIcon = ImageTk.PhotoImage(Image.open("images/delete.png"))
expdeleteLabel = Button(expmidFrame,compound="top", text="Delete\nSelected", relief=RAISED,  command=delete_expense,image=expdeleteIcon,bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
expdeleteLabel.pack(side="left")

e = Canvas(expmidFrame, width=1, height=65, bg="#b3b3b3", bd=0)
e.pack(side="left", padx=5)

expsearchIcon = ImageTk.PhotoImage(Image.open("images/search-icon.png"))
expsearchLabel = Button(expmidFrame,compound="top", text="Search in\nExpenses",relief=RAISED, command=search_expense, image=expsearchIcon,bg="#f8f8f2", fg="black", height=55, bd=1, width=55, )
expsearchLabel.pack(side="left")


lbframe = LabelFrame(expmidFrame, height=60, width=200)
lbframe.pack(side="left", padx=10, pady=0)

lbl_expdt=Label(lbframe,text="Expense date from:",fg='black')
lbl_expdt.grid(row=0, column=0, pady=5, padx=(5, 0))

lbl_expdtt=Label(lbframe,text="Expense date to:" , fg='black')
lbl_expdtt.grid(row=1, column=0, pady=5, padx=(5, 0))

def daterange_expenses(): # Start and stop dates for range
  var1=expdt1.get_date()
  var2=expdtt2.get_date()
  print(var1,var2)
  for record in exp_tree.get_children():
    exp_tree.delete(record)
  
  sqldate='SELECT * FROM Expenses WHERE date BETWEEN %s AND %s'
  valuz=(var1,var2,)
  fbcursor.execute(sqldate,valuz)
  filterdate=fbcursor.fetchall()
  print(filterdate)
  count=0
  for i in filterdate:
    if True:
      if i[13] == '1':
        e = 'Yes'
      else:
        e = 'No'
      exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i [8], i[7], e , i[14], i[11],i[16],i[3]))
    else:
        pass
  count += 1

  
expdt1=DateEntry(lbframe)
expdt1.grid(row=0, column=1)
   
expdtt2=DateEntry(lbframe)
expdtt2.grid(row=1, column=1)
   
checkvar1 = IntVar()
chkbtn1 = Checkbutton(lbframe, text = "Apply filter", variable = checkvar1, onvalue = 1, offvalue =0,   height = 2, width = 8,command=daterange_expenses)
chkbtn1.grid(row=0, column=2, rowspan=2, padx=(5,5))

mainFrameex=Frame(tab8, relief=GROOVE, bg="#f8f8f2")
mainFrameex.pack(side="top", fill=BOTH)
e = Canvas(mainFrameex, width=1, height=55, bg="#b3b3b3", bd=0)
e.pack(side="left", padx=5)

def refresh_expenses():
  for record in exp_tree.get_children():
    exp_tree.delete(record)
  count=0
  fbcursor.execute('SELECT * FROM Expenses;')
  for i in fbcursor:
    if True:
      if i[13] == '1':
        e = 'Yes'
      else:
        e = 'No'
      exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i[7], e , i[14], i[11],i[16],i[3]))
    else:
        pass
  count += 1
  

exprefreshIcon = ImageTk.PhotoImage(Image.open("images/refresh.png"))
exprefreshLabel = Button(expmidFrame,compound="top", text="Refresh\nExpense List",relief=RAISED,  image=exprefreshIcon,bg="#f8f8f2", fg="black", height=55, bd=1, width=63,command=refresh_expenses)
exprefreshLabel.pack(side="left")



invoi1label = Label(expframe, text="Expenses (All)", font=("arial", 18), bg="#f8f8f2")
invoi1label.pack(side="left", padx=(20,0))

def fil(event):
  filt = drop123.get()
  for record in exp_tree.get_children():
    exp_tree.delete(record)
  
  
  
  
  sql = "select * from Expenses where catagory = %s"
  val = (filt,)
  fbcursor.execute(sql, val)
  records = fbcursor.fetchall()
  
  
  count=0
  for i in records:
      if True:
        if i[13] == '1':
           e = 'Yes'
        else:
          e = 'No'
        exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i[7], e , i[14], i[11],i[16],i[3]))
      else:
        pass
  count += 1

sql = "SELECT DISTINCT catagory FROM Expenses"
fbcursor.execute(sql,)
rec = fbcursor.fetchall()
drop123 = ttk.Combobox(expframe,)
drop123['values'] = rec
drop123.pack(side="right", padx=(0,10))
drop123.bind("<<ComboboxSelected>>", fil)


 
invoi1label = Label(expframe, text="Category filter", font=("arial", 15), bg="#f8f8f2")
invoi1label.pack(side="right", padx=(0,10))

# sql= 'SELECT rebillable FROM Expenses '
# fbcursor.execute(sql,)
# c = fbcursor.fetchall()
# print (c[2])
# print(c == 1)
# for e in c:
#   m = e == c
#   m = ("Yes")
#   e != c
#   print("no")
#   else:
#     pass



#table 
s = ttk.Style()
s.configure('Treeview.Heading', background='white', State='DISABLE')


exp_tree=ttk.Treeview(tab6,selectmode='browse')
exp_tree.place(x=0,y=105,height=580)

expverticalbar=ttk.Scrollbar(tab6,orient="vertical",command=exp_tree.yview,)
expverticalbar.place(x=1345,y=102,height=570,)
expverticalbar.place(x=1345,y=102,height=570)
exp_tree["columns"]=("1","2","3","4","5","6","7","8","9","10","11","12")
exp_tree["show"]='headings'
exp_tree.column("1",width=5,anchor='c')
exp_tree.column("2",width=130,anchor='c')
exp_tree.column("3",width=110,anchor='c')
exp_tree.column("4",width=120,anchor='c')
exp_tree.column("5",width=120,anchor='c')
exp_tree.column("6",width=120,anchor='c')
exp_tree.column("7",width=220,anchor='c')
exp_tree.column("8",width=120,anchor='c')
exp_tree.column("9",width=100,anchor='c')
exp_tree.column("10",width=100,anchor='c')
exp_tree.column("11",width=100,anchor='c')
exp_tree.column("12",width=100,anchor='c')
exp_tree.heading("2",text="Client")
exp_tree.heading("3",text="Date")
exp_tree.heading("4",text="Category")
exp_tree.heading("5",text="Vendor")
exp_tree.heading("6",text="Staff member")
exp_tree.heading("7",text="Description")
exp_tree.heading("8",text="Rebillable")
exp_tree.heading("9",text="Invoiced")
exp_tree.heading("10",text="Image")
exp_tree.heading("11",text="Rebill Amount")
exp_tree.heading("12",text="Amount")
exp_tree.bind('<Double-Button-1>' , file_image)





fbcursor.execute('SELECT * FROM Expenses;')

j = 0

for i in fbcursor:
  if i[13] == '1':
    e = 'Yes'
  else:
    e = 'No'
  exp_tree.insert(parent='', index='end', iid=i, text='hello', values=(i[0], i[10], i[4], i[6], i[5], i[8], i[7], e , i[14], i[11],i[16],i[3]))
  j += 1

############################ END OF Expense Module-############################
root.mainloop()
